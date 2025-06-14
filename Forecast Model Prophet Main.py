import pandas as pd
import os
import numpy as np
from collections import Counter
from prophet import Prophet
from prophet.diagnostics import cross_validation, performance_metrics
import matplotlib.pyplot as plt
import warnings
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from statsmodels.tsa.statespace.sarimax import SARIMAX
import joblib
from sklearn.metrics import (
    mean_absolute_error,
    median_absolute_error,
    mean_absolute_error,
    mean_squared_error,
    mean_absolute_percentage_error
)
from math import sqrt
from pykalman import KalmanFilter
from statsmodels.tsa.stattools import adfuller
from sklearn.preprocessing import MinMaxScaler
from sklearn.linear_model import LinearRegression, LogisticRegression
from scipy.optimize import minimize
from statsmodels.tsa.seasonal import STL
from sklearn.cluster import KMeans
from scipy.stats import linregress
from math import ceil
import optuna

warnings.filterwarnings("ignore")

os.environ["TMPDIR"] = "C:/Temp"

forecast_params_used = {}

changepoint_counter = Counter()
seasonality_counter = Counter()
holiday_counter = Counter()
out_of_range_counter = Counter()
out_of_range_stats = {}

# Global dictionaries to store feedback for each model type
prophet_feedback = {}
sarima_feedback = {}
xgboost_feedback = {}
forecast_errors = {}

# Global dictionaries to store the accumulated performance history
sarima_param_history = {}  # key = (asin, (p,d,q,P,D,Q,m)), value = dict with 'score', 'count', etc.
prophet_param_history = {} # key = (asin, (changepoint, seasonality, holiday)), value = dict with 'score', 'count', etc.

##############################
# Seasonal Trend
##############################

def compute_rmse(actual, predicted):
    actual_filled = actual.fillna(0)
    pred_filled   = predicted.fillna(0)
    return sqrt(mean_squared_error(actual_filled, pred_filled))

def determine_seasonal_trend(sales_series):
    """
    Determines the seasonal trend based on weekly sales data.
    
    Logic:
      - Compute the week-to-week percentage change.
      - If there exist at least two consecutive weeks where the percentage
        increase is greater than 40%, return "High Volume Season".
      - If there exist at least two consecutive weeks where the percentage
        decrease is less than -40%, return "Low Volume Season".
      - Otherwise, return "Stable".
    
    Parameters:
      sales_series (list or pandas Series): Weekly sales numbers in chronological order.
    
    Returns:
      str: "High Volume Season", "Low Volume Season", or "Stable".
    """
    sales_array = np.array(sales_series, dtype=float)
    n = len(sales_array)
    if n < 2:
        return "Stable"
    
    pct_changes = []
    for i in range(1, n):
        prev = sales_array[i-1]
        if prev == 0:
            pct_changes.append(0.0)
        else:
            pct_changes.append((sales_array[i] - prev) / prev)
    pct_changes = np.array(pct_changes)
    
    count_increase = 0
    count_decrease = 0
    for change in pct_changes:
        if change > 0.4:
            count_increase += 1
            count_decrease = 0
        elif change < -0.4:
            count_decrease += 1
            count_increase = 0
        else:
            count_increase = 0
            count_decrease = 0
        if count_increase >= 2:
            return "High Volume Season"
        if count_decrease >= 2:
            return "Low Volume Season"
    return "Stable"

def compute_inventory_coverage(current_inventory, weekly_forecast, lead_time=4):
    """
    Calculate inventory coverage in weeks.
    
    Formula: coverage_weeks = current_inventory / average_demand_over_lead_time
    """
    if len(weekly_forecast) == 0:
        return float('inf')
    forecast_for_lead = weekly_forecast[:lead_time] if len(weekly_forecast) >= lead_time else weekly_forecast
    avg_demand = np.mean(forecast_for_lead)
    if avg_demand == 0:
        return float('inf')
    coverage = current_inventory / avg_demand
    return round(coverage, 1)

def compute_stockout_risk(coverage_weeks, lead_time=4):
    """
    Return 'High' if coverage is less than lead_time, else 'Low'.
    """
    if coverage_weeks < lead_time:
        return "High"
    else:
        return "Low"

def compute_reorder_urgency(stockout_risk):
    """
    Return 'Urgent' if stockout risk is high, otherwise 'Normal'.
    """
    return "Urgent" if stockout_risk == "High" else "Normal"

def compute_sales_trend(historical_sales_52):
    """
    Perform linear regression on the provided weekly sales data.
    
    Returns:
      "Increasing (▲)" if the slope > 0,
      "Decreasing (▼)" if the slope < 0,
      "Stable (-)" if the slope is near zero.
    """
    if len(historical_sales_52) < 2:
        return "Stable (-)"
    x = np.arange(len(historical_sales_52))
    slope, intercept, r_val, p_val, std_err = linregress(x, historical_sales_52)
    if slope > 0:
        return "Increasing (▲)"
    elif slope < 0:
        return "Decreasing (▼)"
    else:
        return "Stable (-)"

def compute_seasonality_index(week_number, average_annual_sales):
    """
    Compute a seasonality index as a placeholder.
    
    In practice, compare actual weekly sales to average_annual_sales.
    Here, we return a dummy ratio between 0.8 and 1.2.
    """
    ratio = np.random.uniform(0.8, 1.2)
    return round(ratio, 2)

def compute_sales_volume_rank(asins_total_sales, asin):
    """
    Rank the given ASIN by total historical sales.
    
    asins_total_sales: dict {asin: total_sales, ...}
    """
    sorted_asins = sorted(asins_total_sales.keys(), key=lambda k: asins_total_sales[k], reverse=True)
    rank = sorted_asins.index(asin) + 1
    return rank

def determine_lifecycle_stage(slope):
    """
    Label lifecycle stage based on the sales trend slope.
    
    If slope >= 0.1: 'Growth'
    If slope <= -0.1: 'Decline'
    Else: 'Mature'
    """
    if slope >= 0.1:
        return "Growth"
    elif slope <= -0.1:
        return "Decline"
    else:
        return "Mature"

def forecast_po_order(current_inventory, forecasted_weekly_sales, target_coverage=5):
    avg_forecast_sales = np.mean(forecasted_weekly_sales)
    target_inventory = target_coverage * avg_forecast_sales
    if current_inventory < target_inventory:
        return target_inventory - current_inventory
    else:
        return 0

def generate_restock_suggestions(consolidated_data, runrate_inventory, coverage_threshold=1.0, output_file="restock_suggestions.xlsx", target_weeks=4):
    wb = Workbook()
    ws = wb.active
    ws.title = "Restock Suggestions"

    # Define header row (note the additional column for trigger week PO qty)
    header = [
        "ASIN", "Product Title", "Trigger Week", "Trigger Date",
        "Expected PO Qty (Trigger Week)", "Expected PO Qty (4 Weeks)",
        "Suggested PO Qty (4 Weeks)", "Suggested PO Qty (8 Weeks)",
        "Suggested PO Qty (16 Weeks)"
    ]
    ws.append(header)

    # Process each ASIN’s forecast DataFrame.
    for asin, df in consolidated_data.items():
        if df.empty:
            continue

        # Get the product title (assumed to be in the "Product Title" column of the first row)
        product_title = df["Product Title"].iloc[0] if "Product Title" in df.columns else ""

        # Ensure the DataFrame has the required columns.
        required_cols = {"Week", "Week_Start_Date", "MyForecast"}
        if not required_cols.issubset(set(df.columns)):
            print(f"ASIN {asin} is missing required columns. Skipping restock suggestion for this ASIN.")
            continue

        # Start with the current inventory for this ASIN.
        current_inv = runrate_inventory.get(asin, 0)
        simulated_inv = current_inv

        trigger_week = None
        trigger_date = None
        trigger_index = None
        simulated_inv_at_trigger = None

        # Loop through each week (row) in the forecast DataFrame.
        for idx, row in df.iterrows():
            weekly_forecast = row["MyForecast"]
            if weekly_forecast <= 0:
                continue  # Skip weeks with zero forecast to avoid division by zero.
            coverage = simulated_inv / weekly_forecast
            if coverage < coverage_threshold:
                trigger_week = row["Week"]
                trigger_date = row["Week_Start_Date"]
                trigger_index = idx
                simulated_inv_at_trigger = simulated_inv  # Record the inventory remaining at trigger.
                break
            # Subtract the forecasted sales from inventory.
            simulated_inv -= weekly_forecast

        if trigger_week is None:
            # No trigger was found; set output values to "N/A"
            trigger_week = "N/A"
            trigger_date = "N/A"
            expected_trigger_po = "N/A"
            expected_po_qty = "N/A"
            sug_po_4 = "N/A"
            sug_po_8 = "N/A"
            sug_po_16 = "N/A"
        else:
            # Calculate the expected PO for the trigger week.
            trigger_week_forecast = df["MyForecast"].iloc[trigger_index]
            expected_trigger_po = int(round(max(trigger_week_forecast - simulated_inv_at_trigger, 0)))

            # For the next target_weeks (default 4 weeks) demand:
            trigger_idx_pos = df.index.get_loc(trigger_index)
            forecast_4wk = df["MyForecast"].iloc[trigger_idx_pos : trigger_idx_pos + target_weeks]
            total_demand_4wk = forecast_4wk.sum()
            # Expected PO qty for 4 weeks: additional units required to cover 4-week demand.
            expected_po_qty = int(round(max(total_demand_4wk - simulated_inv, 0)))

            # Suggested PO quantities for extended planning horizons.
            forecast_8wk = df["MyForecast"].iloc[trigger_idx_pos : trigger_idx_pos + 8]
            forecast_16wk = df["MyForecast"].iloc[trigger_idx_pos : trigger_idx_pos + 16]
            sug_po_4 = total_demand_4wk
            sug_po_8 = forecast_8wk.sum() if len(forecast_8wk) > 0 else total_demand_4wk
            sug_po_16 = forecast_16wk.sum() if len(forecast_16wk) > 0 else total_demand_4wk

            expected_trigger_po = int(round(expected_trigger_po))
            expected_po_qty = int(round(expected_po_qty))

        # Append a new row with the computed information.
        ws.append([
            asin,
            product_title,
            trigger_week,
            trigger_date,
            expected_trigger_po,
            expected_po_qty,
            sug_po_4,
            sug_po_8,
            sug_po_16
        ])

    # Save the workbook.
    try:
        wb.save(output_file)
        print(f"Restock suggestions saved to '{output_file}'")
    except Exception as e:
        print(f"Failed to save restock suggestions to '{output_file}': {e}")

##############################
# Reward & Penalty Functions
##############################

def compute_reward(mae, rmse):
    """
    Example reward function that returns higher reward for lower RMSE/MAE.
    Feel free to adjust or add more sophisticated weighting.
    """
    alpha = 0.7  # Weight for MAE
    beta = 0.3   # Weight for RMSE
    
    # Compute a "badness" measure
    badness = alpha * mae + beta * rmse  # Lower is better
    
    # Convert badness to reward (higher is better)
    reward = 1.0 / (1.0 + badness)  # Ensures reward is between 0 and 1
    return reward

def update_param_history(history_dict, asin, param_tuple, rmse, mae):
    """
    Updates the global parameter-history dictionary with a new RMSE/MAE.
    - We compute a new reward
    - We accumulate it into 'score'
    - We track how many times this param set was tried ('count')
    """
    reward = compute_reward(mae, rmse)
    key = (asin, param_tuple)
    if key not in history_dict:
        history_dict[key] = {
            'score': reward,
            'count': 1,
            'avg_rmse': rmse,
            'avg_mae': mae
        }
    else:
        # Weighted average for the RMSE, MAE, and score
        prev = history_dict[key]
        new_count = prev['count'] + 1
        prev['avg_rmse'] = (prev['avg_rmse'] * prev['count'] + rmse) / new_count
        prev['avg_mae'] = (prev['avg_mae'] * prev['count'] + mae) / new_count
        prev['score'] = (prev['score'] * prev['count'] + reward) / new_count
        prev['count'] = new_count

def save_param_histories():
    joblib.dump(sarima_param_history, "sarima_param_history.pkl")
    joblib.dump(prophet_param_history, "prophet_param_history.pkl")

def load_param_histories():
    global sarima_param_history, prophet_param_history
    try:
        sarima_param_history = joblib.load("sarima_param_history.pkl")
    except:
        sarima_param_history = {}
    try:
        prophet_param_history = joblib.load("prophet_param_history.pkl")
    except:
        prophet_param_history = {}

    # PATCH any missing keys (avg_mae, avg_rmse) in old records
    for d in [sarima_param_history, prophet_param_history]:
        for (asin, param_tuple), entry in d.items():
            if 'avg_mae' not in entry or entry['avg_mae'] is None:
                entry['avg_mae'] = 0.0
            if 'avg_rmse' not in entry or entry['avg_rmse'] is None:
                entry['avg_rmse'] = 0.0

def load_runrate_inventory(runrate_file):
    """
    Loads the Runrate.xlsx file, which has columns:
    'ASIN', 'UPC', 'Model Name', 'OTW', 'INV', 'Total Inventory', 'POS (Shipped)', 'Ordered Units'
    
    Returns a dict: { asin: total_inventory }
    """
    runrate_df = pd.read_excel(runrate_file)
    # Ensure the column names match exactly
    runrate_df.columns = runrate_df.columns.str.strip()
    
    inventory_dict = {}
    for _, row in runrate_df.iterrows():
        asin = str(row['ASIN']).strip()
        total_inv = row.get('Total Inventory', 0)
        # Convert or coerce to int if needed
        try:
            total_inv = int(total_inv)
        except:
            total_inv = 0
        inventory_dict[asin] = total_inv
    return inventory_dict


##############################
# Configuration
##############################
# ↓↓↓ For more aggressive fallback to SARIMA, reduce threshold to 20 or another smaller number
FALLBACK_THRESHOLD = 20  # e.g., using 20% or smaller data sets for SARIMA
SARIMA_WEIGHT = 0.4      # Weight for SARIMA forecast if combining with Amazon or other model

##############################
# Counter
##############################

def record_forecast_params(product_id, changepoint_prior_scale, seasonality_prior_scale, holidays_prior_scale):
    """Record the parameters used for a specific product."""
    forecast_params_used[product_id] = {
        'changepoint_prior_scale': changepoint_prior_scale,
        'seasonality_prior_scale': seasonality_prior_scale,
        'holidays_prior_scale': holidays_prior_scale,
    }
    # Update global counters
    changepoint_counter[changepoint_prior_scale] += 1
    seasonality_counter[seasonality_prior_scale] += 1
    holiday_counter[holidays_prior_scale] += 1

def summarize_usage():
    """Summarize parameter usage for all products and the most common settings."""
    print("\n=== Forecast Parameter Usage by Product ===")
    for product_id, params in forecast_params_used.items():
        print(f"Product {product_id}:")
        print(f"  Changepoint Prior Scale: {params['changepoint_prior_scale']}")
        print(f"  Seasonality Prior Scale: {params['seasonality_prior_scale']}")
        print(f"  Holidays Prior Scale: {params['holidays_prior_scale']}")
    
    print("\n=== Most Common Parameters Across All Products ===")
    print("Most Common Changepoint Settings:")
    for value, count in changepoint_counter.most_common():
        print(f"  {value}: {count} times")
    
    print("\nMost Common Seasonality Settings:")
    for value, count in seasonality_counter.most_common():
        print(f"  {value}: {count} times")
    
    print("\nMost Common Holiday Settings:")
    for value, count in holiday_counter.most_common():
        print(f"  {value}: {count} times")
    print("================================")


##############################
# Added Functions for Caching
##############################

def save_model(model, model_type, asin, ts_data):
    model_cache_folder = "model_cache"
    os.makedirs(model_cache_folder, exist_ok=True)
    model_path = os.path.join(model_cache_folder, f"{asin}_{model_type}.pkl")
    
    # Safely set the attribute if the model object supports it
    try:
        model.last_train_date = ts_data['ds'].max()
    except AttributeError:
        print(f"Warning: Model object does not support 'last_train_date' attribute.")
    
    joblib.dump(model, model_path)
    print(f"Model saved to {model_path}")

def load_model(model_type, asin):
    model_cache_folder = "model_cache"
    model_path = os.path.join(model_cache_folder, f"{asin}_{model_type}.pkl")
    exog_path = os.path.join(model_cache_folder, f"{asin}_{model_type}_exog.pkl")

    if os.path.exists(model_path):
        model = joblib.load(model_path)
        if os.path.exists(exog_path):
            exog = joblib.load(exog_path)
            return model, exog
        return model, None
    return None, None

def is_model_up_to_date(model_type, asin, merged_df, model_cache_folder="model_cache"):
    """
    Checks if we have a cached model for the given (model_type, asin),
    and if that model is trained through the latest historical date in merged_df.
    If up to date, return True, else False.
    """
    model_path = os.path.join(model_cache_folder, f"{asin}_{model_type}.pkl")
    if not os.path.exists(model_path):
        return False  
    # Load the cached model
    cached_model = joblib.load(model_path)
    if not hasattr(cached_model, 'last_train_date'):
        return False  # no attribute to compare => treat as not up to date

    # Compare the model's last_train_date to the max ds in merged_df
    last_date_in_data = merged_df['ds'].max()
    model_train_date = cached_model.last_train_date

    # If model was trained through or beyond the last date in your data, it's up to date
    return model_train_date >= last_date_in_data


##############################
# Kalman filter-based Missing Data Handling
##############################

def kalman_smooth(series):
    """Use a Kalman Filter to fill missing values in the series."""
    values = series.values.astype(float)
    initial_state_mean = np.nanmean(values)
    if np.isnan(initial_state_mean):
        initial_state_mean = 0.0

    kf = KalmanFilter(
        initial_state_mean=initial_state_mean,
        n_dim_obs=1,
        n_dim_state=1,
        initial_state_covariance=1,
        transition_matrices=[1],
        observation_matrices=[1],
        transition_covariance=0.1,
        observation_covariance=1.0
    )

    masked_values = np.ma.array(values, mask=np.isnan(values))
    kf = kf.em(masked_values, n_iter=5)
    smoothed_state_means, _ = kf.smooth(masked_values)

    smoothed_state_means_1d = smoothed_state_means[:, 0]
    filled_values = values.copy()
    nan_idx = np.isnan(values)
    filled_values[nan_idx] = smoothed_state_means_1d[nan_idx]

    filled_values = np.round(filled_values).astype(int)
    return pd.Series(filled_values, index=series.index)

def handle_missing_data(data):
    """Handle missing data in 'y' using Kalman smoothing if enough data points exist."""
    if len(data['y']) < 2:
        print("Not enough data points for Kalman smoothing. Using fallback method for missing data.")
        data['y'] = data['y'].fillna(method='bfill').fillna(method='ffill').astype(int)
    else:
        data['y'] = kalman_smooth(data['y'])
    return data

def handle_outliers(data):
    """Clip outliers in 'y' based on IQR."""
    Q1 = data['y'].quantile(0.25)
    Q3 = data['y'].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    clipped = data['y'].clip(lower=lower_bound, upper=upper_bound).round().astype(int)
    data['y'] = clipped
    return data

def min_max_scale_data(df, feature_cols):
    """
    Apply MinMax scaling to specified feature columns.
    Returns the scaled DataFrame and a dictionary of scalers.
    """
    scalers = {}
    for col in feature_cols:
        scaler = MinMaxScaler()
        df[col] = scaler.fit_transform(df[[col]])
        scalers[col] = scaler
    return df, scalers

def preprocess_data(data):
    """
    Preprocess data by handling missing values, outliers, and (optionally) scaling if needed.
    """
    data = handle_missing_data(data)
    data = handle_outliers(data)
    # Example: scale 'y' using MinMax (optional, comment out if not desired)
    # data, scalers = min_max_scale_data(data, ['y'])
    return data


##############################
# Differencing and Stationarity
##############################

def differencing(timeseries, m):
    """Create differenced series for potential stationarity checks."""
    info = []
    tcopy = timeseries.copy()
    original = tcopy.copy()
    for i in range(3):
        original.name = f"d{i}_D0_m0"
        info.append(original)
        original = original.diff().dropna()
    for period in m:
        for j in range(3):
            d_series = info[j].diff(periods=period).dropna()
            d_series.name = f"d{j}_D1_m{period}"
            info.append(d_series)
    for period in m:
        for j in range(3):
            d_series = info[j+3].diff(periods=period).dropna()
            d_series.name = f"d{j}_D2_m{period}"
            info.append(d_series)
    df_info = pd.concat(info, axis=1)
    return df_info

def adf_summary(diff_series):
    """Perform ADF tests on multiple differenced series and return a summary DataFrame."""
    summary = []
    for col in diff_series.columns:
        series = diff_series[col].dropna()
        if len(series) < 3:
            summary.append([np.nan]*7)
            continue
        a, b, c, d, e, f = adfuller(series)
        g, h, i = e.values()
        results = [a, b, c, d, g, h, i]
        summary.append(results)
    columns = ["Test Statistic", "p-value", "#Lags Used", "No. of Obs. Used",
               "Critical Value (1%)", "Critical Value (5%)", "Critical Value (10%)"]
    index = diff_series.columns
    summary = pd.DataFrame(summary, index=index, columns=columns)
    return summary

def create_holiday_regressors(ts_data, holidays):
    """Create exogenous holiday regressors for SARIMA."""
    holiday_names = holidays['holiday'].unique()
    exog = pd.DataFrame(index=ts_data.index)
    exog['ds'] = ts_data['ds']
    for h in holiday_names:
        holiday_dates = holidays[holidays['holiday'] == h]['ds']
        exog[h] = exog['ds'].isin(holiday_dates).astype(int)
    exog.drop(columns=['ds'], inplace=True)
    return exog


##############################
# Custom SARIMA fitting
##############################

def calculate_forecast_metrics(actual, predicted):
    rmse = np.sqrt(mean_squared_error(actual, predicted))
    mae = mean_absolute_error(actual, predicted)
    return rmse, mae

def fit_sarima_model(data, holidays, seasonal_period=52, asin=None):
    """
    Automatically fit a SARIMA model by iterating over a set of p, d, q, P, D, Q, m values.
    We store a reward for each param set in sarima_param_history, so next time 
    we can skip or penalize poor-performing sets for the same ASIN.
    
    Returns:
        (best_model, best_params)
    """
    exog = create_holiday_regressors(data, holidays)
    sample_size = len(data)
    if sample_size < 52:
        if sample_size >= 12:
            seasonal_period = 12
        elif sample_size >= 4:
            seasonal_period = None
        else:
            seasonal_period = 1

    p_values = [0, 1, 2]
    d_values = [0, 1]
    q_values = [0, 1, 2]
    P_values = [0, 1]
    D_values = [0, 1]
    Q_values = [0, 1]

    if seasonal_period is None:
        m_values = [1]
        seasonal = False
    else:
        m_values = [seasonal_period]
        seasonal = True

    best_rmse = float('inf')
    best_model = None
    best_metrics = None

    skip_threshold_score = 0.001  # skip param sets with < 0.001 historical score

    for p in p_values:
        for d in d_values:
            for q in q_values:
                if seasonal:
                    for P in P_values:
                        for D in D_values:
                            for Q in Q_values:
                                for m in m_values:
                                    param_tuple = (p, d, q, P, D, Q, m)
                                    if asin is not None:
                                        key = (asin, param_tuple)
                                        if key in sarima_param_history:
                                            if sarima_param_history[key]['score'] < skip_threshold_score:
                                                continue
                                    try:
                                        model = SARIMAX(
                                            data['y'],
                                            order=(p, d, q),
                                            seasonal_order=(P, D, Q, m),
                                            exog=exog if not exog.empty else None,
                                            enforce_stationarity=False,
                                            enforce_invertibility=False
                                        )
                                        model_fit = model.fit(disp=False)
                                        forecast = model_fit.fittedvalues
                                        actual = data['y']
                                        rmse, mae = calculate_forecast_metrics(actual, forecast)

                                        if asin is not None:
                                            update_param_history(
                                                sarima_param_history, 
                                                asin, 
                                                param_tuple, 
                                                rmse, 
                                                mae
                                            )
                                        if rmse < best_rmse:
                                            best_rmse = rmse
                                            best_model = model_fit
                                            best_metrics = {
                                                'RMSE': rmse,
                                                'MAE': mae,
                                                'p': p, 'd': d, 'q': q,
                                                'P': P, 'D': D, 'Q': Q, 'm': m
                                            }
                                    except:
                                        continue
                else:
                    try:
                        param_tuple = (p, d, q, 0, 0, 0, 1)
                        if asin is not None:
                            key = (asin, param_tuple)
                            if key in sarima_param_history:
                                if sarima_param_history[key]['score'] < skip_threshold_score:
                                    print(f"Skipping SARIMA params {param_tuple} for ASIN {asin} due to low historical score.")
                                    continue
                        model = SARIMAX(
                            data['y'],
                            order=(p, d, q),
                            seasonal_order=(0, 0, 0, 1),
                            exog=exog if not exog.empty else None,
                            enforce_stationarity=False,
                            enforce_invertibility=False
                        )
                        model_fit = model.fit(disp=False)
                        forecast = model_fit.fittedvalues
                        actual = data['y']
                        rmse, mae = calculate_forecast_metrics(actual, forecast)

                        if asin is not None:
                            update_param_history(
                                sarima_param_history, 
                                asin, 
                                param_tuple, 
                                rmse, 
                                mae
                            )
                        if rmse < best_rmse:
                            best_rmse = rmse
                            best_model = model_fit
                            best_metrics = {
                                'RMSE': rmse,
                                'MAE': mae,
                                'p': p, 'd': d, 'q': q,
                                'P': 0, 'D': 0, 'Q': 0, 'm': 1
                            }
                    except:
                        continue

    if best_model is None:
        print("No suitable SARIMA model found.")
        return None, None
    else:
        if best_metrics is not None:
            print(
                f"Best SARIMA Model for {asin}: "
                f"(p,d,q)=({best_metrics['p']},{best_metrics['d']},{best_metrics['q']}), "
                f"(P,D,Q,m)=({best_metrics['P']},{best_metrics['D']},{best_metrics['Q']},{best_metrics['m']}), "
                f"RMSE={best_metrics['RMSE']:.2f}, MAE={best_metrics['MAE']:.2f}"
            )
        return best_model, (best_metrics['p'], 
                            best_metrics['d'], 
                            best_metrics['q'], 
                            best_metrics['P'], 
                            best_metrics['D'], 
                            best_metrics['Q'], 
                            best_metrics['m'])



def sarima_forecast(model_fit, steps, last_date, exog=None):
    """Generate SARIMA forecasts for the specified number of steps ahead."""
    k_exog = model_fit.model.k_exog
    if exog is not None and k_exog > 0:
        # Align exog columns if needed
        exog = exog.iloc[:, :k_exog]

    forecast_values = model_fit.forecast(steps=steps, exog=exog)
    forecast_values = forecast_values.round().astype(int)
    future_dates = pd.date_range(start=last_date + pd.Timedelta(weeks=1), periods=steps, freq='W-SUN')
    # Instead of "SARIMA Forecast", unify to "MyForecast":
    return pd.DataFrame({'ds': future_dates, 'MyForecast': forecast_values})

def generate_future_exog(holidays, steps, last_date):
    """Generate exogenous holiday regressors for future forecasts."""
    future_dates = pd.date_range(start=last_date + pd.Timedelta(weeks=1), periods=steps, freq='W-SUN')
    exog_future = pd.DataFrame(index=future_dates)
    holiday_names = holidays['holiday'].unique()
    for h in holiday_names:
        exog_future[h] = exog_future.index.isin(holidays[holidays['holiday'] == h]['ds'])
    exog_future = exog_future.astype(int)
    return exog_future

def choose_forecast_model(ts_data, threshold=FALLBACK_THRESHOLD, holidays=None):
    """
    Basic decision for model selection:
    - If dataset size <= threshold, use SARIMA.
    - Else, default to Prophet.
    
    We now return either:
      (fitted_sarima_model, "SARIMA")
    or
      ("Prophet", "Prophet")
    so that the calling code can see if it's 'SARIMA' or 'Prophet'.
    """
    asin = ts_data['asin'].iloc[0] if 'asin' in ts_data.columns else "UnknownASIN"

    if len(ts_data) <= threshold:
        print(f"Dataset size ({len(ts_data)}) is <= threshold ({threshold}). Using SARIMA.")
        fitted_model, best_params = fit_sarima_model(ts_data, holidays, seasonal_period=52, asin=asin)
        if fitted_model is not None:
            save_model(fitted_model, "SARIMA", asin, ts_data)
            return fitted_model, "SARIMA"
        else:
            print(f"SARIMA model fitting failed for {asin}, skipping.")
            # Return a placeholder so that the calling code doesn't crash
            return (None, "SARIMA_Failed")
    else:
        print(f"Dataset size ({len(ts_data)}) is > threshold ({threshold}). Using Prophet.")
        return ("Prophet", "Prophet")

def generate_date_from_week(row):
    week_str = row['week']
    year     = row['year']
    week_num = int(week_str[1:])
    # uses %U (Sunday‐based) and then does week_num-1
    return pd.to_datetime(f'{year}-W{week_num - 1}-0', format='%Y-W%U-%w')

def generate_date_from_inventory_week(row):
    """
    Convert (Week, Year) in your inventory_data to a proper datetime 'ds'.
    For example, if row['Week'] = 'W1' and row['Year'] = 2025,
    we interpret it as the Sunday of ISO week 0:
    """
    week_str = str(row['Week']).strip()
    year_str = str(row['Year']).strip()

    if week_str.upper().startswith('W'):
        week_number = int(week_str[1:])
    else:
        week_number = int(week_str)

    date_str = f"{year_str}-W{week_number - 1}-0"  # Sunday of that ISO week
    try:
        dt = pd.to_datetime(date_str, format='%Y-W%U-%w')
    except ValueError:
        dt = pd.NaT
    return dt

def load_inventory_data(inventory_file):
    """
    Loads the inventory_data.xlsx file, which should have columns like:
       'ASIN', 'Model', 'OTW', 'INV', 'Total', 'POS', 'Week', 'Year'
    'Total' is the weekly inventory for each ASIN. We rename it to 'Starting_Inventory'.

    Then we parse Week+Year -> ds using generate_date_from_inventory_week.
    """
    inv_df = pd.read_excel(inventory_file)
    inv_df.columns = inv_df.columns.str.strip()

    # Convert (Week, Year) -> ds
    inv_df['ds'] = inv_df.apply(generate_date_from_inventory_week, axis=1)

    # Rename 'Total' to 'Starting_Inventory'
    inv_df.rename(columns={'Total': 'Starting_Inventory'}, inplace=True)

    # Drop rows that can't parse date
    inv_df.dropna(subset=['ds'], inplace=True)

    return inv_df

def clean_weekly_sales_data(data):
    """Placeholder for additional cleaning steps if needed."""
    return data

def load_weekly_sales_data(file_path):
    """Load weekly sales data from Excel, ensuring required columns are present."""
    data = pd.read_excel(file_path)
    data.columns = data.columns.str.strip().str.lower()

    required_columns = ['product title', 'week', 'year', 'units_sold', 'asin']
    missing_required = [col for col in required_columns if col not in data.columns]
    if missing_required:
        raise ValueError(f"Missing required columns: {missing_required}")

    data['ds'] = data.apply(generate_date_from_week, axis=1)
    data = data.rename(columns={'units_sold': 'y'})
    data['y'] = data['y'].astype(int)
    data = clean_weekly_sales_data(data)
    return data

def load_asins_to_forecast(file_path):
    """Load a list of ASINs from either a text or Excel file."""
    if file_path.endswith('.txt'):
        with open(file_path, 'r') as file:
            asins = [line.strip() for line in file if line.strip()]
    elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
        df = pd.read_excel(file_path)
        asins = df.iloc[:, 0].dropna().astype(str).tolist()
    else:
        raise ValueError("Unsupported file format for ASINs to Forecast file.")
    return asins

def load_amazon_forecasts_from_folder(folder_path, asin):
    """
    Load Amazon forecast data from multiple Excel files in a folder.
    Each file should correspond to a specific forecast type (Mean, P70, etc.).
    """
    forecast_data = {}
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            # Extract forecast type from filename, e.g., 'Mean Forecast.xlsx' -> 'Mean'
            forecast_type = os.path.splitext(file_name)[0].replace('_', ' ').title()
            if forecast_type.endswith(' Forecast'):
                forecast_type = forecast_type.replace(' Forecast', '')
            file_path = os.path.join(folder_path, file_name)
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
            except Exception as e:
                print(f"Error reading {file_path}: {e}")
                continue

            df.columns = df.columns.str.strip().str.upper()
            if 'ASIN' not in df.columns:
                print(f"Error: Column 'ASIN' not found in {file_name}")
                continue

            asin_row = df[df['ASIN'].str.upper() == asin.upper()]
            if asin_row.empty:
                print(f"Warning: ASIN {asin} not found in {file_name}")
                continue

            week_columns = [col for col in df.columns if 'W' in col.upper()]
            if not week_columns:
                print(f"Warning: No week columns found in {file_name}")
                continue

            forecast_values = (
                asin_row.iloc[0][week_columns]
                .astype(str)
                .str.replace(',', '')
                .astype(float)
                .round()
                .astype(int)
                .values
            )
            forecast_data[forecast_type] = forecast_values
    return forecast_data

def add_lag_features(ts_data, lag_weeks=1):
    """Add lag features to the time series DataFrame."""
    ts_data = ts_data.copy()
    ts_data = ts_data.sort_values('ds')
    ts_data[f'lag_{lag_weeks}_week'] = ts_data['y'].shift(lag_weeks)
    ts_data[f'lag_{lag_weeks}_week'].fillna(0, inplace=True)
    return ts_data

def prepare_time_series_with_lags(data, asin, lag_weeks=1):
    """
    Prepare a single ASIN's data for modeling:
    - rename columns
    - sort by date
    - preprocess (missing/outliers)
    - add lag features
    """
    ts_data = data[data['asin'] == asin].copy()
    ts_data = ts_data.sort_values('ds')
    ts_data = preprocess_data(ts_data)
    ts_data = add_lag_features(ts_data, lag_weeks)
    return ts_data

def get_shifted_holidays():
    """
    Example holiday DataFrame with 'ds' as date and 'holiday' as holiday name.
    Adjust to match your actual holiday data.
    """
    holidays_list = [
        ('Prime Day', '2023-06-27'),
        ('Prime Day', '2023-06-28'),
        ('Black Friday', '2023-11-10'),
        ('Cyber Monday', '2023-11-13'),
        ('Christmas', '2023-12-11'),
        ('Prime Day', '2024-07-02'),
        ('Prime Day', '2024-07-03'),
        ('Black Friday', '2024-11-15'),
        ('Cyber Monday', '2024-11-18'),
        ('Christmas', '2024-12-09'),
    ]
    holidays = pd.DataFrame(holidays_list, columns=['holiday', 'ds'])
    holidays['ds'] = pd.to_datetime(holidays['ds'])
    return holidays

def add_holiday_lead_time_features(merged_df, holidays):
    """
    For each row in merged_df (each ds date),
    compute how many days until the next holiday.
    Also mark if we are within 2 weeks prior to a major holiday.
    """
    merged = merged_df.copy()
    holidays_sorted = holidays.sort_values('ds').reset_index(drop=True)

    # For performance, you can do a merge_asof or a nearest join,
    # but here's a simple approach for demonstration:
    merged['days_to_next_holiday'] = None
    for i, row in merged.iterrows():
        current_date = row['ds']
        # find the next holiday date >= current_date
        upcoming = holidays_sorted[holidays_sorted['ds'] >= current_date]
        if upcoming.empty:
            merged.at[i, 'days_to_next_holiday'] = 9999  # no future holiday
        else:
            next_hol_date = upcoming['ds'].iloc[0]
            days_diff = (next_hol_date - current_date).days
            merged.at[i, 'days_to_next_holiday'] = days_diff

    # Optionally create a boolean feature: "is_within_14_days_of_holiday"
    merged['is_within_14_days_of_holiday'] = merged['days_to_next_holiday'] <= 14

    return merged


##############################
# Ensemble Approach (SARIMA, Prophet)
##############################

def ensemble_forecast(sarima_preds, prophet_preds, amazon_mean_preds, weights):
    """
    Weighted ensemble of forecasts from SARIMA, Prophet, and Amazon Mean.
    weights: [w_sarima, w_prophet, w_mean]
    """
    return (
        weights[0] * sarima_preds +
        weights[1] * prophet_preds +
        weights[2] * amazon_mean_preds
    )

def evaluate_forecast(actual, forecast):
    """Compute RMSE for forecast evaluation."""
    rmse = sqrt(mean_squared_error(actual, forecast))
    return rmse

def walk_forward_validation_ensemble(ts_data, n_test, model_sarima, model_prophet, weights):
    """
    Example walk-forward validation for an ensemble approach using SARIMA and Prophet.
    weights: [w_sarima, w_prophet]
    """
    predictions = []
    actuals = []
    train = ts_data.iloc[:-n_test].copy()
    test = ts_data.iloc[-n_test:].copy()

    for i in range(n_test):
        test_point = ts_data.iloc[len(train) + i]

        # SARIMA forecast for next step
        if model_sarima is not None:
            exog_test = None  # Modify if your SARIMA requires exogenous variables
            sarima_value = model_sarima.forecast(steps=1, exog=exog_test)
            sarima_value = int(round(sarima_value.iloc[0]))
        else:
            sarima_value = 0

        # Prophet forecast for next step
        if model_prophet is not None:
            future = pd.DataFrame({'ds': [test_point['ds']]})
            prophet_predict = model_prophet.predict(future)['yhat'].values[0]
            prophet_value = int(round(max(prophet_predict, 0)))
        else:
            prophet_value = 0

        # Ensemble forecast: weighted combination of SARIMA and Prophet
        ensemble_pred = weights[0] * sarima_value + weights[1] * prophet_value
        ensemble_pred = int(round(ensemble_pred))
        predictions.append(ensemble_pred)
        actuals.append(test_point['y'])

    rmse_val = evaluate_forecast(actuals, predictions)
    return rmse_val, predictions, actuals

def print_ensemble_summary(rmse, predictions, actuals):
    """Display a user-friendly summary of the ensemble results."""
    print("\n=== Ensemble Forecast Summary ===")
    print(f"Number of Observations: {len(actuals)}")
    print(f"RMSE: {rmse:.2f}")
    mae_val = mean_absolute_error(actuals, predictions)
    mape_val = mean_absolute_percentage_error(actuals, predictions) * 100
    print(f"MAE: {mae_val:.2f}")
    print(f"MAPE: {mape_val:.2f}%")
    print("=================================\n")

def create_decision_matrix(sarima_rmse, prophet_rmse, ensemble_rmse):
    """
    Create a simple decision matrix showing which model performed best.
    """
    data = {
        'Model': ['SARIMA', 'Prophet', 'Ensemble'],
        'RMSE': [sarima_rmse, prophet_rmse, ensemble_rmse]
    }
    df = pd.DataFrame(data)
    df.sort_values(by='RMSE', inplace=True)
    return df


##############################
# Prophet Customization
##############################

def forecast_with_custom_params(ts_data, forecast_data,
                                changepoint_prior_scale,
                                seasonality_prior_scale,
                                holidays_prior_scale,
                                horizon=16,
                                weights=None):
    """
    Train and forecast with Prophet using user-defined hyperparams and custom weighting for regressors.
    We'll unify the final forecast column to "MyForecast".
    """
    # Default weights if not provided
    weights = weights or {
        'Amazon Mean': 0.7,
        'Amazon P70': 0.2,
        'Amazon P80': 0.1
    }

    future_dates = pd.date_range(start=ts_data['ds'].max() + pd.Timedelta(days=7), periods=horizon, freq='W-SUN')
    future = pd.DataFrame({'ds': future_dates})
    combined_df = pd.concat([ts_data, future], ignore_index=True)

    # Attach forecast_data (Amazon forecasts) as regressors
    for forecast_type, values in forecast_data.items():
        values_to_use = values[:horizon] if len(values) > horizon else values
        extended_values = np.concatenate(
            [
                np.full(len(ts_data), np.nan),
                values_to_use,
                np.full(max(horizon - len(values_to_use), 0),
                        values_to_use[-1] if len(values_to_use) > 0 else 0)
            ]
        )
        extended_values = extended_values[:len(combined_df)]
        combined_df[f'Amazon_{forecast_type} Forecast'] = extended_values

    regressor_cols = [col for col in combined_df.columns if col.startswith('Amazon_')]
    combined_df[regressor_cols] = combined_df[regressor_cols].fillna(0)

    holidays = get_shifted_holidays()
    combined_df['prime_day'] = combined_df['ds'].apply(
        lambda x: 0.2 if x in holidays[holidays['holiday'] == 'Prime Day']['ds'].values else 0
    )

    n = len(ts_data)
    split = int(n * 0.8)
    train_df = combined_df.iloc[:split].dropna(subset=['y']).copy()
    test_df = combined_df.iloc[split:].copy()

    model = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=True,
        seasonality_mode='multiplicative',
        holidays=holidays,
        changepoint_prior_scale=changepoint_prior_scale,
        seasonality_prior_scale=seasonality_prior_scale,
        holidays_prior_scale=holidays_prior_scale,
        growth='linear'
    )

    # Add regressors dynamically
    for regressor in regressor_cols + ['prime_day']:
        model.add_regressor(regressor, mode='multiplicative')

    model.fit(train_df[['ds', 'y'] + regressor_cols + ['prime_day']])

    # Evaluate on test split
    test_forecast = model.predict(test_df.drop(columns='y').copy())
    test_actual = combined_df.iloc[split:].dropna(subset=['y'])
    if not test_actual.empty:
        test_eval = test_forecast[test_forecast['ds'].isin(test_actual['ds'])]
        # We'll do a rough integer rounding for this test evaluation:
        test_eval['MyForecast'] = test_eval['yhat'].round().astype(int)
        mae = mean_absolute_error(test_actual['y'], test_eval['MyForecast'])
        rmse_val = sqrt(mean_squared_error(test_actual['y'], test_eval['MyForecast']))
        print(f"Prophet Test MAE: {mae:.4f}, RMSE: {rmse_val:.4f}")

    future_df = combined_df[combined_df['y'].isna()].drop(columns='y').copy()
    forecast = model.predict(future_df)

    # Apply custom weighting mechanism to adjust forecasts
    for amazon_col, weight in weights.items():
        if f'Amazon_{amazon_col} Forecast' in future_df.columns:
            forecast['yhat'] += weight * future_df[f'Amazon_{amazon_col} Forecast']

    forecast['MyForecast'] = forecast['yhat'].round().astype(int).clip(lower=0)

    return forecast[['ds', 'MyForecast', 'yhat', 'yhat_upper']], model


##############################
# Prophet Parameter Optimization
##############################

PARAM_COUNTER = 0
POOR_PARAM_FOUND = False
EARLY_STOP_THRESHOLD = 10_000

def optimize_prophet_params_with_optuna(ts_data, forecast_data, horizon=16, asin=None, n_trials=30):
    """
    Use Optuna for Bayesian optimization of Prophet hyperparams.
    Returns the best set of hyperparameters, plus an optional dictionary of RMSE values.

    Parameters:
      ts_data (DataFrame): historical time series for this ASIN with columns ['ds','y', ...].
      forecast_data (dict): a dictionary of Amazon forecast arrays, e.g. {"Mean": [...], "P70": [...]}.
      horizon (int): forecast horizon in weeks for cross-validation or hold-out tests.
      asin (str): optional label for logging/feedback.
      n_trials (int): how many optuna trials to run (the bigger, the more thorough).

    Returns:
      best_params (dict): best set of hyperparams from the Optuna study
      best_rmse_values (dict): optional dictionary of model vs. Amazon RMSE metrics
    """

    # Ensure your time series is sorted and has valid ds,y
    ts_data = ts_data.copy()
    ts_data = ts_data.sort_values('ds')

    # Optionally define your cross-validation settings
    cv_initial = '365 days'  # or 52 weeks if weekly data
    cv_period  = '180 days'  
    cv_horizon = '90 days'   # you can adjust these if you have enough data

    # 1) Prepare an objective function for Optuna:
    def objective(trial):
        # Suggest hyperparams from a search space
        changepoint_prior_scale = trial.suggest_loguniform('changepoint_prior_scale', 0.001, 5.0)
        seasonality_prior_scale = trial.suggest_loguniform('seasonality_prior_scale', 0.001, 5.0)
        holidays_prior_scale    = trial.suggest_float('holidays_prior_scale', 0.0, 20.0)
        seasonality_mode        = trial.suggest_categorical('seasonality_mode', ['additive', 'multiplicative'])
        
        # Initialize Prophet with these hyperparams
        model = Prophet(
            yearly_seasonality=True,
            weekly_seasonality=True,
            daily_seasonality=False,
            changepoint_prior_scale=changepoint_prior_scale,
            seasonality_prior_scale=seasonality_prior_scale,
            holidays_prior_scale=holidays_prior_scale,
            seasonality_mode=seasonality_mode,
            growth='linear'
        )

        # If you have regressors in your code, add them here EXACTLY as in your final model:
        # for r in regressor_cols:
        #     model.add_regressor(r, mode='multiplicative')

        # Fit model
        try:
            model.fit(ts_data[['ds','y']])
        except:
            # If fit fails for some reason, return a large value => "bad" trial
            return 1e10

        # Prophet cross-validation can fail if there's not enough data => handle it
        try:
            df_cv = cross_validation(
                model,
                initial=cv_initial,
                period=cv_period,
                horizon=cv_horizon
            )
            df_perf = performance_metrics(df_cv, rolling_window=1)
            rmse_val = df_perf['rmse'].iloc[0]
        except ValueError:
            # Not enough data or something else => penalize
            rmse_val = 1e10

        return rmse_val

    # 2) Run the Optuna study
    study = optuna.create_study(direction='minimize')
    study.optimize(objective, n_trials=n_trials)

    # 3) Extract best hyperparams
    best_params = study.best_trial.params
    print(f"[Optuna] Best trial RMSE={study.best_trial.value:.4f} for ASIN={asin}, params={best_params}")

    # 4) (Optional) Re-train final model on entire ts_data with best_params
    #    and compute a custom "RMSE vs. Amazon forecasts" if you want
    final_model = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=True,
        daily_seasonality=False,
        changepoint_prior_scale=best_params['changepoint_prior_scale'],
        seasonality_prior_scale=best_params['seasonality_prior_scale'],
        holidays_prior_scale=best_params['holidays_prior_scale'],
        seasonality_mode=best_params['seasonality_mode'],
        growth='linear'
    )
    final_model.fit(ts_data[['ds','y']])

    # Quick forecast to compute your custom RMSE with Amazon data:
    future = pd.date_range(
        start=ts_data['ds'].max() + pd.Timedelta(weeks=1),
        periods=horizon,
        freq='W-SUN'
    )
    future_df = pd.DataFrame({'ds': future})
    forecast = final_model.predict(future_df)
    # Force integer forecast or round if you like:
    forecast['MyForecast'] = forecast['yhat'].clip(lower=0).round()

    # Compare "MyForecast" vs. Amazon streams:
    best_rmse_values = {}
    for forecast_type, values in forecast_data.items():
        # pad or slice the array to horizon
        horizon_vals = values[:horizon] if len(values) >= horizon else values
        if len(horizon_vals) < horizon and len(horizon_vals) > 0:
            horizon_vals = np.pad(horizon_vals, (0,horizon-len(horizon_vals)), 'constant', constant_values=horizon_vals[-1])
        elif len(horizon_vals) == 0:
            horizon_vals = np.zeros(horizon, dtype=int)

        # compute RMSE
        rmse = sqrt(mean_squared_error(horizon_vals, forecast['MyForecast']))
        best_rmse_values[forecast_type] = rmse

    print(f"[Optuna] Average RMSE vs. Amazon streams for ASIN {asin}: {np.mean(list(best_rmse_values.values())):.4f}")

    return best_params, best_rmse_values

def forecast_with_optuna_params(ts_data, forecast_data, horizon=16, asin=None, n_trials=30):
    """
    Revised function that calls our Optuna-based hyperparam search,
    then uses those best hyperparams to produce a final forecast.
    
    Returns:
      final_forecast (DataFrame) with columns ['ds','MyForecast', 'yhat','yhat_upper']
      prophet_model (fitted Prophet model)
    """
    # 1) Get best hyperparams
    best_params, _ = optimize_prophet_params_with_optuna(ts_data, forecast_data, horizon, asin, n_trials=n_trials)

    # 2) Build a Prophet model with those best params
    model = Prophet(
        yearly_seasonality      = True,
        weekly_seasonality      = True,
        daily_seasonality       = False,
        changepoint_prior_scale = best_params['changepoint_prior_scale'],
        seasonality_prior_scale = best_params['seasonality_prior_scale'],
        holidays_prior_scale    = best_params['holidays_prior_scale'],
        seasonality_mode        = best_params['seasonality_mode'],
        growth                  = 'linear'
    )

    # 2b) If your code uses regressors, add them here:
    # for r in regressor_cols:
    #     model.add_regressor(r, mode='multiplicative')

    model.fit(ts_data[['ds','y']])

    # 3) Forecast for the next `horizon` weeks
    future_dates = pd.date_range(
        start=ts_data['ds'].max() + pd.Timedelta(weeks=1),
        periods=horizon,
        freq='W-SUN'
    )
    future_df = pd.DataFrame({'ds': future_dates})
    forecast = model.predict(future_df)
    forecast['MyForecast'] = forecast['yhat'].round().clip(lower=0).astype(int)

    return forecast[['ds','MyForecast','yhat','yhat_upper']], model

def calculate_rmse(forecast, forecast_data, horizon):
    """Calculate RMSE comparing our 'MyForecast' vs. Amazon forecast streams."""
    rmse_values = {}
    for forecast_type, values in forecast_data.items():
        values = values[:horizon]
        if len(values) < horizon:
            if len(values) > 0:
                values = np.pad(values, (0, horizon - len(values)), 'constant', constant_values=values[-1])
            else:
                values = np.zeros(horizon, dtype=int)
        rmse = np.sqrt(((forecast['MyForecast'] - values) ** 2).mean())
        rmse_values[forecast_type] = rmse
    return rmse_values

##############################
# Season Detect
##############################

def stl_decomposition(ts_data, period=52):
    """Perform STL decomposition on historical data"""
    decomposition = STL(ts_data['Weekly_Sales'], period=period, robust=True).fit()
    return decomposition

def calculate_seasonal_strength(decomposition):
    """Calculate normalized seasonal strength (0-1)"""
    resid_std = decomposition.resid.std()
    trend_std = decomposition.trend.std()
    return max(0, min(1 - (resid_std / (trend_std + 1e-9)), 1))

def detect_seasonal_periods(ts_data, n_clusters=3):
    """Identify seasonal periods using clustering"""
    decomposition = stl_decomposition(ts_data)
    seasonal = decomposition.seasonal
    
    # Normalize seasonal component
    seasonal_norm = (seasonal - seasonal.min()) / (seasonal.max() - seasonal.min())
    
    # Cluster into seasonal periods
    kmeans = KMeans(n_clusters=n_clusters, random_state=0)
    clusters = kmeans.fit_predict(seasonal_norm.values.reshape(-1, 1))
    
    # Create cluster labels
    cluster_means = pd.Series(seasonal_norm).groupby(clusters).mean()
    sorted_clusters = cluster_means.sort_values().index
    cluster_labels = {sorted_clusters[0]: 'low', 
                     sorted_clusters[1]: 'medium',
                     sorted_clusters[2]: 'high'}
    
    ts_data['seasonal_cluster'] = [cluster_labels[c] for c in clusters]
    return ts_data

def calculate_seasonal_factors(ts_data):
    """Calculate weekly adjustment factors"""
    ts_data = ts_data.copy()
    ts_data['week_of_year'] = ts_data['ds'].dt.isocalendar().week
    
    factors = ts_data.groupby('week_of_year').agg({
        'y': 'mean',
        'seasonal_cluster': lambda x: x.value_counts().index[0]
    }).rename(columns={'y': 'historical_mean'})
    
    # Calculate adjustment factors relative to overall mean
    overall_mean = factors['historical_mean'].mean()
    factors['adjustment_factor'] = factors['historical_mean'] / overall_mean
    
    return factors

def apply_seasonal_adjustment(forecast_df, ts_data, seasonal_factors,
                             override_threshold=0.3, 
                             max_override=2.0):
    """
    Apply seasonal adjustments to forecasts with override logic
    """
    # Add seasonal components to forecast
    forecast_df = forecast_df.copy()
    forecast_df['week_of_year'] = forecast_df['ds'].dt.isocalendar().week
    
    # Merge seasonal factors
    forecast_df = forecast_df.merge(
        seasonal_factors[['adjustment_factor', 'seasonal_cluster']],
        left_on='week_of_year',
        right_index=True,
        how='left'
    )
    
    # Calculate base adjusted forecast
    forecast_df['seasonal_forecast'] = forecast_df['MyForecast'] * forecast_df['adjustment_factor']
    
    # Calculate override ranges
    forecast_df['amazon_deviation'] = (
        forecast_df['MyForecast'] - forecast_df['Amazon Mean Forecast']
    ).abs() / forecast_df['Amazon Mean Forecast']
    
    # Apply conditional override
    override_conditions = (
        (forecast_df['amazon_deviation'] > override_threshold) &
        (forecast_df['seasonal_cluster'].isin(['low', 'high']))
    )
    
    # Limit override magnitude
    forecast_df['final_forecast'] = np.where(
        override_conditions,
        np.clip(
            forecast_df['seasonal_forecast'],
            forecast_df['Amazon Mean Forecast'] * (1 - max_override),
            forecast_df['Amazon Mean Forecast'] * (1 + max_override)
        ),
        forecast_df[['MyForecast', 'Amazon Mean Forecast']].mean(axis=1)
    )
    
    return forecast_df

def validate_seasonal_adjustment(historical_data, adjusted_forecast):
    """Validate seasonal adjustment performance"""
    merged = historical_data.merge(
        adjusted_forecast,
        on='ds',
        suffixes=('_actual', '_forecast'),
        how='inner'
    )
    
    metrics = {
        'mae_before': mean_absolute_error(merged['y_actual'], merged['MyForecast']),
        'mae_after': mean_absolute_error(merged['y_actual'], merged['final_forecast']),
        'seasonal_strength': calculate_seasonal_strength(stl_decomposition(historical_data))
    }
    
    improvement = (metrics['mae_before'] - metrics['mae_after']) / metrics['mae_before']
    metrics['improvement_pct'] = improvement * 100
    
    return metrics

##############################
# Forecast Formatting & Adjustments
##############################

def format_output_with_forecasts(my_forecast_df, forecast_data, horizon=16):
    """
    Combine our 'MyForecast' with Amazon forecast data for comparison.
    Also compute the difference and percentage difference columns.
    """
    comparison = my_forecast_df.copy()
    for forecast_type, values in forecast_data.items():
        values = values[:horizon]
        values = np.array(values, dtype=int)
        if 'Forecast' in forecast_type:
            amazon_col_name = f'Amazon {forecast_type}'
        else:
            amazon_col_name = f'Amazon {forecast_type} Forecast'
        forecast_df = pd.DataFrame({
            'ds': my_forecast_df['ds'],
            amazon_col_name: values
        })
        comparison = comparison.merge(forecast_df, on='ds', how='left')
    comparison.fillna(0, inplace=True)

    for col in comparison.columns:
        if col.startswith('Amazon ') and col.endswith('Forecast'):
            base_name = col.replace('Amazon ', '').replace(' Forecast', '')
            diff_col = f"Diff_{base_name}"
            pct_col = f"Pct_{base_name}"
            comparison[diff_col] = (comparison['MyForecast'] - comparison[col]).round().astype(int)
            comparison[pct_col] = np.where(
                comparison[col] != 0,
                (comparison[diff_col] / comparison[col]) * 100,
                0
            )

    comparison['MyForecast'] = comparison['MyForecast'].round().astype(int)

    for col in comparison.columns:
        if col.startswith("Amazon ") or col.startswith("Diff_") or col.startswith("Pct_"):
            comparison[col] = comparison[col].astype(int, errors='ignore')
    return comparison

def adjust_forecast_weights(forecast, yhat_weight, yhat_upper_weight):
    """Adjust final MyForecast by combining yhat and yhat_upper with specified weights."""
    if 'yhat' not in forecast or 'yhat_upper' not in forecast:
        raise KeyError("'yhat' or 'yhat_upper' not found in forecast DataFrame.")
    adj_forecast = (yhat_weight * forecast['yhat'] + yhat_upper_weight * forecast['yhat_upper']).clip(lower=0)
    adj_forecast = adj_forecast.round().astype(int)
    forecast['MyForecast'] = adj_forecast
    return forecast

def find_best_forecast_weights(forecast, comparison, weights):
    """
    Brute force search over (yhat_weight, yhat_upper_weight) pairs to minimize average RMSE vs. Amazon columns.
    """
    best_rmse = float('inf')
    best_weights = None
    rmse_results = {}
    for yhat_weight, yhat_upper_weight in weights:
        adjusted_forecast = adjust_forecast_weights(forecast.copy(), yhat_weight, yhat_upper_weight)
        rmse_values = {}
        for amazon_col in comparison.columns:
            if amazon_col.startswith('Amazon '):
                rmse = np.sqrt(((comparison[amazon_col] - adjusted_forecast['MyForecast']) ** 2).mean())
                rmse_values[amazon_col] = rmse
        avg_rmse = np.mean(list(rmse_values.values()))
        rmse_results[(yhat_weight, yhat_upper_weight)] = avg_rmse
        if avg_rmse < best_rmse:
            best_rmse = avg_rmse
            best_weights = (yhat_weight, yhat_upper_weight)
    return best_weights, rmse_results

def auto_find_best_weights(forecast, comparison, step=0.05):
    """
    Auto-scan [0.5, 1.0] to find best combination of yhat/yhat_upper weights for final "MyForecast".
    """
    best_rmse = float('inf')
    best_weights = None
    candidates = np.arange(0.5, 1.0 + step, step)
    for w in candidates:
        yhat_weight = w
        yhat_upper_weight = 1 - w
        adjusted_forecast = adjust_forecast_weights(forecast.copy(), yhat_weight, yhat_upper_weight)
        rmse_values = {}
        for amazon_col in comparison.columns:
            if amazon_col.startswith('Amazon '):
                rmse = np.sqrt(((comparison[amazon_col] - adjusted_forecast['MyForecast']) ** 2).mean())
                rmse_values[amazon_col] = rmse
        avg_rmse = np.mean(list(rmse_values.values()))
        if avg_rmse < best_rmse:
            best_rmse = avg_rmse
            best_weights = (yhat_weight, yhat_upper_weight)
    return best_weights, best_rmse


##############################
# Cross Validation
##############################

def try_cross_validation_with_fallback(model, ts_data, horizons, initial='365 days', period='180 days'):
    """
    Attempt multiple horizon values for cross-validation until one works or all fail.
    """
    for horizon in horizons:
        try:
            df_cv = cross_validation(model, initial=initial, period=period, horizon=horizon)
            df_performance = performance_metrics(df_cv)
            print(f"Cross-validation successful with horizon={horizon}")
            return df_cv, df_performance
        except ValueError as e:
            if "Less data than horizon" in str(e):
                print(f"Not enough data for horizon={horizon}, trying smaller horizon.")
            else:
                print(f"Error with horizon={horizon}: {e}")
    print("The ASIN has too little data for cross-validation with any tested horizons.")
    return None, None

def validate_best_params(ts_data, best_params, initial='365 days', period='180 days'):
    """
    Validate the chosen best_params via cross-validation, if possible.
    """
    model = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=True,
        seasonality_mode='multiplicative',
        changepoint_prior_scale=best_params['changepoint_prior_scale'],
        seasonality_prior_scale=best_params['seasonality_prior_scale'],
        holidays_prior_scale=best_params['holidays_prior_scale']
    )
    model.fit(ts_data[['ds','y']])
    candidate_horizons = ['180 days', '90 days', '60 days', '30 days']
    df_cv, df_performance = try_cross_validation_with_fallback(model, ts_data, candidate_horizons, initial=initial, period=period)
    if df_cv is not None and df_performance is not None:
        print("Validation Metrics with chosen best_params:")
        print(df_performance)
        return True
    else:
        print("No valid cross-validation results due to insufficient data.")
        return False


##############################
# Summary Statistics & Visualization
##############################

def calculate_summary_statistics(ts_data, forecast_df, horizon):
    """
    Calculate basic summary stats for historical data and partial sums for 'MyForecast'.
    """
    summary_stats = {
        "min": ts_data["y"].min(),
        "max": ts_data["y"].max(),
        "mean": ts_data["y"].mean(),
        "median": ts_data["y"].median(),
        "std_dev": ts_data["y"].std(),
        "total_sales": ts_data["y"].sum(),
        "data_range": (ts_data["ds"].min(), ts_data["ds"].max())
    }

    total_forecast_16 = forecast_df['MyForecast'][:16].sum() if len(forecast_df) >= 16 else forecast_df['MyForecast'].sum()
    total_forecast_8 = forecast_df['MyForecast'][:8].sum() if len(forecast_df) >= 8 else forecast_df['MyForecast'].sum()
    total_forecast_4 = forecast_df['MyForecast'][:4].sum() if len(forecast_df) >= 4 else forecast_df['MyForecast'].sum()

    max_forecast = forecast_df['MyForecast'].max()
    min_forecast = forecast_df['MyForecast'].min()
    max_week = forecast_df.loc[forecast_df['MyForecast'].idxmax(), 'ds'] if 'ds' in forecast_df.columns else None
    min_week = forecast_df.loc[forecast_df['MyForecast'].idxmin(), 'ds'] if 'ds' in forecast_df.columns else None
    return summary_stats, total_forecast_16, total_forecast_8, total_forecast_4, max_forecast, min_forecast, max_week, min_week

def visualize_forecast_with_comparison(ts_data, comparison, summary_stats,
                                       total_forecast_16, total_forecast_8, total_forecast_4,
                                       max_forecast, min_forecast, max_week, min_week,
                                       asin, product_title, folder_name):
    """
    Visualize historical data vs. forecast (MyForecast + Amazon) and annotate with summary stats.
    """
    fig, ax = plt.subplots(figsize=(16, 10))
    ax.plot(ts_data['ds'], ts_data['y'], label='Historical Sales', marker='o', color='black')
    ax.plot(comparison['ds'], comparison['MyForecast'], label='MyForecast', marker='o', linestyle='--', color='blue')

    for column in comparison.columns:
        if column.startswith('Amazon ') and column.endswith('Forecast'):
            ax.plot(comparison['ds'], comparison[column], label=f'{column}', linestyle=':')

    ax.set_title(f'Sales Forecast Comparison for {product_title}')
    ax.set_xlabel('Date')
    ax.set_ylabel('Units Sold')
    ax.legend(loc='upper left', bbox_to_anchor=(1.05, 1))
    ax.grid()

    summary_text = (
        f"Historical Summary:\n"
        f"Range: {summary_stats['data_range'][0].strftime('%Y-%m-%d')} to {summary_stats['data_range'][1].strftime('%Y-%m-%d')}\n"
        f"Min: {summary_stats['min']:.0f}\n"
        f"Max: {summary_stats['max']:.0f}\n"
        f"Mean: {summary_stats['mean']:.0f}\n"
        f"Median: {summary_stats['median']:.0f}\n"
        f"Std Dev: {summary_stats['std_dev']:.0f}\n"
        f"Total Historical Sales: {summary_stats['total_sales']:.0f} units\n\n"
        f"Forecast Summary:\n"
        f"Total Forecast (16 Weeks): {total_forecast_16:.0f}\n"
        f"Total Forecast (8 Weeks): {total_forecast_8:.0f}\n"
        f"Total Forecast (4 Weeks): {total_forecast_4:.0f}\n"
        f"Max Forecast: {max_forecast:.0f} (Week of {max_week.strftime('%Y-%m-%d') if max_week else 'N/A'})\n"
        f"Min Forecast: {min_forecast:.0f} (Week of {min_week.strftime('%Y-%m-%d') if min_week else 'N/A'})"
    )

    plt.gcf().text(0.78, 0.5, summary_text, fontsize=10, bbox=dict(facecolor='white', alpha=0.8), va='center')
    plt.subplots_adjust(right=0.75)

    import re
    os.makedirs(folder_name, exist_ok=True)

    # sanitize product_title for Windows filenames
    safe_title = re.sub(r'[<>:"/\\|?*\s]+', '_', product_title).strip('_')
    graph_file_path = os.path.join(folder_name, f"{safe_title}_{asin}.png")
    plt.savefig(graph_file_path)
    plt.close()
    print(f"Graph saved to {graph_file_path}")


##############################
# Excel Output
##############################

def save_summary_to_excel(
    comparison,
    summary_stats,
    total_forecast_16,
    total_forecast_8,
    total_forecast_4,
    max_forecast,
    min_forecast,
    max_week,
    min_week,
    output_file_path,
    metrics=None
):
  
    # 1) Ensure we have 'Week_Start_Date' or create from 'ds' if it doesn't exist
    if 'Week_Start_Date' not in comparison.columns:
        if 'ds' in comparison.columns:
            # Convert 'ds' to string date
            comparison['Week_Start_Date'] = pd.to_datetime(
                comparison['ds'], errors='coerce'
            ).dt.strftime('%Y-%m-%d')
        else:
            # Fallback: create empty or placeholder
            comparison['Week_Start_Date'] = ''

    # 2) If "Week" is missing, create it from the sorted index (after sorting by 'Week_Start_Date').
    #    If "Week" already exists, we can still do a sort, but be sure to handle carefully.
    #    We'll unify the approach so everything gets sorted by 'Week_Start_Date' first.
    comparison = comparison.sort_values(
        'Week_Start_Date', na_position='last'
    ).reset_index(drop=True)

    if 'Week' not in comparison.columns:
        # Create "Week" labels as "W1", "W2", ...
        comparison['Week'] = ['W' + str(i + 1) for i in range(len(comparison))]

    # 3) Now that we have a sorted DataFrame with "Week_Start_Date" and "Week", we can proceed.

    # 4) Clean numeric columns if needed. For example, remove inf or NaN from forecast columns.
    forecast_cols = [
        "MyForecast",
        "Amazon Mean Forecast",
        "Amazon P70 Forecast",
        "Amazon P80 Forecast",
        "Amazon P90 Forecast"
    ]
    for fc in forecast_cols:
        if fc in comparison.columns:
            comparison[fc] = comparison[fc].replace([float('inf'), -float('inf')], 0).fillna(0)
            # Convert to int if you want integer forecasts
            comparison[fc] = comparison[fc].astype(int, errors='ignore')

    # 5) Reorder final columns to your old format or desired format:
    desired_cols = [
        "Week",
        "Week_Start_Date",
        "ASIN",
        "MyForecast",
        "Product Title",
        "is_holiday_week",
        # Add any new columns like "Trend", "Inventory Coverage", etc. if you wish:
        "Trend",
        "Inventory Coverage",
        "Stockout Risk",
        "Reorder Urgency",
        "Sales Trend",
        "Seasonality Index",
        "Lifecycle Stage"
    ]
    for col in desired_cols:
        if col not in comparison.columns:
            comparison[col] = ''  # create blank if missing

    # Filter the DataFrame to just the columns we want, in the order we want
    comparison = comparison[desired_cols]

    # 6) Write to Excel
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Forecast Comparison"

    # Write the main DataFrame
    from openpyxl.utils.dataframe import dataframe_to_rows
    for row in dataframe_to_rows(comparison, index=False, header=True):
        ws_main.append(row)

    # 7) Create a Summary sheet
    ws_summary = wb.create_sheet("Summary")
    summary_data = {
        "Metric": [
            "Historical Range",
            "Min Sales",
            "Max Sales",
            "Mean Sales",
            "Median Sales",
            "Std Dev Sales",
            "Total Historical Sales",
            "Total Forecast (16 Weeks)",
            "Total Forecast (8 Weeks)",
            "Total Forecast (4 Weeks)",
            "Max Forecast",
            "Max Forecast Week",
            "Min Forecast",
            "Min Forecast Week"
        ],
        "Value": [
            f"{summary_stats['data_range'][0].strftime('%Y-%m-%d')} to {summary_stats['data_range'][1].strftime('%Y-%m-%d')}" if 'data_range' in summary_stats else '',
            f"{summary_stats.get('min', 0):.0f}",
            f"{summary_stats.get('max', 0):.0f}",
            f"{summary_stats.get('mean', 0):.0f}",
            f"{summary_stats.get('median', 0):.0f}",
            f"{summary_stats.get('std_dev', 0):.0f}",
            f"{summary_stats.get('total_sales', 0):.0f} units",
            f"{total_forecast_16:.0f}",
            f"{total_forecast_8:.0f}",
            f"{total_forecast_4:.0f}",
            f"{max_forecast:.0f}",
            f"{max_week.strftime('%Y-%m-%d') if max_week else 'N/A'}",
            f"{min_forecast:.0f}",
            f"{min_week.strftime('%Y-%m-%d') if min_week else 'N/A'}"
        ]
    }
    if metrics is not None:
        # Append metrics to summary_data
        for k, v in metrics.items():
            summary_data["Metric"].append(k)
            summary_data["Value"].append(str(v))

    summary_df = pd.DataFrame(summary_data)
    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(row)

    # 8) Save workbook
    wb.save(output_file_path)
    print(f"Comparison and summary saved to '{output_file_path}'")


def save_4_8_16_forecast_summary(consolidated_data, output_folder):
    """
    Save the cumulative 4th, 8th, and 16th-week forecasts for each ASIN.
    """
    summary_data = []

    for asin, df in consolidated_data.items():
        if df.empty:
            summary_data.append([asin, "No data", None, None, None])
            continue

        product_title = df['Product Title'].iloc[0] if 'Product Title' in df.columns else ''
        
        df['Cumulative Forecast'] = df['MyForecast'].cumsum()

        four_wk_val = df['Cumulative Forecast'].iloc[3] if len(df) > 3 else None
        eight_wk_val = df['Cumulative Forecast'].iloc[7] if len(df) > 7 else None
        sixteen_wk_val = df['Cumulative Forecast'].iloc[15] if len(df) > 15 else None

        summary_data.append([asin, product_title, four_wk_val, eight_wk_val, sixteen_wk_val])

    summary_df = pd.DataFrame(
        summary_data,
        columns=["ASIN", "Product Title", "4th Week Cumulative Forecast", "8th Week Cumulative Forecast", "16th Week Cumulative Forecast"]
    )

    output_file_path = os.path.join(output_folder, "4-8-16_Weeks_Forecast_Summary.xlsx")
    summary_df.to_excel(output_file_path, index=False)
    print(f"4-8-16 Weeks Forecast Summary saved to {output_file_path}")

def save_forecast_to_excel(output_path, consolidated_data, missing_asin_data):
    """
    Save multiple ASIN forecasts (with "MyForecast") and any missing ASIN data into one Excel file,
    each ASIN in a separate sheet.
    
    This version enumerates from W0, W1, W2... for each row in ascending date order.
    """
    # Desired final column order.
    desired_columns = [
        "Week",
        "Week_Start_Date",
        "ASIN",
        "MyForecast",
        "Amazon Mean Forecast",
        "Amazon P70 Forecast",
        "Amazon P80 Forecast",
        "Amazon P90 Forecast",
        "Product Title",
        "is_holiday_week"
    ]

    # Columns to remove.
    unwanted_cols = [
        "ds", "yhat", "yhat_upper", "Diff_Mean", "Diff_P70", "Diff_P80", "Diff_P90",
        "Pct_Mean", "Pct_P70", "Pct_P80", "Pct_P90", "MyForecast_XGB", "y", "Amazon Mean Forecast",
        "Amazon P70 Forecast",
        "Amazon P80 Forecast",
        "Amazon P90 Forecast",
    ]

    # Columns to be cleaned.
    forecast_cols = [
        "MyForecast",
        "Amazon Mean Forecast",
        "Amazon P70 Forecast",
        "Amazon P80 Forecast",
        "Amazon P90 Forecast"
    ]

    wb = Workbook()

    for asin, forecast_df in consolidated_data.items():
        df_for_excel = forecast_df.copy()

        # --- Step 1: Drop unwanted columns ---
        for col in unwanted_cols:
            if col in df_for_excel.columns:
                df_for_excel.drop(columns=[col], inplace=True, errors='ignore')

        # --- Step 2: If "ds" but no "Week_Start_Date", rename "ds" to "Week_Start_Date" ---
        if "ds" in df_for_excel.columns and "Week_Start_Date" not in df_for_excel.columns:
            df_for_excel["ds"] = pd.to_datetime(df_for_excel["ds"], errors="coerce")
            df_for_excel["Week_Start_Date"] = df_for_excel["ds"].dt.strftime("%Y-%m-%d")
            df_for_excel.drop(columns=["ds"], inplace=True)

        # --- Step 3: If "Week_Start_Date" present, label from W0 upward ---
        if "Week_Start_Date" in df_for_excel.columns:
            df_for_excel["Week_Start_Date"] = pd.to_datetime(df_for_excel["Week_Start_Date"], errors="coerce")
            # Sort by earliest date
            df_for_excel = df_for_excel.sort_values("Week_Start_Date", na_position='last').reset_index(drop=True)
            # Label row 0 as W0, row 1 as W1, etc.
            df_for_excel["Week"] = [f"W{i}" for i in range(len(df_for_excel))]
            # Convert back to string
            df_for_excel["Week_Start_Date"] = df_for_excel["Week_Start_Date"].dt.strftime("%Y-%m-%d")
        else:
            # Fallback if no "Week_Start_Date" => enumerating from W0
            df_for_excel = df_for_excel.reset_index(drop=True)
            df_for_excel["Week"] = [f"W{i}" for i in range(len(df_for_excel))]
            df_for_excel["Week_Start_Date"] = ""

        # --- Step 4: Clean forecast columns ---
        for col in forecast_cols:
            if col in df_for_excel.columns:
                df_for_excel[col] = df_for_excel[col].replace([np.inf, -np.inf], 0).fillna(0)
                df_for_excel[col] = df_for_excel[col].round().astype(int, errors='ignore')

        # --- Step 5: Ensure all desired columns exist ---
        for col in desired_columns:
            if col not in df_for_excel.columns:
                df_for_excel[col] = np.nan

        # --- Step 6: Fill 'ASIN' and 'Product Title' if missing ---
        df_for_excel["ASIN"] = asin
        if "Product Title" in forecast_df.columns and not forecast_df["Product Title"].empty:
            df_for_excel["Product Title"] = forecast_df["Product Title"].iloc[0]
        else:
            df_for_excel["Product Title"] = ""

        # --- Step 7: Reorder columns ---
        df_for_excel = df_for_excel[desired_columns]

        # --- Step 8: Write this ASIN's data to a new worksheet ---
        ws = wb.create_sheet(title=str(asin)[:31])
        for row in dataframe_to_rows(df_for_excel, index=False, header=True):
            ws.append(row)

    # --- Step 9: Missing ASINs as a separate sheet ---
    if not missing_asin_data.empty:
        ws_missing = wb.create_sheet(title="No ASIN")
        for row in dataframe_to_rows(missing_asin_data, index=False, header=True):
            ws_missing.append(row)

    # --- Step 10: Remove default sheet if >1 sheets exist ---
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # --- Step 11: Create a 4/8/16-Week Summary sheet ---
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["ASIN", "Product Title", "4 Week Forecast", "8 Week Forecast", "16 Week Forecast"])

    for asin, df in consolidated_data.items():
        if df.empty:
            ws_summary.append([asin, "No data", None, None, None])
            continue
        product_title = df["Product Title"].iloc[0] if "Product Title" in df.columns else ""
        if "MyForecast" in df.columns:
            df["MyForecast"] = df["MyForecast"].replace([np.inf, -np.inf], 0).fillna(0)
            four_wk_val = df["MyForecast"].iloc[:4].sum() if len(df) >= 4 else df["MyForecast"].sum()
            eight_wk_val = df["MyForecast"].iloc[:8].sum() if len(df) >= 8 else df["MyForecast"].sum()
            sixteen_wk_val = df["MyForecast"].iloc[:16].sum() if len(df) >= 16 else df["MyForecast"].sum()
        else:
            four_wk_val = eight_wk_val = sixteen_wk_val = 0
        ws_summary.append([
            asin,
            product_title,
            int(round(four_wk_val)),
            int(round(eight_wk_val)),
            int(round(sixteen_wk_val))
        ])

    # --- Step 12: Save the workbook ---
    try:
        wb.save(output_path)
        print(f"All forecasts saved to '{output_path}'")
    except Exception as e:
        print(f"Failed to save Excel file '{output_path}': {e}")


def save_feedback_to_excel(feedback_dict, filename):
    """
    Save feedback information from models into an Excel file.
    """
    records = []
    for asin, info in feedback_dict.items():
        record = {'ASIN': asin}
        best_params = info.get('best_params', {})
        for param, value in best_params.items():
            record[param] = value
        if 'rmse_values' in info and info['rmse_values']:
            for k, v in info['rmse_values'].items():
                record[f'RMSE_{k}'] = v
        record['Total Tests'] = info.get('total_tests', None)
        records.append(record)
    
    df_feedback = pd.DataFrame(records)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_feedback.to_excel(writer, index=False, sheet_name='Model Feedback')
    print(f"Feedback saved to {filename}")

def generate_4_week_report(consolidated_forecasts):
    """
    Generate an Excel report for the first 4 weeks of "MyForecast" for each ASIN,
    plus Amazon columns if present.
    """
    report_rows = []
    for asin, comp_df in consolidated_forecasts.items():
        # Determine the model used based on available info
        # (We only unify the final forecast column as "MyForecast")
        product_title = ''
        if 'Product Title' in comp_df.columns and not comp_df.empty:
            product_title = comp_df['Product Title'].iloc[0]

        if 'MyForecast' in comp_df.columns:
            model_used = 'Unified'
            my_forecast_column = 'MyForecast'
        else:
            model_used = 'Unknown'
            my_forecast_column = None

        if my_forecast_column and my_forecast_column in comp_df.columns:
            my_4w_forecast = comp_df[my_forecast_column].iloc[:4].sum()
        else:
            my_4w_forecast = np.nan

        amz_mean = comp_df['Amazon Mean Forecast'].iloc[:4].sum() if 'Amazon Mean Forecast' in comp_df.columns else np.nan
        amz_p70  = comp_df['Amazon P70 Forecast'].iloc[:4].sum() if 'Amazon P70 Forecast' in comp_df.columns else np.nan
        amz_p80  = comp_df['Amazon P80 Forecast'].iloc[:4].sum() if 'Amazon P80 Forecast' in comp_df.columns else np.nan
        amz_p90  = comp_df['Amazon P90 Forecast'].iloc[:4].sum() if 'Amazon P90 Forecast' in comp_df.columns else np.nan

        report_rows.append({
            'ASIN': asin,
            'Product Title': product_title,
            'My 4 Weeks Forecast': my_4w_forecast,
            'AMZ Forecast Mean': amz_mean,
            'AMZ Forecast P70': amz_p70,
            'AMZ Forecast P80': amz_p80,
            'AMZ Forecast P90': amz_p90
        })

    report_df = pd.DataFrame(report_rows)
    report_filename = '4_week_report.xlsx'
    report_df.to_excel(report_filename, index=False)
    print(f"4-week report saved to {report_filename}")

def generate_combined_weekly_report(consolidated_forecasts):
    """
    Generate an Excel report summarizing 4-, 8-, and 16-week forecasts for all ASINs.
    This function ensures that every ASIN in the consolidated_forecasts dictionary is
    included, even if its forecast DataFrame is empty or missing the 'MyForecast' column.
    
    Output Columns: ASIN, Product Title, 4 Weeks Forecast, 8 Weeks Forecast, 16 Weeks Forecast.
    """
    report_rows = []
    for asin, comp_df in consolidated_forecasts.items():
        # Get product name if available; otherwise leave empty.
        product_name = ""
        if 'Product Title' in comp_df.columns and not comp_df.empty:
            product_name = comp_df['Product Title'].iloc[0]
        
        # Ensure that we have a 'MyForecast' column.
        if 'MyForecast' not in comp_df.columns or comp_df.empty:
            # If missing, create a Series of zeros with the same number of rows as comp_df.
            my_forecast = pd.Series([0] * len(comp_df))
        else:
            my_forecast = comp_df['MyForecast']
        
        # Calculate cumulative sums for the first 4, 8, and 16 weeks.
        forecast_4 = int(round(my_forecast.iloc[:4].sum())) if len(my_forecast) >= 4 else int(round(my_forecast.sum()))
        forecast_8 = int(round(my_forecast.iloc[:8].sum())) if len(my_forecast) >= 8 else int(round(my_forecast.sum()))
        forecast_16 = int(round(my_forecast.iloc[:16].sum())) if len(my_forecast) >= 16 else int(round(my_forecast.sum()))
        
        report_rows.append({
            'ASIN': asin,
            'Product Title': product_name,
            '4 Weeks Forecast': forecast_4,
            '8 Weeks Forecast': forecast_8,
            '16 Weeks Forecast': forecast_16
        })
    report_df = pd.DataFrame(report_rows)
    report_filename = 'combined_4_8_16_week_report.xlsx'
    report_df.to_excel(report_filename, index=False)
    print(f"Combined 4-8-16 week report saved to {report_filename}")

def save_consolidated_forecasts(output_path, consolidated_data, base_year=2025):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    desired_columns = [
        'Week', 'Week_Start_Date', 'ASIN', 'MyForecast', 'Amazon Mean Forecast',
        'Amazon P70 Forecast', 'Amazon P80 Forecast', 'Amazon P90 Forecast',
        'Product Title', 'is_holiday_week'
    ]

    def week_label_to_date(week_label, base_year):
        week_num = int(week_label[1:])
        start_date = pd.Timestamp(f'{base_year}-01-01') + pd.Timedelta(weeks=week_num-1)
        return start_date.strftime('%Y-%m-%d')

    consolidated_list = []
    for asin, df in consolidated_data.items():
        df_copy = df.copy()

        # Calculate Week_Start_Date
        if 'ds' in df_copy.columns:
            df_copy['ds'] = pd.to_datetime(df_copy['ds'], errors='coerce')
            df_copy['Week_Start_Date'] = df_copy['ds'].dt.strftime('%Y-%m-%d')
        elif 'Week' in df_copy.columns:
            df_copy['Week_Start_Date'] = df_copy['Week'].apply(lambda w: week_label_to_date(w, base_year))

        # Round forecast columns to integers
        forecast_cols = ['MyForecast', 'Amazon Mean Forecast', 'Amazon P70 Forecast', 
                         'Amazon P80 Forecast', 'Amazon P90 Forecast']
        for col in forecast_cols:
            if col in df_copy.columns:
                df_copy[col] = df_copy[col].round().astype(int)

        # Create Week labels if missing
        if 'Week' not in df_copy.columns:
            df_copy = df_copy.sort_values('Week_Start_Date').reset_index(drop=True)
            df_copy['Week'] = ['W' + str(i+1).zfill(2) for i in range(len(df_copy))]

        # Ensure all desired columns exist
        for col in desired_columns:
            if col not in df_copy.columns:
                df_copy[col] = np.nan

        # Select desired columns only
        df_copy = df_copy[desired_columns]

        consolidated_list.append(df_copy)

    # Combine all ASIN dataframes
    if consolidated_list:
        consolidated_df = pd.concat(consolidated_list, ignore_index=True)
        consolidated_df = consolidated_df.sort_values(['Week_Start_Date', 'ASIN']).reset_index(drop=True)
    else:
        consolidated_df = pd.DataFrame(columns=desired_columns)

    # Save the consolidated DataFrame to Excel
    with pd.ExcelWriter(output_path) as writer:
        consolidated_df.to_excel(writer, sheet_name='Consolidated Forecast', index=False)
    print(f"All forecasts saved to '{output_path}'")


##############################
# Additional Prophet Cross-Validation
##############################

def cross_validate_prophet_model(ts_data, initial='180 days', period='90 days', horizon='90 days'):
    """Run Prophet's built-in cross-validation and performance metrics."""
    model = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=True,
        seasonality_mode='multiplicative'
    )
    model.fit(ts_data[['ds', 'y']])
    df_cv = cross_validation(model, initial=initial, period=period, horizon=horizon)
    df_performance = performance_metrics(df_cv)
    print("Prophet Model Cross-Validation Results:")
    print(df_performance)
    return df_cv, df_performance


##############################
# Analysis of Amazon vs. Prophet
##############################

def analyze_amazon_buying_habits(comparison, holidays):
    """
    Analyze Amazon forecast columns vs. 'MyForecast', highlighting holiday weeks, etc.
    """
    if 'y' not in comparison.columns:
        print("Warning: 'y' column missing in the comparison DataFrame. Skipping analysis.")
        return

    amazon_cols = [
        col for col in comparison.columns
        if col.startswith('Amazon ') and col.endswith('Forecast') and not col.endswith('Forecast Forecast')
    ]
    if not amazon_cols:
        print("No Amazon forecasts found for analysis.")
        return

    myforecast_series = comparison.get('MyForecast', pd.Series(index=comparison.index))
    ds_dates = comparison.get('ds', pd.Series(index=comparison.index))
    holiday_dates = holidays['ds'].values if holidays is not None else []
    comparison['is_holiday_week'] = ds_dates.isin(holiday_dates) if 'ds' in comparison.columns else False

    amazon_types = ['Mean', 'P70', 'P80', 'P90']
    errors = {}
    for forecast_type in amazon_types:
        forecast_col = f"Amazon {forecast_type} Forecast"
        if forecast_col in comparison.columns:
            valid_data = comparison[['y', forecast_col]].dropna()
            if len(valid_data) == 0:
                continue
            current_rmse = np.sqrt(mean_squared_error(comparison['y'], comparison[forecast_col]))
            current_mae = mean_absolute_error(comparison['y'], comparison[forecast_col]) * 100
            errors[forecast_type] = {
                'RMSE': current_rmse,
                'MAE': current_mae
            }

    if errors:
        best_forecast_type = min(errors, key=lambda x: errors[x]['RMSE'])
        print(f"\nBest Amazon forecast type: {best_forecast_type} with RMSE={errors[best_forecast_type]['RMSE']:.4f}")
    
    for forecast_type in amazon_types:
        forecast_col = f"Amazon {forecast_type} Forecast"
        if forecast_col not in comparison.columns:
            continue
        amazon_vals = comparison[forecast_col].values
        safe_myforecast = np.where(myforecast_series == 0, 1e-9, myforecast_series)
        ratio = amazon_vals / safe_myforecast
        diff = amazon_vals - myforecast_series
        avg_ratio = np.mean(ratio)
        avg_diff = np.mean(diff)
        print(f"\nFor {forecast_type}:")
        print(f"  Average Amazon/MyForecast Ratio: {avg_ratio:.2f}")
        print(f"  Average Difference (Amazon - MyForecast): {avg_diff:.2f}")

        if avg_diff > 0:
            print("  Amazon tends to forecast more than MyForecast on average.")
        elif avg_diff < 0:
            print("  Amazon tends to forecast less than MyForecast on average.")
        else:
            print("  Amazon forecasts similarly to MyForecast on average.")

        holiday_mask = comparison['is_holiday_week']
        if holiday_mask.any():
            holiday_ratio = ratio[holiday_mask]
            holiday_diff = diff[holiday_mask]
            if len(holiday_diff) > 0:
                print("  During holiday weeks:")
                print(f"    Avg Ratio (Amazon/MyForecast): {np.mean(holiday_ratio):.2f}")
                print(f"    Avg Diff (Amazon-MyForecast): {np.mean(holiday_diff):.2f}")

        weeks = np.arange(1, len(ratio) + 1)
        segments = {
            'Short-term (Weeks 1-4)': (weeks <= 4),
            'Mid-term (Weeks 5-12)': (weeks >= 5) & (weeks <= 12),
            'Long-term (Weeks 13+)': (weeks > 12)
        }

        for segment_name, mask in segments.items():
            if mask.any():
                seg_ratio = ratio[mask]
                seg_diff = diff[mask]
                print(f"  {segment_name}:")
                print(f"    Avg Ratio (Amazon/MyForecast): {np.mean(seg_ratio):.2f}")
                print(f"    Avg Diff (Amazon-MyForecast): {np.mean(seg_diff):.2f}")


##############################
# Additional/Custom Backtesting Function
##############################

def calculate_extended_metrics(actual, predicted):
    """
    Calculate extended metrics (RMSE, wQL, MASE, MAE, WAPE, etc.) for demonstration.
    """
    actual = np.array(actual)
    predicted = np.array(predicted)
    
    # 1. Root Mean Squared Error (RMSE)
    rmse = np.sqrt(mean_squared_error(actual, predicted))
    
    # 2. Mean Absolute Error (MAE)
    mae = mean_absolute_error(actual, predicted)
    
    # 3. Weighted Quantile Loss (wQL) and Average wQL
    alpha = 0.5  # For simplistic quantile approach; adjust as needed
    residuals = actual - predicted
    quantile_loss = np.where(residuals >= 0, alpha * residuals, (1 - alpha) * -residuals)
    wql = np.sum(np.abs(quantile_loss))
    avg_wql = np.mean(np.abs(quantile_loss)) if len(quantile_loss) > 0 else 0
    
    # 4. Weighted Absolute Percentage Error (WAPE)
    total_actual = np.sum(np.abs(actual))
    if total_actual != 0:
        wape = np.sum(np.abs(actual - predicted)) / total_actual * 100
    else:
        wape = 0
    
    # 5. Mean Absolute Scaled Error (MASE)
    # Naive forecast: using the previous period's actual value
    # Shift actual by one period to get naive forecast
    if len(actual) > 1:
        naive_forecast = actual[:-1]
        actual_shifted = actual[1:]
        naive_error = np.abs(actual_shifted - naive_forecast)
        mae_naive = np.mean(naive_error) if len(naive_error) > 0 else 1
    else:
        mae_naive = 1  # Avoid division by zero
    
    mase = mae / mae_naive if mae_naive != 0 else np.nan
    
    return {
        "RMSE": rmse,
        "MAE": mae,
        "wQL": wql,
        "Average wQL": avg_wql,
        "MASE": mase,
        "WAPE": wape
    }

def backtest_forecast(ts_data, model_function, horizon=16, folder_name="Training and Testing data"):
    """
    Example backtesting function for demonstration.
    """
    os.makedirs(folder_name, exist_ok=True)
    ts_data = ts_data.sort_values('ds').copy()
    n = len(ts_data)
    split_idx = int(n * 0.8)
    train = ts_data.iloc[:split_idx].copy()
    test = ts_data.iloc[split_idx:].copy()

    fitted_model = model_function(train)

    try:
        last_date = train['ds'].iloc[-1]
        steps = len(test)
        test_forecast_df = pd.DataFrame({
            'ds': test['ds'],
            'MyForecast': [train['y'].iloc[-1]] * steps
        })
    except Exception as e:
        print(f"Error in backtest forecasting step: {e}")
        return

    backtest_comparison = test_forecast_df.merge(test[['ds','y']], on='ds', how='left')

    metrics = calculate_extended_metrics(backtest_comparison['y'], backtest_comparison['MyForecast'])
    print("Backtest Metrics:")
    for k, v in metrics.items():
        print(f"  {k}: {v}")

    train_file = os.path.join(folder_name, "training_data.csv")
    test_file = os.path.join(folder_name, "testing_data.csv")
    train.to_csv(train_file, index=False)
    test.to_csv(test_file, index=False)

    plt.figure(figsize=(10,6))
    plt.plot(train['ds'], train['y'], label='Training', color='blue')
    plt.plot(test['ds'], test['y'], label='Testing Actual', color='orange')
    plt.plot(backtest_comparison['ds'], backtest_comparison['MyForecast'], label='Testing Forecast', color='green')
    plt.legend()
    plt.title("Backtest: Training vs Testing Forecast")
    plt.xlabel("Date")
    plt.ylabel("Values")
    backtest_plot_path = os.path.join(folder_name, "training_testing_graph.png")
    plt.savefig(backtest_plot_path)
    plt.close()
    print(f"Training/Testing data and graph saved to '{folder_name}'.")


##############################
# Function to Optimize Ensemble Weights
##############################

def optimize_ensemble_weights(actual, sarima_preds, prophet_preds, amazon_mean_preds):
    """
    Optimize ensemble weights for forecasts from SARIMA, Prophet, and Amazon Mean.
    Returns a dictionary with optimized weights and performance metrics.
    """
    initial_weights = [1/3, 1/3, 1/3]
    constraints = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
    bounds = [(0, 1), (0, 1), (0, 1)]

    def objective(weights):
        ensemble_preds = (
            weights[0] * sarima_preds +
            weights[1] * prophet_preds +
            weights[2] * amazon_mean_preds
        )
        rmse_val = np.sqrt(mean_squared_error(actual, ensemble_preds))
        mae_val = mean_absolute_error(actual, ensemble_preds)
        return rmse_val + mae_val

    result = minimize(objective, initial_weights, constraints=constraints, bounds=bounds, method='SLSQP')
    if result.success:
        optimized_weights = result.x
        ensemble_preds = (
            optimized_weights[0] * sarima_preds +
            optimized_weights[1] * prophet_preds +
            optimized_weights[2] * amazon_mean_preds
        )
        rmse_val = np.sqrt(mean_squared_error(actual, ensemble_preds))
        mae_val = mean_absolute_error(actual, ensemble_preds)
        return {
            'weights': {
                'SARIMA': optimized_weights[0],
                'Prophet': optimized_weights[1],
                'Amazon Mean': optimized_weights[2]
            },
            'metrics': {
                'RMSE': rmse_val,
                'MAE': mae_val
            },
            'ensemble_preds': ensemble_preds
        }
    else:
        raise ValueError("Optimization failed: " + result.message)

##############################
# Adjust Forecast If Out of Range
##############################

def adjust_forecast_if_out_of_range(
    comparison,
    asin,
    forecast_col_name='MyForecast',
    adjustment_threshold=0.3,
    bypass_seasonality=True,
    window=6,                # use 6 weeks to compute recent average sales
    min_forecast=None,       # optional: lower bound for final forecast
    max_forecast=None,       # optional: upper bound for final forecast
    amazon_forecast_series=None  # optional: external Amazon forecast series if not in comparison
):
    """
    Adjust 'MyForecast' using a dynamic adjustment factor based on the ratio of
    the recent actual sales average (over the last 'window' weeks, from column 'y')
    to the overall average Amazon Mean Forecast.
    
    The dynamic factor is computed as:
    
        dynamic_factor = (recent_actual_avg) / (mean_amz_forecast)
    
    This factor is bounded between min_factor (e.g. 0.1) and max_factor (e.g. 2.0) to avoid extreme adjustments.
    
    The new forecast is then:
    
         New Forecast = round(Amazon Mean Forecast × dynamic_factor)
    
    Optionally, the new forecast is clipped between min_forecast and max_forecast if provided.
    
    If the Amazon Mean Forecast column is missing and an external Series is provided, it is added.
    Otherwise, if there isn’t enough data to compute a dynamic factor, the function falls back to
    the existing seasonal adjustment logic.
    """
    global out_of_range_counter
    global out_of_range_stats

    # Ensure the Amazon Mean Forecast column exists.
    if 'Amazon Mean Forecast' not in comparison.columns:
        if amazon_forecast_series is not None:
            comparison['Amazon Mean Forecast'] = amazon_forecast_series
        else:
            print(f"Warning: 'Amazon Mean Forecast' column not found for ASIN {asin}. Skipping adjustment.")
            return comparison

    # STEP A: Compute recent actual average over the last 'window' weeks from column 'y'
    recent_actual_avg = 0
    if 'y' in comparison.columns:
        recent_rows = comparison.dropna(subset=['y']).sort_values('ds').tail(window)
        if not recent_rows.empty:
            recent_actual_avg = recent_rows['y'].mean()

    # STEP B: Compute overall average Amazon Mean Forecast from the DataFrame
    mean_amz_forecast = comparison['Amazon Mean Forecast'].mean() if len(comparison) > 0 else 0

    # Only compute dynamic factor if we have valid numbers.
    if mean_amz_forecast > 0 and recent_actual_avg > 0:
        dynamic_factor = recent_actual_avg / mean_amz_forecast
        # Bound the factor to avoid extreme changes.
        min_factor = 0.1
        max_factor = 2.0
        dynamic_factor = max(min_factor, min(dynamic_factor, max_factor))
        # Only apply if the factor is meaningfully different from 1 (say <0.9 or >1.1)
        if dynamic_factor < 0.9 or dynamic_factor > 1.1:
            print(f"For ASIN {asin}: Recent {window}-week actual avg = {recent_actual_avg:.1f}, "
                  f"Amazon Mean forecast avg = {mean_amz_forecast:.1f}. "
                  f"Applying dynamic adjustment factor {dynamic_factor:.2f}.")
            new_forecast = (comparison['Amazon Mean Forecast'] * dynamic_factor).round().clip(lower=0)
            if min_forecast is not None:
                new_forecast = new_forecast.clip(lower=min_forecast)
            if max_forecast is not None:
                new_forecast = new_forecast.clip(upper=max_forecast)
            comparison[forecast_col_name] = new_forecast
            return comparison

    # If dynamic factor is near 1 (i.e. recent actuals similar to Amazon forecast), 
    # proceed with the existing seasonal adjustment logic.
    if bypass_seasonality and 'seasonal_cluster' in comparison.columns:
        comparison['bypass_adjustment'] = comparison['seasonal_cluster'].apply(lambda x: x in ['high', 'low'])
    else:
        comparison['bypass_adjustment'] = False

    if 'y' in comparison.columns:
        recent_5 = comparison.dropna(subset=['y']).sort_values('ds').tail(5)
        last_5_avg = recent_5['y'].mean() if len(recent_5) > 0 else 0
        overall_avg = comparison['y'].mean()
    else:
        last_5_avg = 0
        overall_avg = 0

    local_threshold = 0.6 if last_5_avg < overall_avg else adjustment_threshold
    if last_5_avg < overall_avg:
        print(f"Recent 5-week avg is below overall avg, using threshold=60% for ASIN {asin}")

    condition_adjust = (
        ((comparison[forecast_col_name] < comparison['Amazon Mean Forecast'] * (1 - local_threshold)) |
         (comparison[forecast_col_name] > comparison['Amazon Mean Forecast'] * (1 + local_threshold)))
        & (~comparison['bypass_adjustment'])
    )
    comparison['is_out_of_range'] = condition_adjust

    if comparison['is_out_of_range'].any():
        num_adjustments = comparison['is_out_of_range'].sum()
        print(f"Adjusting {num_adjustments} out-of-range forecasts for ASIN {asin}.")
        out_of_range_counter[asin] += num_adjustments
        out_of_range_stats.setdefault(asin, {'total': 0, 'adjusted': 0})
        out_of_range_stats[asin]['total'] += len(comparison)
        out_of_range_stats[asin]['adjusted'] += num_adjustments

        idx = comparison['is_out_of_range']
        # This logic clips the ratio between (1 ± local_threshold) and then multiplies by the Amazon Mean Forecast.
        adjusted_values = (
            comparison.loc[idx, 'Amazon Mean Forecast']
            * comparison.loc[idx, forecast_col_name]
            / comparison.loc[idx, 'Amazon Mean Forecast']
        ).clip(lower=(1 - local_threshold), upper=(1 + local_threshold)) * comparison.loc[idx, 'Amazon Mean Forecast']
        comparison.loc[idx, forecast_col_name] = adjusted_values.clip(lower=0)
    
    comparison.drop(columns=['bypass_adjustment'], inplace=True, errors='ignore')
    return comparison


def log_fallback_triggers(comparison, asin, product_title, fallback_file="fallback_triggers.csv"):
    """
    Logs products where the fallback mechanism was triggered to a separate file,
    based on the 'is_out_of_range' flag set in adjust_forecast_if_out_of_range().
    """
    # Look for the flag your adjustment function actually creates:
    if 'is_out_of_range' not in comparison.columns:
        print(f"No 'is_out_of_range' column present in DataFrame for ASIN: {asin}. No fallback to log.")
        return

    # Rows where we adjusted the forecast
    fallback_rows = comparison[comparison['is_out_of_range']]
    if not fallback_rows.empty:
        print(f"Outlier detected for ASIN: {asin}, Product: {product_title}")
        fallback_info = {
            "ASIN": [asin],
            "Product Title": [product_title],
            "Outlier Weeks": [fallback_rows.get('ds', fallback_rows.get('Week')).tolist()],
            "Total Adjustments": [len(fallback_rows)]
        }

        fallback_df = pd.DataFrame(fallback_info)

        if os.path.exists(fallback_file):
            existing_data = pd.read_csv(fallback_file)
            fallback_df = pd.concat([existing_data, fallback_df], ignore_index=True)

        fallback_df.to_csv(fallback_file, index=False)
        print(f"Fallback log updated: {fallback_file}")
    else:
        print(f"No outliers detected for ASIN: {asin}")


##############################
# Feedback Loop Functions
##############################

def record_forecast_errors(asin, forecast_df, actual_df):
    """
    Compare 'MyForecast' vs. actual data for a given ASIN and record errors.
    """
    merged = forecast_df.merge(actual_df[['ds', 'y']], on='ds', how='left')
    merged = merged.dropna(subset=['y'])
    if merged.empty:
        print(f"No actual data available yet to compare for ASIN {asin}.")
        return
    mae = mean_absolute_error(merged['y'], merged['MyForecast'])
    mape = mean_absolute_error(merged['y'], merged['MyForecast']) * 100
    forecast_errors.setdefault(asin, []).append({'mae': mae, 'mape': mape})
    print(f"Recorded forecast errors for ASIN {asin}: MAE={mae}, MAPE={mape}%")


def update_prophet_model_with_feedback(
    asin,
    ts_data,
    forecast_data,
    horizon=16,
    current_model=None,
    n_trials=30
):
    """
    Update the Prophet model for a given ASIN using new actual data,
    leveraging Optuna-based hyperparameter optimization.

    Steps:
      1) Use Optuna to find best hyperparams for the given ts_data + forecast_data.
      2) Train a final Prophet model on ALL data using those best hyperparams.
      3) Generate a horizon-week forecast (or however many weeks you want).
      4) Save the model to disk (or model_cache) for future use.
    """
    print(f"[Optuna-Update] Updating Prophet model for ASIN {asin} with new data...")

    # 1) Run Bayesian optimization via Optuna
    best_params, _ = optimize_prophet_params_with_optuna(
        ts_data      = ts_data,
        forecast_data= forecast_data,
        horizon      = horizon,
        asin         = asin,
        n_trials     = n_trials
    )

    # 2) Build a final Prophet model with the best hyperparams
    updated_model = Prophet(
        yearly_seasonality      = True,
        weekly_seasonality      = True,
        daily_seasonality       = False,
        changepoint_prior_scale = best_params['changepoint_prior_scale'],
        seasonality_prior_scale = best_params['seasonality_prior_scale'],
        holidays_prior_scale    = best_params['holidays_prior_scale'],
        seasonality_mode        = best_params['seasonality_mode'],
        growth                  = 'linear'
    )

    updated_model.fit(ts_data[['ds','y']])

    # 3) Forecast for the next 'horizon' weeks
    future_dates = pd.date_range(
        start=ts_data['ds'].max() + pd.Timedelta(weeks=1),
        periods=horizon,
        freq='W-SUN'
    )
    future_df = pd.DataFrame({'ds': future_dates})
    forecast = updated_model.predict(future_df)
    forecast['MyForecast'] = forecast['yhat'].clip(lower=0).round()

    # 4) Save the updated model if desired
    save_model(updated_model, "Prophet", asin, ts_data)
    print(f"[Optuna-Update] Prophet model for ASIN {asin} updated with best_params={best_params}.")

    return forecast, updated_model


##############################
# Analyze PO Order
##############################

def analyze_po_data_by_asin(po_file="po database.xlsx", product_filter=None, output_folder="po_analysis_by_asin"):
    """
    Analyze PO orders by each ASIN, generating weekly/monthly requested quantity charts,
    and fit a Prophet model to forecast future PO volume.
    Writes an Excel file for each ASIN (with sheets for Weekly Quantity, Monthly Trend, and PO Forecast),
    and returns a dictionary (po_forecasts_dict) containing each ASIN's PO forecast (columns: [ds, PO_Forecast]).
    """
    po_df = pd.read_excel(po_file)
    po_df.columns = po_df.columns.str.strip()

    for date_col in ['Order date', 'Expected date']:
        if date_col in po_df.columns:
            po_df[date_col] = pd.to_datetime(po_df[date_col], errors='coerce')

    if product_filter:
        po_df = po_df[po_df['ASIN'] == product_filter]

    po_df['Requested quantity'] = pd.to_numeric(po_df['Requested quantity'], errors='coerce').fillna(0)
    os.makedirs(output_folder, exist_ok=True)

    po_forecasts_dict = {}
    unique_asins = po_df['ASIN'].dropna().unique()

    for asin in unique_asins:
        asin_df = po_df[po_df['ASIN'] == asin].copy()
        if asin_df.empty:
            continue

        # Weekly aggregation
        asin_df['Order_Week_Period'] = asin_df['Order date'].dt.to_period('W-SUN')
        asin_df['Order Week'] = asin_df['Order_Week_Period'].apply(lambda p: p.end_time)
        weekly_agg = (asin_df.groupby('Order Week', as_index=False)['Requested quantity']
                      .sum().rename(columns={'Requested quantity': 'Weekly_PO_Qty'}))
        # Monthly aggregation (if needed)
        asin_df['Order_Month_Period'] = asin_df['Order date'].dt.to_period('M')
        asin_df['Order Month'] = asin_df['Order_Month_Period'].apply(lambda p: p.end_time)
        monthly_trend = (asin_df.groupby('Order Month', as_index=False)['Requested quantity']
                         .sum().rename(columns={'Requested quantity': 'Monthly_PO_Qty'}))

        # Plot weekly and monthly charts (unchanged)
        plt.figure(figsize=(10,6))
        plt.plot(weekly_agg['Order Week'], weekly_agg['Weekly_PO_Qty'], marker='o')
        plt.title(f'Weekly Requested Quantities for ASIN {asin}')
        plt.xlabel('Week')
        plt.ylabel('Total Requested Quantity')
        plt.grid()
        plt.xticks(rotation=45)
        plt.tight_layout()
        weekly_chart = os.path.join(output_folder, f"{asin}_weekly_requested_quantity.png")
        plt.savefig(weekly_chart)
        plt.close()

        plt.figure(figsize=(10,6))
        plt.plot(monthly_trend['Order Month'], monthly_trend['Monthly_PO_Qty'], marker='o', color='green')
        plt.title(f'Monthly Requested Quantities for ASIN {asin}')
        plt.xlabel('Month')
        plt.ylabel('Total Requested Quantity')
        plt.grid()
        plt.xticks(rotation=45)
        plt.tight_layout()
        monthly_chart = os.path.join(output_folder, f"{asin}_monthly_requested_quantity.png")
        plt.savefig(monthly_chart)
        plt.close()

        # Prophet forecast on weekly data
        forecast_po = pd.DataFrame()
        if len(weekly_agg) >= 2:
            try:
                weekly_for_prophet = weekly_agg.rename(columns={'Order Week': 'ds', 'Weekly_PO_Qty': 'y'}).copy()
                weekly_for_prophet['ds'] = pd.to_datetime(weekly_for_prophet['ds'], errors='coerce')
                weekly_for_prophet = weekly_for_prophet.dropna(subset=['ds','y']).sort_values('ds')
                m = Prophet(weekly_seasonality=True)
                m.fit(weekly_for_prophet[['ds','y']])
                future = m.make_future_dataframe(periods=8, freq='W-SUN')
                forecast_po = m.predict(future)
                forecast_po['PO_Forecast'] = forecast_po['yhat'].clip(lower=0).round().astype(int)
                # Optionally plot forecast
                fig = m.plot(forecast_po, xlabel='Date', ylabel='PO Qty')
                plt.title(f'Prophet Forecast - PO Qty for ASIN {asin}')
                forecast_chart = os.path.join(output_folder, f"{asin}_po_forecast.png")
                fig.savefig(forecast_chart)
                plt.close(fig)
            except Exception as ex:
                print(f"Prophet forecast failed for ASIN {asin}: {ex}")
        po_forecasts_dict[asin] = forecast_po[['ds','PO_Forecast']].copy() if not forecast_po.empty else pd.DataFrame()

        # Write an Excel file for this ASIN (optional)
        excel_path = os.path.join(output_folder, f"{asin}_po_data.xlsx")
        with pd.ExcelWriter(excel_path) as writer:
            weekly_agg.to_excel(writer, sheet_name='Weekly Quantity', index=False)
            monthly_trend.to_excel(writer, sheet_name='Monthly Trend', index=False)
            if not forecast_po.empty:
                cols = ['ds','PO_Forecast']
                forecast_po[cols].to_excel(writer, sheet_name='PO Forecast', index=False)

    print(f"Generated PO analysis Excel reports for {len(unique_asins)} ASINs in folder '{output_folder}'.")
    print("Returning po_forecasts_dict for in-memory usage.")
    return po_forecasts_dict

##############################
# Compare PO Orders with Forecasts
##############################

def compare_historical_sales_po(asin, sales_df, po_df, output_folder="po_forecast_comparison"):
    """
    Compare historical sales with PO requested quantities for a given ASIN.
    Overlays historical sales and PO trends, computes correlation, 
    calculates growth percentages and volume insights,
    and provides a basic prediction for future PO quantities.
    Saves overlay graphs and merged data.
    
    Parameters:
      asin: The ASIN to analyze.
      sales_df: DataFrame containing weekly aggregated historical sales with columns ['ds', 'y'].
                - 'ds' is assumed to be a Sunday date (week-end).
      po_df: DataFrame containing historical PO data with columns ['ASIN', 'Order date', 'Requested quantity'].
      output_folder: Directory where results will be saved.
    
    Returns:
      merged_df: DataFrame merging historical sales and PO data.
      correlation: Correlation coefficient between sales and PO requested quantities.
      growth_info: Dictionary containing growth percentages, volume insights, and predicted next week PO quantity.
    """
    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    
    # Ensure numeric conversion for Requested quantity; handle commas and spaces
    po_df = po_df.copy()
    po_df['Requested quantity'] = (
         po_df['Requested quantity']
         .astype(str)
         .str.replace(',', '')
         .str.strip()
         .replace('', '0')
    )
    po_df['Requested quantity'] = pd.to_numeric(po_df['Requested quantity'], errors='coerce').fillna(0)
    
    asin_po = po_df[po_df['ASIN'] == asin].copy()
    asin_po['Order date'] = pd.to_datetime(asin_po['Order date'], errors='coerce')
    asin_po.sort_values('Order date', inplace=True)

    # 3) Create a daily PO DataFrame (raw, no summation)
    # Rename columns to be consistent
    asin_po.rename(columns={'Order date': 'Date', 'Requested quantity': 'Daily_PO_Qty'}, inplace=True)

    # 4) Basic volume stats on the raw/daily POs
    total_po_qty = asin_po['Daily_PO_Qty'].sum()
    avg_po_qty   = asin_po['Daily_PO_Qty'].mean() if len(asin_po) else 0
    max_po_qty   = asin_po['Daily_PO_Qty'].max() if len(asin_po) else 0
    min_po_qty   = asin_po['Daily_PO_Qty'].min() if len(asin_po) else 0

    volume_insights = {
        'Total_PO_Quantity': total_po_qty,
        'Average_PO_Quantity': avg_po_qty,
        'Max_PO_Quantity': max_po_qty,
        'Min_PO_Quantity': min_po_qty
    }

    # 5) Basic linear model (optional) to predict next daily PO
    predicted_next_po_qty = 0
    if len(asin_po) > 1:
        asin_po = asin_po.reset_index(drop=True)
        asin_po['Index'] = np.arange(1, len(asin_po) + 1)
        X = asin_po[['Index']]
        y = asin_po['Daily_PO_Qty']
        model = LinearRegression()
        model.fit(X, y)
        next_index = asin_po['Index'].iloc[-1] + 1
        predicted_next_po_qty = max(model.predict([[next_index]])[0], 0)

    prediction_info = {
        'Predicted_Next_Daily_PO_Quantity': predicted_next_po_qty
    }

    # 6) Merge weekly sales with daily PO if you want a combined DataFrame
    #    This is purely optional. The DS column is Sunday-based in sales_df, daily in PO.
    #    We'll do a left join on date. Many rows won't match exactly unless a daily date
    #    equals a Sunday from your sales.
    merged_df = pd.merge(
        sales_df.rename(columns={'ds': 'Date'}), 
        asin_po[['Date','Daily_PO_Qty']], 
        on='Date', 
        how='outer'
    ).sort_values('Date')
    merged_df['Daily_PO_Qty'] = merged_df['Daily_PO_Qty'].fillna(0)
    merged_df['y'] = merged_df['y'].fillna(0)

    # 7) Plot overlay: we can try to plot daily PO vs. weekly sales on the same axis
    plt.figure(figsize=(12,6))
    # Weekly sales
    plt.plot(
        merged_df['Date'], 
        merged_df['y'], 
        marker='o', 
        label='Weekly Sales (y)'
    )
    # Daily PO
    plt.plot(
        merged_df['Date'], 
        merged_df['Daily_PO_Qty'], 
        marker='x',
        label='Daily PO Qty'
    )
    plt.title(f"Daily POs vs Weekly Sales for ASIN {asin}")
    plt.xlabel("Date")
    plt.ylabel("Quantity")
    plt.legend()
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()

    chart_path = os.path.join(output_folder, f"{asin}_dailyPO_vs_weeklySales.png")
    plt.savefig(chart_path)
    plt.close()
    print(f"Chart saved to {chart_path}")

    # 8) Correlation
    # We'll only check correlation on days that have data in both columns:
    valid_idx = (merged_df['y'] > 0) & (merged_df['Daily_PO_Qty'] > 0)
    if valid_idx.sum() > 1:
        correlation = merged_df.loc[valid_idx, ['y','Daily_PO_Qty']].corr().iloc[0, 1]
        print(f"Correlation between weekly sales and daily PO for ASIN {asin}: {correlation:.2f}")
    else:
        correlation = None
        print(f"Not enough overlapping data for correlation on ASIN {asin}.")

    # 9) Write to Excel: two sheets - "Weekly Sales" and "Daily PO"
    # Also put merged in a third sheet if desired.
    excel_filename = os.path.join(output_folder, f"{asin}_sales_po_comparison.xlsx")
    with pd.ExcelWriter(excel_filename) as writer:
        # Sheet 1: Weekly Sales
        sales_df.to_excel(writer, sheet_name='Weekly Sales', index=False)

        # Sheet 2: Daily PO
        asin_po.to_excel(writer, sheet_name='Daily PO', index=False)

        # Optionally a third sheet with merged data
        merged_df.to_excel(writer, sheet_name='Merged (Optional)', index=False)

        # Another sheet with basic stats or growth info
        stats_df = pd.DataFrame([volume_insights])
        stats_df.to_excel(writer, sheet_name='PO Volume Insights', index=False)

        pred_df = pd.DataFrame([prediction_info])
        pred_df.to_excel(writer, sheet_name='PO Prediction', index=False)

    print(f"Excel saved to {excel_filename}")

    return merged_df, correlation, {
        'volume_insights': volume_insights,
        'prediction_info': prediction_info
    }

##############################
# Forecast Sales with PO Regressor
##############################

def forecast_sales_with_po_regressor(sales_file='weekly_sales_data.xlsx',
                                     po_file='po database.xlsx',
                                     horizon_weeks=16,
                                     amazon_start_date='2025-02-09', #Change dates here
                                     cv_initial='730 days',
                                     cv_period='180 days',
                                     cv_horizon='365 days'):
    """
    Forecast future sales using Prophet while incorporating historical PO orders as a regressor.
    The idea is that the model will learn the relationship between the weekly aggregated PO orders
    and the actual sales. Backtesting (via Prophet’s cross_validation) is run to help assess accuracy.
    
    Parameters:
      sales_file (str): Path to Excel file containing historical sales data. Must have at least:
                        - ds: dates (or convertable to datetime)
                        - y: sales numbers.
      po_file (str): Path to Excel file containing historical PO data. Must include columns:
                     [ASIN, Order date, Requested quantity].
      horizon_weeks (int): Number of future weeks to forecast.
      amazon_start_date (str): A string (or Timestamp) that defines the start date for the future forecast.
      cv_initial (str): Initial training period for cross-validation (e.g., '730 days').
      cv_period (str): Period between CV cutoffs (e.g., '180 days').
      cv_horizon (str): Forecast horizon for CV (e.g., '365 days').
      
    Returns:
      forecast (DataFrame): Prophet forecast for the future (including the PO regressor).
      model (Prophet object): The fitted Prophet model.
      cv_metrics (DataFrame): Backtesting performance metrics (MAE, RMSE, etc.).
    """

    # ----- STEP 1. Load and prepare sales data -----
    sales_df = pd.read_excel(sales_file)
    sales_df.columns = sales_df.columns.str.strip()
    # Assume sales_df has columns 'ds' (dates) and 'y' (sales)
    sales_df['ds'] = pd.to_datetime(sales_df['ds'], errors='coerce')
    sales_df = sales_df.dropna(subset=['ds', 'y'])

    # ----- STEP 2. Load and aggregate PO data by week -----
    po_df = pd.read_excel(po_file)
    po_df.columns = po_df.columns.str.strip()
    # Ensure required columns exist:
    required_cols = {'ASIN', 'Order date', 'Requested quantity'}
    if not required_cols.issubset(po_df.columns):
        raise ValueError(f"PO data must have columns at least: {required_cols}")
    po_df['Order date'] = pd.to_datetime(po_df['Order date'], errors='coerce')
    po_df = po_df.dropna(subset=['ASIN', 'Order date', 'Requested quantity'])
    # Here we aggregate PO orders by week.
    # We choose a weekly period that matches your sales data frequency. For example, if your sales
    # data is aggregated weekly with weeks starting on Sunday, do:
    po_df['Week'] = po_df['Order date'].dt.to_period('W-SUN').apply(lambda r: r.start_time)
    weekly_po = po_df.groupby('Week', as_index=False)['Requested quantity'].sum()
    weekly_po = weekly_po.rename(columns={'Requested quantity': 'Weekly_PO_Qty'})
    
    # ----- STEP 3. Merge PO regressor into sales data -----
    # For each sales date, determine its corresponding week (using the same rule as above).
    sales_df['Week'] = sales_df['ds'].dt.to_period('W-SUN').apply(lambda r: r.start_time)
    # Merge on the "Week" column (left join so that every sales record gets a PO value)
    sales_df = pd.merge(sales_df, weekly_po, on='Week', how='left')
    # If a week has no PO data, fill with 0 (or you might use forward‐fill or another treatment)
    sales_df['Weekly_PO_Qty'] = sales_df['Weekly_PO_Qty'].fillna(0)
    
    # ----- STEP 4. Initialize Prophet with PO regressor -----
    model = Prophet(yearly_seasonality=True,
                    weekly_seasonality=True,
                    daily_seasonality=False,
                    seasonality_mode='multiplicative')
    model.add_regressor('Weekly_PO_Qty')
    
    # Fit the model using the columns ds, y, and Weekly_PO_Qty
    print(f"Training Prophet model on {len(sales_df)} data points with PO regressor...")
    model.fit(sales_df[['ds', 'y', 'Weekly_PO_Qty']])
    
    # ----- STEP 5. Create future dataframe for forecasting -----
    # Convert amazon_start_date to datetime if necessary.
    amazon_start_date = pd.to_datetime(amazon_start_date, errors='coerce')
    if pd.isna(amazon_start_date):
        raise ValueError(f"Invalid amazon_start_date: {amazon_start_date}")
    
    # Build future dataframe: the forecast will start on amazon_start_date and run for horizon_weeks weeks.
    future = model.make_future_dataframe(periods=horizon_weeks, freq='W-SUN')
    
    # For the regressor column in the future, you need to supply a value.
    # One simple approach is to use the historical average of Weekly_PO_Qty.
    future['Weekly_PO_Qty'] = sales_df['Weekly_PO_Qty'].mean()
    
    # ----- STEP 6. Generate forecast -----
    forecast = model.predict(future)
    
    # For output, keep only the forecast dates from the amazon_start_date onward.
    forecast = forecast[forecast['ds'] >= amazon_start_date].copy()
    
    # Create a week label: for example, if you want "W0, W1, ..." starting with amazon_start_date:
    forecast = forecast.sort_values('ds').reset_index(drop=True)
    forecast['Week'] = [f"W{i}" for i in range(len(forecast))]
    forecast['Week_Start_Date'] = forecast['ds'].dt.strftime("%Y-%m-%d")
    
    # Keep only the columns you need; here we show the integrated forecast (for sales) that used the PO data.
    forecast_out = forecast[['Week', 'Week_Start_Date', 'ds', 'yhat', 'Weekly_PO_Qty']].copy()
    forecast_out = forecast_out.rename(columns={'yhat': 'Forecast_Sales'})
    
    # ----- STEP 7. Backtesting / Cross-validation -----
    # (Optional but recommended) Evaluate model performance via Prophet's cross_validation.
    try:
        from prophet.diagnostics import cross_validation, performance_metrics
        print("Running cross-validation to evaluate forecast accuracy...")
        df_cv = cross_validation(model, initial=cv_initial, period=cv_period, horizon=cv_horizon)
        cv_metrics = performance_metrics(df_cv)
        print("Cross-validation metrics:")
        print(cv_metrics[['mae', 'rmse']])
    except Exception as e:
        print(f"Cross-validation failed: {e}")
        cv_metrics = None

    return forecast_out, model, cv_metrics

##############################
#life cycle functions
##############################
def lifecycle_stage_from_slope(slope):
    """
    Lifecycle Stage = 
      - Growth   if slope > 0.1
      - Decline  if slope < -0.1
      - Mature   otherwise
    """
    if slope > 0.1:
        return "Growth"
    elif slope < -0.1:
        return "Decline"
    else:
        return "Mature"

def compute_sales_slope(historical_sales_52):
    """
    Returns the numeric slope from the last 52 weeks of actual sales.
    If len < 2, returns 0.0 for slope.
    """
    if len(historical_sales_52) < 2:
        return 0.0
    x = np.arange(len(historical_sales_52))
    slope, _, _, _, _ = linregress(x, historical_sales_52)
    return slope


def generate_low_coverage_report(consolidated_data, coverage_threshold=1.0, output_file="low_inventory_report.xlsx"):
    
    flagged_rows = []
    
    for asin, df_forecast in consolidated_data.items():
        if 'Inventory Coverage' not in df_forecast.columns or 'Reorder Urgency' not in df_forecast.columns:
            # If these columns are missing for some reason, skip or continue
            continue
        
        # Filter rows
        # e.g. coverage < 1.0 AND Reorder Urgency = 'Urgent'
        condition = (
            (df_forecast['Inventory Coverage'] < coverage_threshold) &
            (df_forecast['Reorder Urgency'].str.lower() == 'urgent')
        )
        flagged = df_forecast[condition].copy()
        if not flagged.empty:
            flagged['ASIN'] = asin
            flagged_rows.append(flagged)
    
    if len(flagged_rows) == 0:
        print(f"No ASINs found with coverage < {coverage_threshold} and Urgent reorder.")
        return
    
    # Concatenate all flagged rows into one DataFrame
    flagged_df = pd.concat(flagged_rows, ignore_index=True)
    
    # Sort by ASIN, then by Week (assuming 'Week' is a column)
    sort_cols = []
    if 'ASIN' in flagged_df.columns:
        sort_cols.append('ASIN')
    if 'Week' in flagged_df.columns:
        sort_cols.append('Week')
    flagged_df = flagged_df.sort_values(sort_cols).reset_index(drop=True)
    
    # Save to Excel (or you can to_csv if you prefer)
    flagged_df.to_excel(output_file, index=False)
    print(f"Low Coverage/High Urgency rows saved to {output_file}. Rows: {len(flagged_df)}")


##############################
# PO Order Data Analysis
##############################
def generate_po_coverage_report(
    merged_df,
    output_folder="po_coverage_analysis",
    output_file="po_coverage_summary.xlsx"
):

    # 1) Ensure the output folder exists
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(output_folder, output_file)

    df = merged_df.copy()
    mapping = {
        'high': 'High Volume Season',
        'low': 'Low Volume Season',
        'medium': 'Regular Season',
        # If your code calls it "stable" or anything else, map it to "Regular Season"
        'stable': 'Regular Season'
    }
    
    # If you named the column differently, adjust as needed:
    if 'seasonal_cluster' not in df.columns:
        df['seasonal_cluster'] = 'regular'
        mapping = {'regular': 'Regular Season'}

    df['SeasonLabel'] = df['seasonal_cluster'].map(mapping).fillna('Regular Season')

    group_cols = ['ASIN', 'SeasonLabel']
    agg_df = (df
              .groupby(group_cols, as_index=False)
              .agg({
                  'Weekly_PO_Qty': 'mean',
                  'coverage': 'mean',
                  'ds': [ 'min', 'max' ]
              }))

    # Flatten multi-level columns
    agg_df.columns = [
        'ASIN', 'SeasonLabel',
        'avg_po', 'avg_cov',
        'start_date', 'end_date'
    ]

    asin_titles = df[['ASIN', 'Model Name']].drop_duplicates()
    agg_df = pd.merge(agg_df, asin_titles, on='ASIN', how='left')

    pivoted = agg_df.pivot(
        index=['ASIN','Model Name'],
        columns='SeasonLabel',
        values=['avg_po','avg_cov','start_date','end_date']
    )

    # pivoted will have a multi-level column structure.
    # We'll flatten them and rename to a simpler pattern.
    pivoted.columns = [
        f"{col[0]} ({col[1]})" for col in pivoted.columns
    ]

    pivoted.reset_index(inplace=True)

    rename_map = {
        'avg_po (Regular Season)': 'Average PO order (regular season)',
        'avg_cov (Regular Season)': 'average Inventory coverage (regular season)',
        'start_date (Regular Season)': 'start_date_regular',
        'end_date (Regular Season)': 'end_date_regular',

        'avg_po (High Volume Season)': 'Average PO order (High Volume season)',
        'avg_cov (High Volume Season)': 'average Inventory coverage (High volume season)',
        'start_date (High Volume Season)': 'start_date_high',
        'end_date (High Volume Season)': 'end_date_high',

        'avg_po (Low Volume Season)': 'Average PO order (low volume season)',
        'avg_cov (Low Volume Season)': 'average Inventory coverage (low volume season)',
        'start_date (Low Volume Season)': 'start_date_low',
        'end_date (Low Volume Season)': 'end_date_low',
    }

    # Some columns may be missing if an ASIN never had a "high" or "low" cluster, so we do a safe rename:
    pivoted.rename(columns=rename_map, inplace=True)

    # 7) Create final columns for date range (regular, high, low).
    #    We'll combine the start_date_X and end_date_X into one string:
    def make_date_range(row, start_col, end_col):
        s = row.get(start_col, pd.NaT)
        e = row.get(end_col, pd.NaT)
        if pd.isna(s) or pd.isna(e):
            return ""
        return f"{pd.to_datetime(s).date()} to {pd.to_datetime(e).date()}"

    for label in ['regular','high','low']:
        start_col = f"start_date_{label}"
        end_col   = f"end_date_{label}"
        range_col = f"Date range  ({label} season)"
        pivoted[range_col] = pivoted.apply(
            lambda row: make_date_range(row, start_col, end_col),
            axis=1
        )

    # 8) Drop the start_date_X/end_date_X helper columns if you don’t want them in final output
    drop_cols = [
        'start_date_regular','end_date_regular',
        'start_date_high','end_date_high',
        'start_date_low','end_date_low'
    ]
    for c in drop_cols:
        if c in pivoted.columns:
            pivoted.drop(columns=[c], inplace=True)

    # 9) Final column order
    final_columns = [
        'ASIN',
        'Model Name',

        'Average PO order (regular season)',
        'average Inventory coverage (regular season)',
        'Date range  (regular season)',

        'Average PO order (High Volume season)',
        'average Inventory coverage (High volume season)',
        'Date range  (high volume season)',

        'Average PO order (low volume season)',
        'average Inventory coverage (low volume season)',
        'Date range  (low volume season)',
    ]
    for c in final_columns:
        if c not in pivoted.columns:
            pivoted[c] = ""

    pivoted = pivoted[final_columns]

    # 10) Optionally round numeric columns
    numeric_cols = [
        'Average PO order (regular season)',
        'average Inventory coverage (regular season)',
        'Average PO order (High Volume season)',
        'average Inventory coverage (High volume season)',
        'Average PO order (low volume season)',
        'average Inventory coverage (low volume season)'
    ]
    for col in numeric_cols:
        if col in pivoted.columns:
            pivoted[col] = pivoted[col].astype(float).round(2)

    # 11) Save to Excel
    pivoted.to_excel(output_path, index=False)
    print(f"PO coverage analysis saved to: {output_path}")

def merge_historical_data(sales_file, inventory_file, po_file):
    """
    1) Load weekly sales data (via load_weekly_sales_data), which yields columns:
       'ds','asin','product title','y' (renamed to 'Weekly_Sales'), etc.
    2) Load inventory data (via load_inventory_data), which yields:
       'ds','ASIN','Starting_Inventory',...
    3) Load PO data from Excel, aggregate to weekly.

    Merge them so the final DataFrame has:
      - 'ASIN'
      - 'ds'
      - 'Weekly_Sales'
      - 'Starting_Inventory'
      - 'Weekly_PO_Qty'
      - 'coverage' (computed as Starting_Inventory / Weekly_Sales)
      etc.
    """
    # 1) Load weekly sales data
    sales_df = load_weekly_sales_data(sales_file)
    sales_df = sales_df.rename(columns={
        'asin': 'ASIN',
        'y': 'Weekly_Sales',
        'product title': 'Model Name'
    })

    # 2) Load inventory data
    inv_df = load_inventory_data(inventory_file)

    # 3) Load and process PO data
    po_df = pd.read_excel(po_file)
    po_df.columns = po_df.columns.str.strip()
    po_df['Order date'] = pd.to_datetime(po_df['Order date'], errors='coerce')
    po_df = po_df.dropna(subset=['ASIN', 'Order date', 'Requested quantity'])
    po_df['Requested quantity'] = (
        po_df['Requested quantity']
        .astype(str)
        .str.replace(',', '')
        .str.strip()
        .replace('', '0')
    )
    po_df['Requested quantity'] = pd.to_numeric(po_df['Requested quantity'], errors='coerce').fillna(0)

    # Aggregate PO data weekly (Sunday-end)
    po_df['ds'] = po_df['Order date'].dt.to_period('W-SUN').apply(lambda r: r.end_time)
    po_weekly = (
        po_df.groupby(['ASIN', 'ds'], as_index=False)['Requested quantity']
        .sum()
        .rename(columns={'Requested quantity': 'Weekly_PO_Qty'})
    )

    # Merge sales + inventory on (ASIN, ds)
    merged = pd.merge(
        sales_df,
        inv_df[['ASIN','ds','Starting_Inventory']],
        on=['ASIN','ds'],
        how='left'
    )

    # Merge with weekly PO data
    merged = pd.merge(merged, po_weekly, on=['ASIN','ds'], how='left')
    merged['Weekly_PO_Qty'] = merged['Weekly_PO_Qty'].fillna(0)

    # Compute coverage
    merged['coverage'] = merged.apply(
        lambda row: row['Starting_Inventory'] / (row['Weekly_Sales'] + 1e-9)
                    if pd.notna(row['Starting_Inventory']) and pd.notna(row['Weekly_Sales'])
                    else 0,
        axis=1
    )

    return merged

def train_amazon_po_model(merged_df):
    """
    Train a 2-stage model:
      - logistic -> did Amazon place an order? (Weekly_PO_Qty > 0)
      - linear regression -> how big was the order if placed
    Using coverage, holiday lead-time features, etc.
    """
    df = merged_df.copy()

    # Classification label
    df['Order_Placed'] = (df['Weekly_PO_Qty'] > 0).astype(int)
    classes = df['Order_Placed'].unique()
    if len(classes) < 2:
        print("Only one class (no PO placed). Returning dummy model.")
        return None, None

    # If you have holiday lead time features: 'days_to_next_holiday','is_within_14_days_of_holiday'
    # Otherwise, just coverage
    features = []
    if 'days_to_next_holiday' in df.columns and 'is_within_14_days_of_holiday' in df.columns:
        features = ['coverage', 'days_to_next_holiday', 'is_within_14_days_of_holiday']
        # convert bool
        df['is_within_14_days_of_holiday'] = df['is_within_14_days_of_holiday'].astype(int)
    else:
        features = ['coverage']

    X_class = df[features].copy()
    X_class = X_class.fillna(0)
    y_class = df['Order_Placed']

    print("Sample merged data:\n", df[['ASIN','ds','Weekly_PO_Qty','coverage']].head(30))

    # Logistic
    log_reg = LogisticRegression()
    log_reg.fit(X_class, y_class)

    # For regression, only rows where an order was placed
    df_orders = df[df['Order_Placed'] == 1].copy()
    if df_orders.empty:
        print("Warning: No positive PO data found, returning logistic model + None for regression.")
        return log_reg, None

    X_reg = df_orders[features].copy()
    y_reg = df_orders['Weekly_PO_Qty']

    lin_reg = LinearRegression()
    lin_reg.fit(X_reg, y_reg)

    return log_reg, lin_reg

def forecast_amazon_po_orders(future_df, logistic_model, regression_model, initial_inventory):
    """
    Stepwise approach for each future date:
      coverage = current_inventory / forecasted sales
      logistic_model -> place_order?
      if place_order => regression_model for PO qty
      update inventory
    """
    if logistic_model is None:
        print("No logistic model was trained (only one class). Returning an empty PO forecast.")
        return pd.DataFrame({
            'ds': future_df['ds'],
            'Predicted_PO_Qty': 0,
            'Coverage_Sim': 999,
            'Inventory_After_PO': initial_inventory,
            'Order_Probability': 0
        })

    pred_rows = []
    current_inventory = initial_inventory

    for i in range(len(future_df)):
        row = future_df.iloc[i]
        ds_date = row['ds']
        sales_fc = row.get('MyForecast', 0)

        # coverage
        coverage = current_inventory / (sales_fc + 1e-9)

        # logistic features (just coverage if no holiday lead-time)
        # if you have holiday lead-time in future_df, do:
        days_next = row.get('days_to_next_holiday', 9999)
        is_holiday_14 = int(row.get('is_within_14_days_of_holiday', False))

        feat = [coverage, days_next, is_holiday_14] \
            if 'days_to_next_holiday' in future_df.columns else [coverage]
        
        # logistic predict
        order_prob = logistic_model.predict_proba([feat])[0,1] if logistic_model else 0
        place_order = (order_prob >= 0.5)

        if place_order and regression_model is not None:
            po_qty = regression_model.predict([feat])[0]
            po_qty = max(po_qty, 0)
        else:
            po_qty = 0

        # update inventory
        current_inventory = current_inventory - sales_fc + po_qty
        if current_inventory < 0:
            current_inventory = 0

        pred_rows.append({
            'ds': ds_date,
            'Predicted_PO_Qty': round(po_qty),
            'Coverage_Sim': round(coverage, 2),
            'Inventory_After_PO': round(current_inventory, 2),
            'Order_Probability': round(order_prob, 2)
        })

    return pd.DataFrame(pred_rows)

def apply_inventory_constraints(forecast_df, runrate_inventory):
    df = forecast_df.copy()
    starting_inv = runrate_inventory
    influences = []
    adj_forecasts = []

    for i in range(len(df)):
        row = df.iloc[i]
        predicted_sales = row.get('MyForecast', np.nan)

        # If we can't get a numeric forecast, skip or treat as zero
        if pd.isna(predicted_sales):
            influences.append(0)
            adj_forecasts.append(0)
            continue  # or handle differently

        coverage = starting_inv / (predicted_sales + 1e-9)

        # Example logic: if coverage < 0.8 => reduce forecast 10%, etc.
        # Simplified example:
        if coverage < 0.8:
            adj_sales = predicted_sales * 0.90
        else:
            adj_sales = predicted_sales

        # If either predicted_sales or adj_sales is NaN, handle it
        if pd.isna(adj_sales):
            adj_sales = 0
        if pd.isna(predicted_sales):
            predicted_sales = 0

        influences.append(round(adj_sales - predicted_sales))  # Now safe
        adj_forecasts.append(round(adj_sales))

        # reduce inventory by the adjusted sales
        starting_inv = max(0, starting_inv - adj_sales)

    df['AdjustedForecast'] = adj_forecasts
    df['inventory_influence'] = influences
    return df

def generate_future_forecast(training_data, periods=16):
    """
    Fit Prophet on the training data (which must have columns 'ds' and 'y'),
    then generate a future DataFrame with exactly 'periods' rows.
    We set include_history=False to return only forecast periods.
    """
    model = Prophet(yearly_seasonality=True, weekly_seasonality=True)
    model.fit(training_data[['ds', 'y']])
    # include_history=False ensures only future dates are returned
    future = model.make_future_dataframe(periods=periods, freq='W-SUN', include_history=False)
    forecast = model.predict(future)
    forecast.rename(columns={'yhat': 'MyForecast'}, inplace=True)
    return forecast

def generate_sales_forecast_for_asin(asin_df, periods=16):
    """
    Fit Prophet on the given asin_df and produce a forecast for 'periods' weeks.
    If Prophet fails (e.g., insufficient data variance), we return an empty DataFrame.
    """
    # 1) Prepare training data
    train_df = asin_df[['ds', 'Weekly_Sales']].dropna().rename(columns={'Weekly_Sales': 'y'})
    
    # Quick check: if too few points or no variance, skip
    if len(train_df) < 2 or train_df['y'].nunique() < 2:
        print("  Not enough valid data or no variance, skipping forecast.")
        return pd.DataFrame(columns=['ds','MyForecast'])
    
    # 2) Create and fit model in a try-except
    model = Prophet(yearly_seasonality=True, weekly_seasonality=True)
    try:
        model.fit(train_df[['ds', 'y']])
    except RuntimeError as e:
        # If Prophet fails, log and return empty
        print("  Prophet model fitting failed for this ASIN. Skipping.")
        return pd.DataFrame(columns=['ds','MyForecast'])
    
    # 3) Generate future dates & forecast
    future = model.make_future_dataframe(periods=periods, freq='W-SUN', include_history=False)
    forecast = model.predict(future)
    
    # 4) Rename yhat => MyForecast, clip negative
    forecast.rename(columns={'yhat': 'MyForecast'}, inplace=True)
    forecast['MyForecast'] = forecast['MyForecast'].clip(lower=0).round()
    
    return forecast[['ds','MyForecast']]

def main_po_forecast_pipeline():
    # 1) Merge historical data (sales, inventory, PO)
    merged_df = merge_historical_data(
        sales_file="weekly_sales_data.xlsx",
        inventory_file="inventory_data.xlsx",
        po_file="po database.xlsx"
    )

    # 2) Get a list of unique ASINs
    asins = merged_df['ASIN'].dropna().unique()
    if len(asins) == 0:
        print("No valid ASINs found in merged data. Exiting.")
        return

    # Load PO data separately for checking if an ASIN has any PO orders
    po_df = pd.read_excel("po database.xlsx")
    po_df.columns = po_df.columns.str.strip()

    # 3) Prepare an Excel writer to save all forecasts in one file using a context manager
    output_file = "All_ASIN_PO_Forecasts.xlsx"
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # 4) Also store a DataFrame with historical coverage for all ASINs in one sheet
        coverage_df = merged_df[['ds', 'ASIN', 'coverage']].dropna()

        # 5) Loop over each ASIN individually
        for asin in asins:
            # Skip ASIN if no PO orders exist
            if not has_po_orders(asin, po_df):
                print(f"Skipping ASIN {asin} since there are no PO orders.")
                continue

            print(f"\n=== Processing ASIN: {asin} ===")
            # Subset the merged data for this ASIN
            asin_df = merged_df[merged_df['ASIN'] == asin].copy()
            if asin_df.empty:
                print(f"  ASIN {asin} is empty after filtering. Skipping.")
                continue

            # (Optional) Add holiday lead time features to historical data
            holidays = get_shifted_holidays()
            if 'ds' in asin_df.columns:
                asin_df = add_holiday_lead_time_features(asin_df, holidays)

            # 6) Train logistic + regression model for this ASIN
            logistic_model, regression_model = train_amazon_po_model(asin_df)

            # Check if there are at least 2 valid 'Weekly_Sales' data points
            sales_data_count = len(asin_df.dropna(subset=['Weekly_Sales']))
            if sales_data_count < 2:
                print(f"ASIN {asin} has only {sales_data_count} data points for Prophet. Creating dummy forecast.")
                future_forecast_df = pd.DataFrame(columns=['ds', 'MyForecast'])
            else:
                # Wrap the call to Prophet in try-except to skip if initialization fails
                try:
                    future_forecast_df = generate_sales_forecast_for_asin(asin_df, periods=16)
                except RuntimeError as e:
                    print(f"Prophet model fitting failed for ASIN {asin}. Skipping. Error: {e}")
                    continue

            # 7) Branch based on whether a logistic model was trained
            if logistic_model is None:
                print(f"  No logistic model for ASIN {asin}. Creating dummy PO forecast.")
                po_prediction = forecast_amazon_po_orders(
                    future_forecast_df, None, None, initial_inventory=1000
                )
                final_result = future_forecast_df.merge(
                    po_prediction[['ds', 'Predicted_PO_Qty', 'Coverage_Sim', 'Inventory_After_PO', 'Order_Probability']],
                    on='ds', how='left'
                )
            else:
                # (Optional) Add holiday lead time features to the future forecast
                if 'ds' in future_forecast_df.columns:
                    future_forecast_df = add_holiday_lead_time_features(future_forecast_df, holidays)

                # 8) Forecast Amazon PO Orders using the trained models
                po_prediction = forecast_amazon_po_orders(
                    future_forecast_df, logistic_model, regression_model, initial_inventory=1000
                )
                # Merge PO predictions with the future forecast
                final_result = future_forecast_df.merge(
                    po_prediction[['ds', 'Predicted_PO_Qty', 'Coverage_Sim', 'Inventory_After_PO', 'Order_Probability']],
                    on='ds', how='left'
                )

            # 9) Optionally apply inventory constraints
            final_result = apply_inventory_constraints(final_result, 1000)

            # 10) Add ASIN and Model columns (detected from data)
            final_result['ASIN'] = asin
            final_result['Model'] = "LogReg + Regr (PO Forecast)"

            # 11) Keep only the desired columns
            columns_to_keep = [
                'ds', 
                'ASIN',
                'Model',
                'MyForecast',         
                'Predicted_PO_Qty',
                'Coverage_Sim',
                'Inventory_After_PO',
                'Order_Probability'
            ]
            for c in columns_to_keep:
                if c not in final_result.columns:
                    final_result[c] = np.nan
            final_result = final_result[columns_to_keep]

            # 12) Write this ASIN's forecast to a separate sheet
            sheet_name = str(asin)[:31]  # limit to 31 characters
            final_result.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  Forecast for ASIN {asin} saved to sheet '{sheet_name}'.")

        # 13) Also write historical coverage for all ASINs to a separate sheet
        coverage_df.to_excel(writer, sheet_name="HistoricalCoverage", index=False)

    # When the 'with' block ends, the file is automatically saved and closed.
    print(f"\nAll forecasts saved to '{output_file}'")
    print("PO Forecast pipeline complete.")

def has_po_orders(asin, po_df):
    """
    Returns True if the given ASIN has any PO orders with a positive 'Requested quantity'.
    """
    asin_po_df = po_df[po_df['ASIN'] == asin]
    if asin_po_df.empty:
        return False
    try:
        total = asin_po_df['Requested quantity'].astype(float).sum()
        return total > 0
    except Exception as e:
        return False
    
def dynamic_prophet_setup(ts_data):
    # Determine the number of weeks of data
    num_weeks = ts_data['ds'].nunique()
    if num_weeks < 52:
        # For short series, disable yearly seasonality
        model = Prophet(
            yearly_seasonality=False,
            weekly_seasonality=True,
            seasonality_mode='multiplicative'
        )
        # Optionally, add a custom long-term seasonality if needed:
        # model.add_seasonality(name='custom', period=num_weeks, fourier_order=3)
    else:
        # For series with over a year of data, use the default yearly seasonality
        model = Prophet(
            yearly_seasonality=True,
            weekly_seasonality=True,
            seasonality_mode='multiplicative'
        )
    return model
    

##############################
# Main
##############################

def main():
    # ----------------------------------------------------------------------------
    # 1) Load existing parameter histories, runrate data, sales data, config
    # ----------------------------------------------------------------------------
    load_param_histories()

    runrate_file = "Runrate.xlsx"  
    runrate_inventory = load_runrate_inventory(runrate_file) 
    sales_file = 'weekly_sales_data.xlsx'
    forecasts_folder = 'forecasts_folder'
    asins_to_forecast_file = 'ASINs to Forecast.xlsx'
    horizon = 16
    output_file = 'consolidated_forecast_WITH_PO.xlsx'

    # Load weekly sales data
    data = load_weekly_sales_data(sales_file)
    valid_data = data[data['asin'].notna() & (data['asin'] != '#N/A')]
    missing_asin_data = data[data['asin'].isna() | (data['asin'] == '#N/A')]

    if not missing_asin_data.empty:
        print("The following entries have no ASIN and will be noted in the forecast file:")
        print(missing_asin_data[['product title', 'week', 'year', 'y']].to_string())

    # Load the ASINs to forecast
    asins_to_forecast = load_asins_to_forecast(asins_to_forecast_file)
    print(f"ASINs to forecast: {asins_to_forecast}")

    # Filter valid_data to only ASINs in asins_to_forecast
    asin_list = valid_data['asin'].unique()
    asin_list = [asin for asin in asin_list if asin in asins_to_forecast]

    # Dictionary to store final forecasts (WITH PO)
    consolidated_forecasts_with_po = {}

    # Get holiday data
    holidays = get_shifted_holidays()

    # Optional cross-validation on first ASIN
    if len(asin_list) > 0:
        test_asin = asin_list[0]
        test_ts_data = prepare_time_series_with_lags(valid_data, test_asin, lag_weeks=1)
        if not test_ts_data.empty and len(test_ts_data.dropna()) >= 2:
            print(f"Performing cross-validation on ASIN {test_asin} Prophet model (for test only).")
            cross_validate_prophet_model(test_ts_data, initial='180 days', period='90 days', horizon='90 days')
        else:
            print(f"Not enough data for {test_asin} to perform cross-validation test.")

    # Create output folders
    insufficient_data_folder = "Insufficient data"
    with_po_folder = "Sufficient data WITH_PO"
    os.makedirs(insufficient_data_folder, exist_ok=True)
    os.makedirs(with_po_folder, exist_ok=True)

    global PARAM_COUNTER

    # ----------------------------------------------------------------------------
    # 2) Load and prepare global PO data (for merging per ASIN)
    # ----------------------------------------------------------------------------
    po_file = "po database.xlsx"
    inventory_file = "inventory_data.xlsx"  # Pending file
    # ^ Make sure you actually need 'po_file' duplicated below or if it's just the same variable
    po_file = "po database.xlsx"
    po_df = pd.read_excel(po_file)
    po_df.columns = po_df.columns.str.strip()
    po_df['Order date'] = pd.to_datetime(po_df['Order date'], errors='coerce')
    po_df = po_df.dropna(subset=['ASIN', 'Order date', 'Requested quantity'])
    po_df['Requested quantity'] = (
        po_df['Requested quantity']
        .astype(str)
        .str.replace(',', '')
        .str.strip()
        .replace('', '0')
    )
    po_df['Requested quantity'] = pd.to_numeric(po_df['Requested quantity'], errors='coerce').fillna(0)

    merged_df = merge_historical_data(sales_file, inventory_file, po_file)

    merged_df = detect_seasonal_periods(merged_df)  
    # Now call generate_po_coverage_report:
    generate_po_coverage_report(
        merged_df,
        output_folder="po_coverage_analysis",
        output_file="po_coverage_summary.xlsx"
    )
    # ------------------------------------------------------------------------

    # ----------------------------------------------------------------------------
    # 3) For each ASIN, generate forecast WITH PO integration
    # ----------------------------------------------------------------------------
    for asin in asin_list:
        if pd.isna(asin) or asin == '#N/A':
            print(f"Skipping invalid ASIN: {asin}")
            continue

        product_title = valid_data[valid_data['asin'] == asin]['product title'].iloc[0]
        print(f"\nProcessing ASIN: {asin} - {product_title}")

        # Load Amazon forecasts from folder
        forecast_data = load_amazon_forecasts_from_folder(forecasts_folder, asin)
        if not forecast_data:
            print(f"No forecast data found for ASIN {asin}, skipping.")
            continue

        # -------------------------
        # 3.1) Prepare Base Time Series
        # -------------------------
        ts_data_base = prepare_time_series_with_lags(valid_data, asin, lag_weeks=1)
        print(f"Time series base data for ASIN {asin} prepared. Size: {len(ts_data_base)}")

        # -------------------------
        # 3.2) Merge PO Data
        # -------------------------
        asin_po_df = po_df[po_df['ASIN'] == asin].copy()
        asin_po_df['OrderWeek'] = asin_po_df['Order date'].dt.to_period('W-SUN').apply(lambda p: p.start_time)
        weekly_po = asin_po_df.groupby('OrderWeek', as_index=False)['Requested quantity'].sum()
        weekly_po.rename(columns={'Requested quantity': 'Weekly_PO_Qty'}, inplace=True)

        ts_data_with_po = ts_data_base.copy()
        ts_data_with_po['OrderWeek'] = ts_data_with_po['ds'].dt.to_period('W-SUN').apply(lambda p: p.start_time)
        ts_data_with_po = pd.merge(ts_data_with_po, weekly_po, on='OrderWeek', how='left')
        ts_data_with_po['Weekly_PO_Qty'] = ts_data_with_po['Weekly_PO_Qty'].fillna(0)
        print(f"Time series data (WITH PO) for ASIN {asin} prepared. Size: {len(ts_data_with_po)}")

        # 4) If insufficient data (<10 => fallback), else normal pipeline
        non_nan_count_with_po = len(ts_data_with_po.dropna())
        if non_nan_count_with_po < 10:
            print(f"ASIN {asin} has only {non_nan_count_with_po} data points (WITH PO). Using fallback approach.")
            if not ts_data_with_po.empty:
                last_y_value_po = ts_data_with_po['y'].iloc[-1]
                last_date_po = ts_data_with_po['ds'].iloc[-1]
            else:
                last_y_value_po = 0
                last_date_po = pd.to_datetime("today")

            fallback_dates_with_po = pd.date_range(
                start=last_date_po + pd.Timedelta(weeks=1),
                periods=horizon,
                freq='W'
            )
            fallback_sarima_vals_with_po = [last_y_value_po]*horizon

            fallback_df_with_po = pd.DataFrame({
                'ds': fallback_dates_with_po,
                'Fallback_SARIMA': fallback_sarima_vals_with_po
            })

            amz_mean_key_po = None
            for ftype in forecast_data.keys():
                if 'mean' in ftype.lower():
                    amz_mean_key_po = ftype
                    break

            if amz_mean_key_po is None:
                print("Warning: No 'Mean' forecast found in Amazon data (WITH PO). Using 0.")
                amazon_mean_with_po = [0]*horizon
            else:
                raw_arr_with_po = forecast_data[amz_mean_key_po]
                if len(raw_arr_with_po) >= horizon:
                    amazon_mean_with_po = raw_arr_with_po[:horizon]
                else:
                    pad_needed_with_po = horizon - len(raw_arr_with_po)
                    amazon_mean_with_po = list(raw_arr_with_po) + [raw_arr_with_po[-1]]*pad_needed_with_po

            fallback_df_with_po['Amazon Mean Forecast'] = amazon_mean_with_po

            # Weighted fallback (simple example)
            weight_amz_po = 0.9
            weight_sarima_po = 0.1
            final_vals_po = []
            for i in range(horizon):
                mean_val_po = fallback_df_with_po['Amazon Mean Forecast'].iloc[i]
                sarima_val_po = fallback_df_with_po['Fallback_SARIMA'].iloc[i]
                combined_po = (weight_amz_po * mean_val_po) + (weight_sarima_po * sarima_val_po)
                final_vals_po.append(int(round(max(combined_po, 0))))

            fallback_df_with_po['MyForecast'] = final_vals_po
            fallback_df_with_po['ASIN'] = asin
            fallback_df_with_po['Product Title'] = product_title

            # Plot
            plt.figure(figsize=(10,6))
            plt.plot(fallback_df_with_po['ds'], fallback_df_with_po['MyForecast'], marker='o', label='MyForecast')
            plt.title(f'Fallback (WITH PO) Forecast for ASIN {asin}')
            plt.xlabel('Date')
            plt.ylabel('Requested Quantity')
            plt.legend()
            plt.grid()
            plt.xticks(rotation=45)
            plt.tight_layout()
            fallback_chart_with_po = os.path.join(with_po_folder, f"{asin}_fallback_forecast_WITH_PO.png")
            plt.savefig(fallback_chart_with_po)
            plt.close()
            print(f"Fallback (WITH PO) chart saved to {fallback_chart_with_po}")

            def iso_week_label(d):
                if pd.isna(d):
                    return "W?"
                return f"W{d.isocalendar().week + 1}"

            fallback_df_with_po.rename(columns={'ds': 'Week_Start_Date'}, inplace=True)
            fallback_df_with_po['Week'] = pd.to_datetime(fallback_df_with_po['Week_Start_Date']).dt.isocalendar().week.apply(
                lambda w: 'W' + str(w)
            )
            fallback_df_with_po['is_holiday_week'] = False

            desired_columns_po = [
                'Week','Week_Start_Date','ASIN','MyForecast','Amazon Mean Forecast',
                'Amazon P70 Forecast','Amazon P80 Forecast','Amazon P90 Forecast',
                'Product Title','is_holiday_week'
            ]
            for col in ['Amazon P70 Forecast','Amazon P80 Forecast','Amazon P90 Forecast']:
                if col not in fallback_df_with_po.columns:
                    fallback_df_with_po[col] = 0
            fallback_df_with_po = fallback_df_with_po[desired_columns_po]

            consolidated_forecasts_with_po[asin] = fallback_df_with_po
            fallback_with_po_path = os.path.join(with_po_folder, f"{asin}_fallback_WITH_PO.xlsx")
            fallback_df_with_po.to_excel(fallback_with_po_path, index=False)
            print(f"Saved fallback (WITH PO) forecast for ASIN {asin} to {fallback_with_po_path}")

        else:
            # 5) Normal Pipeline
            model_with_po, model_type_with_po = choose_forecast_model(
                ts_data_with_po,
                threshold=FALLBACK_THRESHOLD,
                holidays=holidays
            )

            # === SARIMA (WITH PO)
            if model_type_with_po == "SARIMA":
                n_with_po = len(ts_data_with_po)
                split_with_po = int(n_with_po * 0.8)
                train_sarima_with_po = ts_data_with_po.iloc[:split_with_po]
                test_sarima_with_po = ts_data_with_po.iloc[split_with_po:]
                
                # --- Generate exogenous data for test forecasting ---
                exog_test_with_po = create_holiday_regressors(test_sarima_with_po, holidays)
                if 'Weekly_PO_Qty' in test_sarima_with_po.columns:
                    exog_test_with_po['Weekly_PO_Qty'] = test_sarima_with_po['Weekly_PO_Qty'].values

                if model_with_po is None:
                    print(f"[WITH PO][SARIMA] Model is None for {asin}, skipping.")
                    continue

                try:
                    # Fit SARIMA model with parameter optimization.
                    best_sarima_model_with_po, best_params_with_po = fit_sarima_model(
                        data=ts_data_with_po,
                        holidays=holidays,
                        seasonal_period=52,
                        asin=asin
                    )
                    if best_sarima_model_with_po is None:
                        print(f"[WITH PO][SARIMA] Fitting failed for {asin}. Using fallback: Amazon mean forecast.")
                        raise ValueError("No suitable SARIMA model found.")
                    
                    # --- Evaluate on the test set ---
                    steps_test = len(test_sarima_with_po)
                    sarima_test_forecast_df_with_po = sarima_forecast(
                        model_fit=best_sarima_model_with_po,
                        steps=steps_test,
                        last_date=train_sarima_with_po['ds'].iloc[-1],
                        exog=exog_test_with_po
                    )
                    sarima_preds_with_po = sarima_test_forecast_df_with_po['MyForecast'].values
                    sarima_mae_with_po = mean_absolute_error(test_sarima_with_po['y'], sarima_preds_with_po)
                    sarima_rmse_with_po = sqrt(mean_squared_error(test_sarima_with_po['y'], sarima_preds_with_po))
                    print(f"[WITH PO][SARIMA] Test MAE={sarima_mae_with_po:.4f}, RMSE={sarima_rmse_with_po:.4f} for ASIN {asin}")
                    
                    update_param_history(
                        history_dict=sarima_param_history,
                        asin=asin,
                        param_tuple=best_params_with_po,
                        rmse=sarima_rmse_with_po,
                        mae=sarima_mae_with_po
                    )
                    
                    # --- Generate exogenous data for final forecast ---
                    last_date_full_with_po = ts_data_with_po['ds'].iloc[-1]
                    exog_future_with_po = generate_future_exog(holidays, steps=horizon, last_date=last_date_full_with_po)
                    exog_future_with_po['Weekly_PO_Qty'] = ts_data_with_po['Weekly_PO_Qty'].mean()
                    
                    final_forecast_df_with_po = sarima_forecast(
                        model_fit=best_sarima_model_with_po,
                        steps=horizon,
                        last_date=train_sarima_with_po['ds'].iloc[-1],
                        exog=exog_future_with_po
                    )
                    if final_forecast_df_with_po.empty:
                        print(f"[WITH PO][SARIMA] Final forecast empty for {asin}. Using fallback: Amazon mean forecast.")
                        raise ValueError("Final forecast empty.")
                
                except Exception as e:
                    print(f"[WITH PO][SARIMA] Error for ASIN {asin}: {e}. Using fallback: Amazon mean forecast.")
                    # Fallback: use Amazon Mean forecast if available; otherwise, repeat the last observed actual value.
                    if "Mean" in forecast_data:
                        amazon_mean_series = forecast_data["Mean"]
                    else:
                        last_value = ts_data_with_po['y'].iloc[-1] if len(ts_data_with_po) > 0 else 0
                        amazon_mean_series = np.array([last_value] * horizon)
                    # Pad series if needed
                    if len(amazon_mean_series) < horizon:
                        amazon_mean_series = np.pad(amazon_mean_series, (0, horizon - len(amazon_mean_series)),
                                                    'constant', constant_values=amazon_mean_series[-1])
                    fallback_dates_with_po = pd.date_range(start=ts_data_with_po['ds'].iloc[-1] + pd.Timedelta(weeks=1),
                                                        periods=horizon, freq='W-SUN')
                    # IMPORTANT: Create a DataFrame with 'ds' and 'MyForecast' columns.
                    final_forecast_df_with_po = pd.DataFrame({
                        'ds': fallback_dates_with_po,
                        'MyForecast': amazon_mean_series
                    })
                
                # --- Prepare final comparison DataFrame ---
                comparison_with_po = final_forecast_df_with_po.copy()
                comparison_with_po['ASIN'] = asin
                comparison_with_po['Product Title'] = product_title
                comparison_with_po = comparison_with_po.merge(
                    ts_data_with_po[['ds', 'y']], on='ds', how='left'
                )
                
                # --- Blend with Amazon forecasts using RMSE-based weights (if available) ---
                if forecast_data:
                    for ftype, values in forecast_data.items():
                        horizon_vals = values[:horizon] if len(values) >= horizon else values
                        if len(horizon_vals) < horizon and len(horizon_vals) > 0:
                            horizon_vals = np.pad(horizon_vals, (0, horizon - len(horizon_vals)),
                                                'constant', constant_values=horizon_vals[-1])
                        elif len(horizon_vals) == 0:
                            horizon_vals = np.zeros(horizon, dtype=int)
                        ftype_lower = ftype.lower()
                        if 'mean' in ftype_lower:
                            comparison_with_po['Amazon Mean Forecast'] = horizon_vals
                        elif 'p70' in ftype_lower:
                            comparison_with_po['Amazon P70 Forecast'] = horizon_vals
                        elif 'p80' in ftype_lower:
                            comparison_with_po['Amazon P80 Forecast'] = horizon_vals
                        elif 'p90' in ftype_lower:
                            comparison_with_po['Amazon P90 Forecast'] = horizon_vals
                        else:
                            print(f"Warning: Unrecognized forecast type '{ftype}'. Skipping.")
                    
                    # If enough overlapping actuals exist, compute dynamic weights.
                    window_df = comparison_with_po.dropna(subset=['y'])
                    if len(window_df) >= horizon:
                        rmse_dict = {}
                        for col in ['Amazon Mean Forecast', 'Amazon P70 Forecast', 'Amazon P80 Forecast', 'Amazon P90 Forecast']:
                            if col in comparison_with_po.columns:
                                rmse_dict[col] = compute_rmse(window_df['y'], window_df[col])
                        inv_rmse = {col: (1.0 / val if val >= 1e-9 else 1e9) for col, val in rmse_dict.items()}
                        sum_inv = sum(inv_rmse.values())
                        weights_dict = {col: inv_rmse[col] / sum_inv for col in rmse_dict}
                        print("\n=== Dynamic Weights Based on RMSE ===")
                        for col, weight in weights_dict.items():
                            print(f"  {col}: weight={weight:.3f}  (RMSE={rmse_dict[col]:.2f})")
                        blended_amz = np.zeros(len(comparison_with_po))
                        for col in weights_dict:
                            blended_amz += weights_dict[col] * comparison_with_po[col].fillna(0)
                        blended_amz = np.maximum(blended_amz, 0)
                        comparison_with_po['Ensemble_Amazon'] = blended_amz
                        FALLBACK_RATIO = 0.3
                        comparison_with_po['MyForecast'] = (
                            (1 - FALLBACK_RATIO) * comparison_with_po['MyForecast'] +
                            FALLBACK_RATIO * comparison_with_po['Ensemble_Amazon']
                        ).clip(lower=0)
                    else:
                        print("Not enough overlapping actual 'y' values for dynamic weighting. Skipping dynamic weighting.")
                
                # --- Adjust out-of-range forecasts ---
                comparison_with_po = adjust_forecast_if_out_of_range(
                    comparison_with_po, asin, forecast_col_name='MyForecast', adjustment_threshold=0.3
                )
                
                # --- Adjust forecast if overall mean is too high relative to past 8 weeks ---
                past_8_weeks_with_po = ts_data_with_po.sort_values('ds').tail(8)
                if not past_8_weeks_with_po.empty and 'MyForecast' in comparison_with_po.columns:
                    past8_avg_with_po = past_8_weeks_with_po['y'].mean()
                    fc_mean_with_po = comparison_with_po['MyForecast'].mean()
                    if fc_mean_with_po > 1.5 * past8_avg_with_po:
                        print(f"[WITH PO][SARIMA] Adjusting forecast for {asin}, 8wk avg={past8_avg_with_po}, fc mean={fc_mean_with_po}")
                        comparison_with_po['MyForecast'] = (
                            0.8 * past8_avg_with_po + 0.2 * comparison_with_po['MyForecast']
                        ).clip(lower=0)
                
                # --- Seasonal adjustment ---
                try:
                    ts_data_with_po = detect_seasonal_periods(ts_data_with_po)
                    seasonal_factors_with_po = calculate_seasonal_factors(ts_data_with_po)
                    comparison_with_po = apply_seasonal_adjustment(
                        comparison_with_po,
                        ts_data_with_po,
                        seasonal_factors_with_po,
                        override_threshold=0.3,
                        max_override=1.5
                    )
                    metrics_season_with_po = validate_seasonal_adjustment(ts_data_with_po, comparison_with_po)
                    print(f"[WITH PO][SARIMA] Seasonal: MAE Improve={metrics_season_with_po['improvement_pct']:.1f}%, Strength={metrics_season_with_po['seasonal_strength']:.2f}")
                    if metrics_season_with_po['improvement_pct'] < -5:
                        print("[WITH PO][SARIMA] Reverting forecast due to poor seasonal adjustment.")
                        comparison_with_po['final_forecast'] = comparison_with_po['MyForecast']
                except Exception as e:
                    print(f"[WITH PO][SARIMA] Seasonal adjust failed for {asin}: {e}")
                    comparison_with_po['final_forecast'] = comparison_with_po['MyForecast']
                
                # Finalize MyForecast column (ensure it exists)
                comparison_with_po['MyForecast'] = comparison_with_po.get('final_forecast', comparison_with_po.get('MyForecast', 0))
                
                # --- Append Additional Inventory & Performance Columns ---
                # Guarantee 'Week_Start_Date'
                if 'Week_Start_Date' not in comparison_with_po.columns:
                    if 'ds' in comparison_with_po.columns:
                        comparison_with_po['Week_Start_Date'] = pd.to_datetime(comparison_with_po['ds'], errors='coerce').dt.strftime('%Y-%m-%d')
                    else:
                        comparison_with_po['Week_Start_Date'] = ''
                
                # Create ISO week labels
                comparison_with_po['Week_Start_Date'] = pd.to_datetime(comparison_with_po['Week_Start_Date'], errors='coerce')
                def iso_week_label(d):
                    if pd.isna(d):
                        return "W??"
                    return f"W{d.isocalendar().week + 1}"
                comparison_with_po['Week'] = comparison_with_po['Week_Start_Date'].apply(iso_week_label)
                comparison_with_po['Week_Start_Date'] = comparison_with_po['Week_Start_Date'].dt.strftime('%Y-%m-%d')
                
                # Append Trend, Inventory Coverage, etc.
                comparison_with_po['Trend'] = determine_seasonal_trend(comparison_with_po['MyForecast'])
                
                starting_inv = runrate_inventory.get(asin, 0)
                coverage_list = []
                reorder_urg_list = []
                stockout_risk_list = []
                for i in range(len(comparison_with_po)):
                    fc_val = comparison_with_po.loc[i, 'MyForecast']
                    if fc_val <= 0:
                        coverage_list.append(float('inf'))
                        reorder_urg_list.append("Normal")
                        stockout_risk_list.append("Low")
                        continue
                    cov = starting_inv / float(fc_val)
                    coverage_list.append(round(cov, 2))
                    reorder_urg_list.append("Urgent" if cov < 1.0 else "Normal")
                    stockout_risk_list.append("High" if cov < 0.5 else "Low")
                    starting_inv = max(0, starting_inv - fc_val)
                comparison_with_po['Inventory Coverage'] = coverage_list
                comparison_with_po['Reorder Urgency'] = reorder_urg_list
                comparison_with_po['Stockout Risk'] = stockout_risk_list
                
                hist_sales_52 = ts_data_with_po.sort_values('ds')['y'].tail(52).tolist()
                slope_val = compute_sales_slope(hist_sales_52)
                if slope_val > 0:
                    sales_trend_label = "Increasing (▲)"
                elif slope_val < 0:
                    sales_trend_label = "Decreasing (▼)"
                else:
                    sales_trend_label = "Stable"
                comparison_with_po['Sales Trend'] = sales_trend_label
                
                comparison_with_po = comparison_with_po.sort_values('Week_Start_Date', na_position='last').reset_index(drop=True)
                def sea_idx_func(row):
                    w_i = row.name + 1
                    return compute_seasonality_index(w_i, average_annual_sales=10000.0)
                comparison_with_po['Seasonality Index'] = comparison_with_po.apply(sea_idx_func, axis=1)
                
                lcs_stage = "Growth" if slope_val > 0.1 else ("Decline" if slope_val < -0.1 else "Mature")
                comparison_with_po['Lifecycle Stage'] = lcs_stage
                
                desired_cols = [
                    "Week", "Week_Start_Date", "ASIN", "MyForecast", "Amazon Mean Forecast",
                    "Amazon P70 Forecast", "Amazon P80 Forecast", "Amazon P90 Forecast",
                    "Product Title", "is_holiday_week",
                    "Trend", "Inventory Coverage", "Stockout Risk", "Reorder Urgency",
                    "Sales Trend", "Seasonality Index", "Lifecycle Stage"
                ]
                existing_cols = [c for c in desired_cols if c in comparison_with_po.columns]
                comparison_with_po = comparison_with_po[existing_cols].copy()
                comparison_with_po['ASIN'] = asin
                comparison_with_po['Product Title'] = product_title
                
                # Save individual forecast Excel file into the WITH_PO folder.
                output_file_name_with_po = f'forecast_summary_{asin}_WITH_PO.xlsx'
                output_file_path_with_po = os.path.join(with_po_folder, output_file_name_with_po)
                comparison_with_po.to_excel(output_file_path_with_po, index=False)
                print(f"[WITH PO][SARIMA] Forecast for ASIN {asin} saved to {output_file_path_with_po}")
                
                summary_stats_with_po, total16_with_po, total8_with_po, total4_with_po, \
                max_fc_with_po, min_fc_with_po, max_week_with_po, min_week_with_po = calculate_summary_statistics(
                    ts_data_with_po, comparison_with_po, horizon=horizon
                )
                
                save_summary_to_excel(
                    comparison_with_po,
                    summary_stats_with_po,
                    total16_with_po,
                    total8_with_po,
                    total4_with_po,
                    max_fc_with_po,
                    min_fc_with_po,
                    max_week_with_po,
                    min_week_with_po,
                    output_file_path_with_po,
                    metrics=None
                )
                
                consolidated_forecasts_with_po[asin] = comparison_with_po

            else:
                # == PROPHET WITH PO block ==
                if model_with_po is None:
                    print(f"[WITH PO][Prophet] Model is None for {asin}, skipping.")
                    continue
                if model_with_po == "Prophet_Failed":
                    print(f"[WITH PO][Prophet] No valid prophet model for {asin}, skipping.")
                    continue

                print(f"[WITH PO][Prophet] Training a new model with Optuna for ASIN {asin} ...")
                forecast_po, trained_prophet_model_po = forecast_with_optuna_params(
                    ts_data_with_po,
                    forecast_data,
                    horizon=horizon,
                    asin=asin,
                    n_trials=30  
                )

                if trained_prophet_model_po is not None:
                    save_model(trained_prophet_model_po, "Prophet_WITH_PO", asin, ts_data_with_po)
                else:
                    print(f"[WITH PO][Prophet] Failed to train new model for {asin}.")
                    continue

                if 'forecast_po' not in locals() or forecast_po.empty:
                    print(f"[WITH PO][Prophet] Forecast is empty for ASIN {asin}, skipping.")
                    continue

                comparison_with_po = format_output_with_forecasts(forecast_po, forecast_data, horizon=horizon)
                best_weights_po, best_rmse_po = auto_find_best_weights(forecast_po, comparison_with_po, step=0.05)
                print(f"[WITH PO][Prophet] Auto best weights for ASIN {asin}: {best_weights_po} (RMSE={best_rmse_po})")

                forecast_po = adjust_forecast_weights(forecast_po.copy(), *best_weights_po)
                comparison_with_po = format_output_with_forecasts(forecast_po, forecast_data, horizon=horizon)

                print("\n--- Forecast Before Out-of-Range Adjustment (WITH PO) ---")
                print(comparison_with_po[['MyForecast']].head(10))
                print("---------------------------------------------------------\n")

                comparison_with_po = adjust_forecast_if_out_of_range(
                    comparison_with_po, asin, adjustment_threshold=0.3
                )
                log_fallback_triggers(comparison_with_po, asin, product_title)

                past_8_weeks_po = ts_data_with_po.sort_values('ds').tail(8)
                if not past_8_weeks_po.empty and 'MyForecast' in comparison_with_po.columns:
                    past8_avg_po = past_8_weeks_po['y'].mean()
                    fc_mean_po = comparison_with_po['MyForecast'].mean()
                    if fc_mean_po > 1.5 * past8_avg_po:
                        print(f"[WITH PO][Prophet] Adjusting forecast for {asin}: 8wk avg={past8_avg_po}, fc mean={fc_mean_po}")
                        comparison_with_po['MyForecast'] = (
                            0.8 * past8_avg_po + 
                            0.2 * comparison_with_po['MyForecast']
                        ).clip(lower=0)

                try:
                    ts_data_with_po = detect_seasonal_periods(ts_data_with_po)
                    seasonal_factors_with_po = calculate_seasonal_factors(ts_data_with_po)
                    comparison_with_po = apply_seasonal_adjustment(
                        comparison_with_po,
                        ts_data_with_po,
                        seasonal_factors_with_po,
                        override_threshold=0.3,
                        max_override=1.5
                    )
                    metrics_season_with_po = validate_seasonal_adjustment(ts_data_with_po, comparison_with_po)
                    print(
                        f"[WITH PO][Prophet] Seasonal adjust: MAE Improve={metrics_season_with_po['improvement_pct']:.1f}%, "
                        f"Strength={metrics_season_with_po['seasonal_strength']:.2f}"
                    )
                    if metrics_season_with_po['improvement_pct'] < -5:
                        print("[WITH PO][Prophet] Reverting forecast due to poor seasonal adjust.")
                        comparison_with_po['final_forecast'] = comparison_with_po['MyForecast']
                except Exception as e:
                    print(f"[WITH PO][Prophet] Seasonal adjust failed for {asin}: {e}")
                    comparison_with_po['final_forecast'] = comparison_with_po['MyForecast']

                comparison_with_po['MyForecast'] = comparison_with_po['final_forecast']

                if 'ds' in comparison_with_po.columns and 'y' not in comparison_with_po.columns:
                    comparison_with_po = comparison_with_po.merge(ts_data_with_po[['ds','y']], on='ds', how='left')

                # Summaries & Visualization
                summary_stats_with_po, total16_with_po, total8_with_po, total4_with_po, \
                max_fc_with_po, min_fc_with_po, max_week_with_po, min_week_with_po = calculate_summary_statistics(
                    ts_data_with_po, comparison_with_po, horizon=horizon
                )
                visualize_forecast_with_comparison(
                    ts_data_with_po,
                    comparison_with_po,
                    summary_stats_with_po,
                    total16_with_po,
                    total8_with_po,
                    total4_with_po,
                    max_fc_with_po,
                    min_fc_with_po,
                    max_week_with_po,
                    min_week_with_po,
                    asin,
                    product_title,
                    with_po_folder
                )
                log_fallback_triggers(comparison_with_po, asin, product_title)

                # -------------------------
                # Append Additional Inventory & Performance Columns
                # -------------------------
                # 1) Guarantee 'Week_Start_Date'
                if 'Week_Start_Date' not in comparison_with_po.columns:
                    if 'ds' in comparison_with_po.columns:
                        comparison_with_po['Week_Start_Date'] = pd.to_datetime(
                            comparison_with_po['ds'], errors='coerce'
                        ).dt.strftime('%Y-%m-%d')
                    else:
                        comparison_with_po['Week_Start_Date'] = ''

                # 2) Guarantee 'Week'
                comparison_with_po['Week_Start_Date'] = pd.to_datetime(comparison_with_po['Week_Start_Date'], errors='coerce')

                # Now compute the ISO week
                def iso_week_label(d):
                    if pd.isna(d):
                        return "W?"
                    return f"W{d.isocalendar().week + 1}"

                comparison_with_po['Week'] = comparison_with_po['Week_Start_Date'].apply(iso_week_label)
                comparison_with_po['Week_Start_Date'] = comparison_with_po['Week_Start_Date'].dt.strftime('%Y-%m-%d')


                # 3) Trend
                comparison_with_po['Trend'] = determine_seasonal_trend(comparison_with_po['MyForecast'])

                # 4) Row-by-row coverage
                starting_inv = runrate_inventory.get(asin, 0)
                coverage_list = []
                reorder_urg_list = []
                stockout_risk_list = []

                for irow in range(len(comparison_with_po)):
                    fc_val = comparison_with_po.loc[irow,'MyForecast']
                    if fc_val <=0:
                        coverage_list.append(float('inf'))
                        reorder_urg_list.append("Normal")
                        stockout_risk_list.append("Low")
                        continue
                    
                    coverage = starting_inv / float(fc_val)
                    coverage_list.append(round(coverage,2))

                    # Mark reorder urgency if coverage < 1.0
                    if coverage < 1.0:
                        reorder_urg_list.append("Urgent")
                    else:
                        reorder_urg_list.append("Normal")

                    # Mark stockout risk if coverage < 0.5
                    if coverage < 0.5:
                        stockout_risk_list.append("High")
                    else:
                        stockout_risk_list.append("Low")

                    starting_inv = max(0, starting_inv - fc_val)

                comparison_with_po['Inventory Coverage'] = coverage_list
                comparison_with_po['Reorder Urgency'] = reorder_urg_list
                comparison_with_po['Stockout Risk'] = stockout_risk_list

                # 5) Sales Trend => arrow labels
                hist_sales_52 = ts_data_with_po.sort_values('ds')['y'].tail(52).tolist()
                slope_val = compute_sales_slope(hist_sales_52)  # numeric slope
                if slope_val > 0:
                    sales_trend_label = "Increasing (▲)"
                elif slope_val < 0:
                    sales_trend_label = "Decreasing (▼)"
                else:
                    sales_trend_label = "Stable"
                comparison_with_po['Sales Trend'] = sales_trend_label

                # 6) Seasonality
                comparison_with_po = comparison_with_po.sort_values('Week_Start_Date', na_position='last').reset_index(drop=True)
                def sea_idx_func(row):
                    w_i = row.name + 1
                    return compute_seasonality_index(w_i, average_annual_sales=10000.0)
                comparison_with_po['Seasonality Index'] = comparison_with_po.apply(sea_idx_func, axis=1)

                # 7) Lifecycle Stage from slope
                if slope_val > 0.1:
                    lcs_stage = "Growth"
                elif slope_val < -0.1:
                    lcs_stage = "Decline"
                else:
                    lcs_stage = "Mature"
                comparison_with_po['Lifecycle Stage'] = lcs_stage

                # 8) Reorder columns (no "Sales Volume Rank" now)
                desired_cols = [
                    "Week", "Week_Start_Date", "ASIN", "MyForecast", "Amazon Mean Forecast",
                    "Amazon P70 Forecast", "Amazon P80 Forecast", "Amazon P90 Forecast",
                    "Product Title", "is_holiday_week",
                    "Trend", "Inventory Coverage", "Stockout Risk", "Reorder Urgency",
                    "Sales Trend", "Seasonality Index", "Lifecycle Stage"
                ]
                existing_cols = [c for c in desired_cols if c in comparison_with_po.columns]
                comparison_with_po = comparison_with_po[existing_cols].copy()

                comparison_with_po['ASIN'] = asin
                comparison_with_po['Product Title'] = product_title

                # Final save
                output_file_name_with_po = f'forecast_summary_{asin}_WITH_PO.xlsx'
                output_file_path_with_po = os.path.join(with_po_folder, output_file_name_with_po)
                comparison_with_po.to_excel(output_file_path_with_po, index=False)

                # Summaries & save
                summary_stats_with_po, total16_with_po, total8_with_po, total4_with_po, \
                max_fc_with_po, min_fc_with_po, max_week_with_po, min_week_with_po = calculate_summary_statistics(
                    ts_data_with_po, comparison_with_po, horizon=horizon
                )

                save_summary_to_excel(
                    comparison_with_po,
                    summary_stats_with_po,
                    total16_with_po,
                    total8_with_po,
                    total4_with_po,
                    max_fc_with_po,
                    min_fc_with_po,
                    max_week_with_po,
                    min_week_with_po,
                    output_file_path_with_po,
                    metrics=None
                )

                consolidated_forecasts_with_po[asin] = comparison_with_po

    # ----------------------------------------------------------------------------
    # 6) After processing ALL ASINs, save consolidated data
    # ----------------------------------------------------------------------------
    final_output_path_with_po = "consolidated_forecast_WITH_PO.xlsx"
    save_forecast_to_excel(final_output_path_with_po, consolidated_forecasts_with_po, missing_asin_data)
    print(f"[WITH PO] All per-ASIN forecasts saved to {final_output_path_with_po}")

    generate_restock_suggestions(
        consolidated_forecasts_with_po,
        runrate_inventory,
        coverage_threshold=1.0,
        output_file="restock_suggestions.xlsx",
        target_weeks=4
    )
    print("\n Restock Suggestion completed. \n")

    save_feedback_to_excel(prophet_feedback, "prophet_feedback.xlsx")
    generate_4_week_report(consolidated_forecasts_with_po)
    generate_combined_weekly_report(consolidated_forecasts_with_po)
    generate_low_coverage_report(
        consolidated_forecasts_with_po,
        coverage_threshold=1.0,
        output_file="ASINs_LowCoverage_Urgent.xlsx"
    )

    save_param_histories()
    print(f"Total number of parameter sets tested: {PARAM_COUNTER}")
    if POOR_PARAM_FOUND:
        print("Note: Early stopping occurred for some ASINs due to poor parameter performance.")


##############################
# Run the Main Script
##############################
if __name__ == '__main__':
    main()
    summarize_usage()
    main_po_forecast_pipeline()

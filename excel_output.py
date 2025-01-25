# excel_output.py

import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def save_summary_to_excel(comparison,
                          summary_stats,
                          total_forecast_16,
                          total_forecast_8,
                          total_forecast_4,
                          max_forecast,
                          min_forecast,
                          max_week,
                          min_week,
                          output_file_path,
                          metrics=None,
                          base_year=2025):

    def week_label_to_date(week_label, base_year):
        # Convert week label like 'W01' to a start date
        week_num = int(week_label[1:])
        # Assume first week of base_year starts on January 1st
        start_date = pd.Timestamp(f'{base_year}-01-01') + pd.Timedelta(weeks=week_num-1)
        return start_date.strftime('%Y-%m-%d')

    # Create Week_Start_Date based on available information
    if 'ds' in comparison.columns:
        comparison['ds'] = pd.to_datetime(comparison['ds'], errors='coerce')
        comparison['Week_Start_Date'] = comparison['ds'].dt.strftime('%Y-%m-%d')
    elif 'Week' in comparison.columns:
        comparison['Week_Start_Date'] = comparison['Week'].apply(lambda w: week_label_to_date(w, base_year))

    # Create Week label column if not present
    if 'Week' not in comparison.columns:
        # Sort by Week_Start_Date to ensure correct order
        comparison = comparison.sort_values('Week_Start_Date').reset_index(drop=True)
        comparison['Week'] = ['W' + str(i+1) for i in range(len(comparison))]

    # Round forecast values to integers for specified columns
    for col in ['MyForecast', 'Amazon Mean Forecast', 'Amazon P70 Forecast',
                'Amazon P80 Forecast', 'Amazon P90 Forecast']:
        if col in comparison.columns:
            comparison[col] = comparison[col].round().astype(int)

    # Drop 'ds' column if it exists since we now use Week and Week_Start_Date
    if 'ds' in comparison.columns:
        comparison.drop(columns=['ds'], inplace=True)

    # Define the desired column order, including the new columns
    columns_to_include = [
        'Week', 'Week_Start_Date', 'ASIN', 'MyForecast', 'Amazon Mean Forecast',
        'Amazon P70 Forecast', 'Amazon P80 Forecast', 'Amazon P90 Forecast',
        'Product Title', 'is_holiday_week'
    ]

    # Ensure all desired columns are present in the DataFrame
    for col in columns_to_include:
        if col not in comparison.columns:
            comparison[col] = np.nan

    comparison = comparison[columns_to_include]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Forecast Comparison"

    for r in dataframe_to_rows(comparison, index=False, header=True):
        ws1.append(r)

    # Prepare summary sheet as before
    ws2 = wb.create_sheet(title="Summary")
    summary_data = {
        "Metric": [
            "Historical Range",
            "Min Sales",
            "Max Sales",
            "Mean Sales",
            "Median Sales",
            "Std Dev Sales",
            "Total Historical Sales",
            "Total Forecast (16 Weeks)",
            "Total Forecast (8 Weeks)",
            "Total Forecast (4 Weeks)",
            "Max Forecast",
            "Max Forecast Week",
            "Min Forecast",
            "Min Forecast Week"
        ],
        "Value": [
            f"{summary_stats['data_range'][0].strftime('%Y-%m-%d')} to {summary_stats['data_range'][1].strftime('%Y-%m-%d')}",
            f"{summary_stats['min']:.0f}",
            f"{summary_stats['max']:.0f}",
            f"{summary_stats['mean']:.0f}",
            f"{summary_stats['median']:.0f}",
            f"{summary_stats['std_dev']:.0f}",
            f"{summary_stats['total_sales']:.0f} units",
            f"{total_forecast_16:.0f}",
            f"{total_forecast_8:.0f}",
            f"{total_forecast_4:.0f}",
            f"{max_forecast:.0f}",
            f"{max_week.strftime('%Y-%m-%d') if max_week else 'N/A'}",
            f"{min_forecast:.0f}",
            f"{min_week.strftime('%Y-%m-%d') if min_week else 'N/A'}"
        ]
    }
    if metrics is not None:
        for k, v in metrics.items():
            summary_data["Metric"].append(k)
            summary_data["Value"].append(str(v))

    summary_df = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws2.append(r)

    wb.save(output_file_path)
    print(f"Comparison and summary saved to '{output_file_path}'")

def save_forecast_to_excel(output_path, consolidated_data, missing_asin_data, base_year=2025):
    """
    Save multiple ASIN forecasts (with "MyForecast") and any missing ASIN data into one Excel file,
    each ASIN in a separate sheet.
    """
    desired_columns = [
        'Week', 'Week_Start_Date', 'ASIN', 'MyForecast', 'Amazon Mean Forecast',
        'Amazon P70 Forecast', 'Amazon P80 Forecast', 'Amazon P90 Forecast',
        'Product Title', 'is_holiday_week'
    ]

    def week_label_to_date(week_label, base_year):
        week_num = int(week_label[1:])
        start_date = pd.Timestamp(f'{base_year}-01-01') + pd.Timedelta(weeks=week_num-1)
        return start_date.strftime('%Y-%m-%d')

    wb = Workbook()

    for asin, df in consolidated_data.items():
        df_copy = df.copy()

        # Calculate Week_Start_Date
        if 'ds' in df_copy.columns:
            df_copy['ds'] = pd.to_datetime(df_copy['ds'], errors='coerce')
            df_copy['Week_Start_Date'] = df_copy['ds'].dt.strftime('%Y-%m-%d')
        elif 'Week' in df_copy.columns:
            df_copy['Week_Start_Date'] = df_copy['Week'].apply(lambda w: week_label_to_date(w, base_year))

        # Round forecast columns to integers
        forecast_cols = ['MyForecast', 'Amazon Mean Forecast', 'Amazon P70 Forecast', 
                         'Amazon P80 Forecast', 'Amazon P90 Forecast']
        for col in forecast_cols:
            if col in df_copy.columns:
                df_copy[col] = df_copy[col].round().astype(int)

        # Create Week labels if missing
        if 'Week' not in df_copy.columns:
            df_copy = df_copy.sort_values('Week_Start_Date').reset_index(drop=True)
            df_copy['Week'] = ['W' + str(i+1).zfill(2) for i in range(len(df_copy))]

        # Ensure all desired columns exist
        for col in desired_columns:
            if col not in df_copy.columns:
                df_copy[col] = np.nan

        # Fill 'ASIN' and 'Product Title' if missing
        df_copy['ASIN'] = asin
        df_copy['Product Title'] = df['Product Title'].iloc[0] if 'Product Title' in df.columns else ''

        # Select desired columns only
        df_copy = df_copy[desired_columns]

        ws = wb.create_sheet(title=str(asin)[:31])  # Excel sheet names limited to 31 characters

        for r in dataframe_to_rows(df_copy, index=False, header=True):
            ws.append(r)

    # Handle missing ASIN data
    if not missing_asin_data.empty:
        ws_missing = wb.create_sheet(title="No ASIN")
        for r in dataframe_to_rows(missing_asin_data, index=False, header=True):
            ws_missing.append(r)

    # Remove default 'Sheet' if it exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Add a Summary Sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["ASIN", "Product Title", "4 Week Forecast", "8 Week Forecast", "16 Week Forecast"])

    for asin, df in consolidated_data.items():
        if df.empty:
            ws_summary.append([asin, "No data", None, None, None])
            continue

        product_title = df['Product Title'].iloc[0] if 'Product Title' in df.columns else ''
        four_wk_val = df['MyForecast'].iloc[:4].sum() if len(df) >= 4 else df['MyForecast'].sum()
        eight_wk_val = df['MyForecast'].iloc[:8].sum() if len(df) >= 8 else df['MyForecast'].sum()
        sixteen_wk_val = df['MyForecast'].iloc[:16].sum() if len(df) >= 16 else df['MyForecast'].sum()

        ws_summary.append([asin, product_title, four_wk_val, eight_wk_val, sixteen_wk_val])

    wb.save(output_path)
    print(f"All forecasts saved to '{output_path}'")

def save_feedback_to_excel(feedback_dict, filename):
    """
    Save feedback information from models into an Excel file.
    """
    records = []
    for asin, info in feedback_dict.items():
        record = {'ASIN': asin}
        best_params = info.get('best_params', {})
        for param, value in best_params.items():
            record[param] = value
        if 'rmse_values' in info and info['rmse_values']:
            for k, v in info['rmse_values'].items():
                record[f'RMSE_{k}'] = v
        record['Total Tests'] = info.get('total_tests', None)
        records.append(record)
    
    df_feedback = pd.DataFrame(records)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_feedback.to_excel(writer, index=False, sheet_name='Model Feedback')
    print(f"Feedback saved to {filename}")

def generate_4_week_report(consolidated_forecasts):
    """
    Generate an Excel report for the first 4 weeks of "MyForecast" for each ASIN,
    plus Amazon columns if present.
    """
    report_rows = []
    for asin, comp_df in consolidated_forecasts.items():
        # Determine the model used based on available info
        # (We only unify the final forecast column as "MyForecast")
        product_title = ''
        if 'Product Title' in comp_df.columns and not comp_df.empty:
            product_title = comp_df['Product Title'].iloc[0]

        if 'MyForecast' in comp_df.columns:
            model_used = 'Unified'
            my_forecast_column = 'MyForecast'
        else:
            model_used = 'Unknown'
            my_forecast_column = None

        if my_forecast_column and my_forecast_column in comp_df.columns:
            my_4w_forecast = comp_df[my_forecast_column].iloc[:4].sum()
        else:
            my_4w_forecast = np.nan

        amz_mean = comp_df['Amazon Mean Forecast'].iloc[:4].sum() if 'Amazon Mean Forecast' in comp_df.columns else np.nan
        amz_p70  = comp_df['Amazon P70 Forecast'].iloc[:4].sum() if 'Amazon P70 Forecast' in comp_df.columns else np.nan
        amz_p80  = comp_df['Amazon P80 Forecast'].iloc[:4].sum() if 'Amazon P80 Forecast' in comp_df.columns else np.nan
        amz_p90  = comp_df['Amazon P90 Forecast'].iloc[:4].sum() if 'Amazon P90 Forecast' in comp_df.columns else np.nan

        report_rows.append({
            'ASIN': asin,
            'Product Title': product_title,
            'My 4 Weeks Forecast': my_4w_forecast,
            'AMZ Forecast Mean': amz_mean,
            'AMZ Forecast P70': amz_p70,
            'AMZ Forecast P80': amz_p80,
            'AMZ Forecast P90': amz_p90
        })

    report_df = pd.DataFrame(report_rows)
    report_filename = '4_week_report.xlsx'
    report_df.to_excel(report_filename, index=False)
    print(f"4-week report saved to {report_filename}")

def generate_combined_weekly_report(consolidated_forecasts):
    """
    Generate an Excel report summarizing 4-, 8-, and 16-week forecasts for each ASIN.
    
    Output Columns:
    ASIN, Product Name, 4 Weeks Forecast, 8 Weeks Forecast, 16 Weeks Forecast,
    """
    report_rows = []
    for asin, df in consolidated_forecasts.items():
        # Extract product name if available
        product_name = ""
        if 'Product Title' in df.columns and not df.empty:
            product_name = df['Product Title'].iloc[0]
        
        # Use 'MyForecast' column; if missing, default to zeros
        my_forecast = df['MyForecast'] if 'MyForecast' in df.columns else pd.Series([0]*len(df))
        # Use 'Amazon Mean Forecast' column; if missing, default to zeros
        #amz_forecast = df['Amazon Mean Forecast'] if 'Amazon Mean Forecast' in df.columns else pd.Series([0]*len(df))

        # Calculate cumulative sums for specified weeks
        forecast_4 = int(round(my_forecast.iloc[:4].sum())) if len(my_forecast) >= 4 else int(round(my_forecast.sum()))
        forecast_8 = int(round(my_forecast.iloc[:8].sum())) if len(my_forecast) >= 8 else int(round(my_forecast.sum()))
        forecast_16 = int(round(my_forecast.iloc[:16].sum())) if len(my_forecast) >= 16 else int(round(my_forecast.sum()))

        report_rows.append({
            'ASIN': asin,
            'Product Name': product_name,
            '4 Weeks Forecast': forecast_4,
            '8 Weeks Forecast': forecast_8,
            '16 Weeks Forecast': forecast_16
        })

    report_df = pd.DataFrame(report_rows)
    report_filename = 'combined_4_8_16_week_report.xlsx'
    report_df.to_excel(report_filename, index=False)
    print(f"Combined 4-8-16 week report saved to {report_filename}")

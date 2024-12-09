import pandas as pd
import numpy as np
from prophet import Prophet
from prophet.diagnostics import cross_validation, performance_metrics
import matplotlib.pyplot as plt
import warnings
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")  # Ignore warnings for cleaner output

def generate_date_from_week(row):
    week_str = row['week']
    year = row['year']
    week_number = int(week_str[1:])
    return pd.to_datetime(f'{year}-W{week_number - 1}-0', format='%Y-W%U-%w')

def load_weekly_sales_data(file_path):
    data = pd.read_excel(file_path)
    data.columns = data.columns.str.strip().str.lower()

    required_columns = ['product title', 'week', 'year', 'units_sold', 'asin']
    missing_required = [col for col in required_columns if col not in data.columns]
    if missing_required:
        raise ValueError(f"Missing required columns: {missing_required}")

    data['date'] = data.apply(generate_date_from_week, axis=1)
    data = data.rename(columns={'units_sold': 'y'})
    return data

def load_asins_to_forecast(file_path):
    if file_path.endswith('.txt'):
        with open(file_path, 'r') as file:
            asins = [line.strip() for line in file if line.strip()]
    elif file_path.endswith('.xlsx') or file_path.endswith('.xls'):
        df = pd.read_excel(file_path)
        asins = df.iloc[:, 0].dropna().astype(str).tolist()
    else:
        raise ValueError("Unsupported file format for ASINs to Forecast file.")
    return asins

def load_amazon_forecasts_from_folder(folder_path, asin):
    forecast_data = {}
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            forecast_type = os.path.splitext(file_name)[0].replace('_', ' ').title()
            file_path = os.path.join(folder_path, file_name)
            df = pd.read_excel(file_path)

            df.columns = df.columns.str.strip().str.upper()
            if 'ASIN' not in df.columns:
                print(f"Error: Column 'ASIN' not found in {file_name}")
                continue

            asin_row = df[df['ASIN'].str.upper() == asin.upper()]
            if asin_row.empty:
                print(f"Warning: ASIN {asin} not found in {file_name}")
                continue

            week_columns = [col for col in df.columns if 'W' in col.upper()]
            if not week_columns:
                print(f"Warning: No week columns found in {file_name}")
                continue

            forecast_values = (
                asin_row.iloc[0][week_columns]
                .astype(str)
                .str.replace(',', '')
                .astype(float)
                .round()
                .astype(int)
                .values
            )
            forecast_data[forecast_type] = forecast_values

    return forecast_data

def add_lag_features(ts_data, lag_weeks=1):
    ts_data = ts_data.copy()
    ts_data = ts_data.sort_values('ds')
    ts_data[f'lag_{lag_weeks}_week'] = ts_data['y'].shift(lag_weeks)
    ts_data[f'lag_{lag_weeks}_week'].fillna(0, inplace=True)
    return ts_data

def prepare_time_series_with_lags(data, asin, lag_weeks=1):
    ts_data = data[data['asin'] == asin].rename(columns={'date': 'ds', 'y': 'y'})
    ts_data = ts_data.sort_values('ds')
    ts_data['y'] = ts_data['y'].interpolate().bfill().clip(lower=0)
    ts_data = add_lag_features(ts_data, lag_weeks)
    return ts_data

def get_shifted_holidays():
    holidays_list = [
        ('Prime Day', '2023-06-27'),
        ('Prime Day', '2023-06-28'),
        ('Black Friday', '2023-11-10'),
        ('Cyber Monday', '2023-11-13'),
        ('Christmas', '2023-12-11'),
        ('Prime Day', '2024-07-02'),
        ('Prime Day', '2024-07-03'),
        ('Black Friday', '2024-11-15'),
        ('Cyber Monday', '2024-11-18'),
        ('Christmas', '2024-12-09'),
    ]
    holidays = pd.DataFrame(holidays_list, columns=['holiday', 'ds'])
    holidays['ds'] = pd.to_datetime(holidays['ds'])
    return holidays

def forecast_with_custom_params(ts_data, forecast_data, changepoint_prior_scale, seasonality_prior_scale, holidays_prior_scale, horizon=20):
    future_dates = pd.date_range(start=ts_data['ds'].max() + pd.Timedelta(days=7), periods=horizon, freq='W')
    future = pd.DataFrame({'ds': future_dates})

    combined_df = pd.concat([ts_data, future], ignore_index=True)

    for forecast_type, values in forecast_data.items():
        values_to_use = values[:horizon] if len(values) > horizon else values
        extended_values = np.concatenate(
            [
                np.full(len(ts_data), np.nan),
                values_to_use,
                np.full(max(horizon - len(values_to_use), 0), values_to_use[-1] if len(values_to_use) > 0 else 0)
            ]
        )
        extended_values = extended_values[:len(combined_df)]
        combined_df[f'Amazon_{forecast_type}'] = extended_values

    regressor_cols = [col for col in combined_df.columns if col.startswith('Amazon_')]
    combined_df[regressor_cols] = combined_df[regressor_cols].fillna(0)

    holidays = get_shifted_holidays()
    combined_df['prime_day'] = combined_df['ds'].apply(
        lambda x: 0.2 if x in holidays[holidays['holiday'] == 'Prime Day']['ds'].values else 0
    )

    train_df = combined_df[~combined_df['y'].isna()].copy()
    future_df = combined_df[combined_df['y'].isna()].drop(columns='y').copy()

    model = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=True,
        seasonality_mode='multiplicative',
        holidays=holidays,
        changepoint_prior_scale=changepoint_prior_scale,
        seasonality_prior_scale=seasonality_prior_scale,
        holidays_prior_scale=holidays_prior_scale
    )

    for regressor in regressor_cols + ['prime_day']:
        model.add_regressor(regressor, mode='multiplicative')

    try:
        model.fit(train_df)
        forecast = model.predict(future_df)
        forecast['Prophet Forecast'] = forecast['yhat']
        return forecast[['ds', 'Prophet Forecast', 'yhat', 'yhat_upper']], model
    except Exception as e:
        print(f"Error during forecasting: {e}")
        return pd.DataFrame(columns=['ds', 'Prophet Forecast']), None

def optimize_prophet_params(ts_data, forecast_data, param_grid, horizon=20):
    best_rmse = float('inf')
    best_params = None
    best_rmse_values = None

    for changepoint_prior_scale in param_grid['changepoint_prior_scale']:
        for seasonality_prior_scale in param_grid['seasonality_prior_scale']:
            for holidays_prior_scale in param_grid['holidays_prior_scale']:
                try:
                    forecast, _ = forecast_with_custom_params(
                        ts_data, forecast_data,
                        changepoint_prior_scale,
                        seasonality_prior_scale,
                        holidays_prior_scale,
                        horizon
                    )
                    if forecast.empty or 'Prophet Forecast' not in forecast.columns:
                        raise ValueError("Forecast failed to generate 'Prophet Forecast'.")

                    rmse_values = calculate_rmse(forecast, forecast_data, horizon)
                    avg_rmse = np.mean(list(rmse_values.values()))

                    if avg_rmse < best_rmse:
                        best_rmse = avg_rmse
                        best_params = {
                            'changepoint_prior_scale': changepoint_prior_scale,
                            'seasonality_prior_scale': seasonality_prior_scale,
                            'holidays_prior_scale': holidays_prior_scale
                        }
                        best_rmse_values = rmse_values
                except Exception as e:
                    print(f"Error during optimization: {e}")
                    continue

    if best_params is None:
        print("Optimization failed. Using default parameters.")
        best_params = {'changepoint_prior_scale': 0.1, 'seasonality_prior_scale': 1, 'holidays_prior_scale': 10}
        best_rmse_values = {}

    return best_params, best_rmse_values

def calculate_rmse(forecast, forecast_data, horizon):
    rmse_values = {}
    for forecast_type, values in forecast_data.items():
        values = values[:horizon]
        rmse = np.sqrt(((forecast['Prophet Forecast'] - values) ** 2).mean())
        rmse_values[forecast_type] = rmse
    return rmse_values

def format_output_with_forecasts(prophet_forecast, forecast_data, horizon=20):
    comparison = prophet_forecast.copy()
    for forecast_type, values in forecast_data.items():
        values = values[:horizon]
        forecast_df = pd.DataFrame({
            'ds': prophet_forecast['ds'],
            f'Amazon {forecast_type}': values
        })
        comparison = comparison.merge(forecast_df, on='ds', how='left')

    comparison.fillna(0, inplace=True)
    return comparison

def adjust_forecast_weights(forecast, yhat_weight, yhat_upper_weight):
    if 'yhat' not in forecast or 'yhat_upper' not in forecast:
        raise KeyError("'yhat' or 'yhat_upper' not found in forecast DataFrame.")

    forecast['Prophet Forecast'] = (
        yhat_weight * forecast['yhat'] + yhat_upper_weight * forecast['yhat_upper']
    ).clip(lower=0).round().astype(int)
    return forecast

def find_best_forecast_weights(forecast, comparison, weights):
    best_rmse = float('inf')
    best_weights = None
    rmse_results = {}

    for yhat_weight, yhat_upper_weight in weights:
        adjusted_forecast = adjust_forecast_weights(forecast.copy(), yhat_weight, yhat_upper_weight)
        rmse_values = {}
        for amazon_col in comparison.columns:
            if amazon_col.startswith('Amazon '):
                rmse = np.sqrt(((comparison[amazon_col] - adjusted_forecast['Prophet Forecast']) ** 2).mean())
                rmse_values[amazon_col] = rmse

        avg_rmse = np.mean(list(rmse_values.values()))
        rmse_results[(yhat_weight, yhat_upper_weight)] = avg_rmse

        if avg_rmse < best_rmse:
            best_rmse = avg_rmse
            best_weights = (yhat_weight, yhat_upper_weight)

    return best_weights, rmse_results

def auto_find_best_weights(forecast, comparison, step=0.05):
    best_rmse = float('inf')
    best_weights = None

    candidates = np.arange(0.5, 1.0 + step, step)
    for w in candidates:
        yhat_weight = w
        yhat_upper_weight = 1 - w
        adjusted_forecast = adjust_forecast_weights(forecast.copy(), yhat_weight, yhat_upper_weight)

        rmse_values = {}
        for amazon_col in comparison.columns:
            if amazon_col.startswith('Amazon '):
                rmse = np.sqrt(((comparison[amazon_col] - adjusted_forecast['Prophet Forecast']) ** 2).mean())
                rmse_values[amazon_col] = rmse

        avg_rmse = np.mean(list(rmse_values.values()))
        if avg_rmse < best_rmse:
            best_rmse = avg_rmse
            best_weights = (yhat_weight, yhat_upper_weight)

    return best_weights, best_rmse

def try_cross_validation_with_fallback(model, ts_data, horizons, initial='365 days', period='180 days'):
    for horizon in horizons:
        try:
            df_cv = cross_validation(model, initial=initial, period=period, horizon=horizon)
            df_performance = performance_metrics(df_cv)
            print(f"Cross-validation successful with horizon={horizon}")
            return df_cv, df_performance
        except ValueError as e:
            if "Less data than horizon" in str(e):
                print(f"Not enough data for horizon={horizon}, trying smaller horizon.")
            else:
                print(f"Error with horizon={horizon}: {e}")
    print("The ASIN has too little data for cross-validation with any tested horizons.")
    return None, None

def validate_best_params(ts_data, best_params, initial='365 days', period='180 days'):
    model = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=True,
        seasonality_mode='multiplicative',
        changepoint_prior_scale=best_params['changepoint_prior_scale'],
        seasonality_prior_scale=best_params['seasonality_prior_scale'],
        holidays_prior_scale=best_params['holidays_prior_scale']
    )
    model.fit(ts_data[['ds','y']])

    candidate_horizons = ['180 days', '90 days', '60 days', '30 days']
    df_cv, df_performance = try_cross_validation_with_fallback(model, ts_data, candidate_horizons, initial=initial, period=period)

    if df_cv is not None and df_performance is not None:
        print("Validation Metrics with chosen best_params:")
        print(df_performance)
    else:
        print("No valid cross-validation results due to insufficient data.")

def calculate_summary_statistics(ts_data, forecast_df, horizon):
    summary_stats = {
        "min": ts_data["y"].min(),
        "max": ts_data["y"].max(),
        "mean": ts_data["y"].mean(),
        "median": ts_data["y"].median(),
        "std_dev": ts_data["y"].std(),
        "total_sales": ts_data["y"].sum(),
        "data_range": (ts_data["ds"].min(), ts_data["ds"].max())
    }
    total_forecast_16 = forecast_df['Prophet Forecast'][:16].sum()
    total_forecast_8 = forecast_df['Prophet Forecast'][:8].sum()
    total_forecast_4 = forecast_df['Prophet Forecast'][:4].sum()
    max_forecast = forecast_df['Prophet Forecast'].max()
    min_forecast = forecast_df['Prophet Forecast'].min()
    max_week = forecast_df.loc[forecast_df['Prophet Forecast'].idxmax(), 'ds']
    min_week = forecast_df.loc[forecast_df['Prophet Forecast'].idxmin(), 'ds']
    return summary_stats, total_forecast_16, total_forecast_8, total_forecast_4, max_forecast, min_forecast, max_week, min_week

def visualize_forecast_with_comparison(ts_data, comparison, summary_stats, total_forecast_16, total_forecast_8, total_forecast_4, max_forecast, min_forecast, max_week, min_week, asin, product_title):
    # Use ds from comparison and ts_data as datetime for proper plotting
    fig, ax = plt.subplots(figsize=(16, 10))
    ax.plot(ts_data['ds'], ts_data['y'], label='Historical Sales', marker='o', color='black')
    ax.plot(comparison['ds'], comparison['Prophet Forecast'], label='Prophet Forecast', marker='o', linestyle='--', color='blue')

    for column in comparison.columns:
        if column.startswith('Amazon '):
            ax.plot(comparison['ds'], comparison[column], label=f'{column}', linestyle=':')

    ax.set_title(f'Sales Forecast Comparison for {product_title}')
    ax.set_xlabel('Date')
    ax.set_ylabel('Units Sold')
    ax.legend(loc='upper left', bbox_to_anchor=(1.05, 1))
    ax.grid()

    summary_text = (
        f"Historical Summary:\n"
        f"Range: {summary_stats['data_range'][0].strftime('%Y-%m-%d')} to {summary_stats['data_range'][1].strftime('%Y-%m-%d')}\n"
        f"Min: {summary_stats['min']:.0f}\n"
        f"Max: {summary_stats['max']:.0f}\n"
        f"Mean: {summary_stats['mean']:.0f}\n"
        f"Median: {summary_stats['median']:.0f}\n"
        f"Std Dev: {summary_stats['std_dev']:.0f}\n"
        f"Total Historical Sales: {summary_stats['total_sales']:.0f} units\n\n"
        f"Forecast Summary:\n"
        f"Total Forecast (16 Weeks): {total_forecast_16:.0f}\n"
        f"Total Forecast (8 Weeks): {total_forecast_8:.0f}\n"
        f"Total Forecast (4 Weeks): {total_forecast_4:.0f}\n"
        f"Max Forecast: {max_forecast:.0f} (Week of {max_week.strftime('%Y-%m-%d')})\n"
        f"Min Forecast: {min_forecast:.0f} (Week of {min_week.strftime('%Y-%m-%d')})"
    )

    plt.gcf().text(0.78, 0.5, summary_text, fontsize=10, bbox=dict(facecolor='white', alpha=0.8), va='center')
    plt.subplots_adjust(right=0.75)

    graph_folder = 'forecast_graphs'
    os.makedirs(graph_folder, exist_ok=True)
    graph_file_path = os.path.join(graph_folder, f"{product_title.replace('/', '_')}_{asin}.png")
    plt.savefig(graph_file_path)
    plt.close()
    print(f"Graph saved to {graph_file_path}")

def save_summary_to_excel(comparison, summary_stats, total_forecast_16, total_forecast_8, total_forecast_4, max_forecast, min_forecast, max_week, min_week, output_file_path):
    # Convert ds to Week labels here
    comparison_for_excel = comparison.copy()
    if 'yhat' in comparison_for_excel.columns:
        comparison_for_excel.drop(columns=['yhat'], inplace=True)
    if 'yhat_upper' in comparison_for_excel.columns:
        comparison_for_excel.drop(columns=['yhat_upper'], inplace=True)

    # Replace ds with Week i
    for i in range(len(comparison_for_excel)):
        comparison_for_excel.loc[i, 'Week'] = f"Week {i}"
    if 'ds' in comparison_for_excel.columns:
        comparison_for_excel.drop(columns=['ds'], inplace=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Forecast Comparison"

    for r in dataframe_to_rows(comparison_for_excel, index=False, header=True):
        ws1.append(r)

    ws2 = wb.create_sheet(title="Summary")
    summary_data = {
        "Metric": [
            "Historical Range",
            "Min Sales",
            "Max Sales",
            "Mean Sales",
            "Median Sales",
            "Std Dev Sales",
            "Total Historical Sales",
            "Total Forecast (16 Weeks)",
            "Total Forecast (8 Weeks)",
            "Total Forecast (4 Weeks)",
            "Max Forecast",
            "Max Forecast Week",
            "Min Forecast",
            "Min Forecast Week"
        ],
        "Value": [
            f"{summary_stats['data_range'][0].strftime('%Y-%m-%d')} to {summary_stats['data_range'][1].strftime('%Y-%m-%d')}",
            f"{summary_stats['min']:.0f}",
            f"{summary_stats['max']:.0f}",
            f"{summary_stats['mean']:.0f}",
            f"{summary_stats['median']:.0f}",
            f"{summary_stats['std_dev']:.0f}",
            f"{summary_stats['total_sales']:.0f} units",
            f"{total_forecast_16:.0f}",
            f"{total_forecast_8:.0f}",
            f"{total_forecast_4:.0f}",
            f"{max_forecast:.0f}",
            f"{max_week.strftime('%Y-%m-%d')}",
            f"{min_forecast:.0f}",
            f"{min_week.strftime('%Y-%m-%d')}"
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws2.append(r)

    wb.save(output_file_path)
    print(f"Comparison and summary saved to '{output_file_path}'")

def save_forecast_to_excel(output_path, consolidated_data, missing_asin_data):
    wb = Workbook()
    for asin, forecast_df in consolidated_data.items():
        df_for_excel = forecast_df.copy()
        if 'yhat' in df_for_excel.columns:
            df_for_excel.drop(columns=['yhat'], inplace=True)
        if 'yhat_upper' in df_for_excel.columns:
            df_for_excel.drop(columns=['yhat_upper'], inplace=True)

        if 'ds' in df_for_excel.columns:
            for i in range(len(df_for_excel)):
                df_for_excel.loc[i, 'Week'] = f"Week {i}"
            df_for_excel.drop(columns=['ds'], inplace=True)

        ws = wb.create_sheet(title=str(asin)[:31])
        for r in dataframe_to_rows(df_for_excel, index=False, header=True):
            ws.append(r)

    if not missing_asin_data.empty:
        ws_missing = wb.create_sheet(title="No ASIN")
        for r in dataframe_to_rows(missing_asin_data, index=False, header=True):
            ws_missing.append(r)

    del wb['Sheet']
    wb.save(output_path)
    print(f"All forecasts saved to {output_path}")

def cross_validate_prophet_model(ts_data, initial='365 days', period='180 days', horizon='180 days'):
    model = Prophet(
        yearly_seasonality=True,
        weekly_seasonality=True,
        seasonality_mode='multiplicative'
    )
    model.fit(ts_data[['ds', 'y']])
    df_cv = cross_validation(model, initial=initial, period=period, horizon=horizon)
    df_performance = performance_metrics(df_cv)
    print("Prophet Model Cross-Validation Results:")
    print(df_performance)
    return df_cv, df_performance

def analyze_amazon_buying_habits(comparison, holidays):
    amazon_cols = [col for col in comparison.columns if col.startswith('Amazon ')]
    if not amazon_cols:
        print("No Amazon forecasts found for analysis.")
        return

    prophet_forecast = comparison['Prophet Forecast'].values
    ds_dates = comparison.get('ds', pd.Series(index=comparison.index))
    holiday_dates = holidays['ds'].values if holidays is not None else []
    comparison['is_holiday_week'] = comparison.get('ds', pd.Series(index=comparison.index)).isin(holiday_dates) if 'ds' in comparison.columns else False

    for col in amazon_cols:
        amazon_forecast = comparison[col].values
        safe_prophet = np.where(prophet_forecast == 0, 1e-9, prophet_forecast)

        ratio = amazon_forecast / safe_prophet
        diff = amazon_forecast - prophet_forecast

        avg_ratio = np.mean(ratio)
        avg_diff = np.mean(diff)
        print(f"\nFor {col}:")
        print(f"  Average Amazon/Prophet Ratio: {avg_ratio:.2f}")
        print(f"  Average Difference (Amazon - Prophet): {avg_diff:.2f}")

        if avg_diff > 0:
            print("  Amazon tends to forecast more than Prophet on average.")
        elif avg_diff < 0:
            print("  Amazon tends to forecast less than Prophet on average.")
        else:
            print("  Amazon forecasts similarly to Prophet on average.")

        holiday_mask = comparison['is_holiday_week']
        if holiday_mask.any():
            holiday_ratio = (amazon_forecast[holiday_mask] / safe_prophet[holiday_mask])
            holiday_diff = amazon_forecast[holiday_mask] - prophet_forecast[holiday_mask]
            if len(holiday_diff) > 0:
                print("  During holiday weeks:")
                print(f"    Avg Ratio (Amazon/Prophet): {np.mean(holiday_ratio):.2f}")
                print(f"    Avg Diff (Amazon-Prophet): {np.mean(holiday_diff):.2f}")

        weeks = np.arange(1, len(ratio)+1)
        segments = {
            'Short-term (Weeks 1-4)': (weeks <= 4),
            'Mid-term (Weeks 5-12)': (weeks >=5) & (weeks <=12),
            'Long-term (Weeks 13+)': (weeks > 12)
        }

        for segment_name, mask in segments.items():
            if mask.any():
                seg_ratio = ratio[mask]
                seg_diff = diff[mask]
                print(f"  {segment_name}:")
                print(f"    Avg Ratio (Amazon/Prophet): {np.mean(seg_ratio):.2f}")
                print(f"    Avg Diff (Amazon-Prophet): {np.mean(seg_diff):.2f}")

def main():
    sales_file = 'weekly_sales_data.xlsx'
    forecasts_folder = 'forecasts_folder'
    asins_to_forecast_file = 'ASINs to Forecast.xlsx'
    output_file = 'consolidated_forecast.xlsx'
    horizon = 20

    data = load_weekly_sales_data(sales_file)
    valid_data = data[data['asin'].notna() & (data['asin'] != '#N/A')]
    missing_asin_data = data[data['asin'].isna() | (data['asin'] == '#N/A')]
    if not missing_asin_data.empty:
        print("The following entries have no ASIN and will be noted in the forecast file:")
        # Printing with 'errors=ignore' in print to handle unicode
        # Or just encode and decode to avoid errors:
        print(missing_asin_data[['product title', 'week', 'year', 'y']].to_string())

    asins_to_forecast = load_asins_to_forecast(asins_to_forecast_file)
    print(f"ASINs to forecast: {asins_to_forecast}")

    asin_list = valid_data['asin'].unique()
    asin_list = [asin for asin in asin_list if asin in asins_to_forecast]

    consolidated_forecasts = {}
    param_grid = {
        'changepoint_prior_scale': [0.05, 0.1, 0.2, 0.3, 0.4, 0.5],
        'seasonality_prior_scale': [1, 2, 3, 4, 5, 10],
        'holidays_prior_scale': [5, 10, 20]
    }

    holidays = get_shifted_holidays()

    # Optional: Cross-validate on one ASIN
    if len(asin_list) > 0:
        test_asin = asin_list[0]
        test_ts_data = prepare_time_series_with_lags(valid_data, test_asin, lag_weeks=1)
        if not test_ts_data.empty:
            print(f"Performing cross-validation on ASIN {test_asin} Prophet model (no Amazon regressors)...")
            cross_validate_prophet_model(test_ts_data, initial='365 days', period='180 days', horizon='180 days')

    for asin in asin_list:
        product_title = valid_data[valid_data['asin'] == asin]['product title'].iloc[0]
        print(f"\nProcessing ASIN: {asin} - {product_title}")

        forecast_data = load_amazon_forecasts_from_folder(forecasts_folder, asin)
        if not forecast_data:
            print(f"No forecast data found for ASIN {asin}, skipping.")
            continue

        ts_data = prepare_time_series_with_lags(valid_data, asin, lag_weeks=1)
        if ts_data.empty:
            print(f"No historical data found for ASIN {asin}, skipping.")
            continue

        best_params, best_rmse_values = optimize_prophet_params(ts_data, forecast_data, param_grid, horizon=horizon)
        print(f"Best parameters for ASIN {asin}: {best_params}")
        print(f"Best RMSE values for ASIN {asin}: {best_rmse_values}")

        validate_best_params(ts_data, best_params)

        forecast, _ = forecast_with_custom_params(
            ts_data, forecast_data,
            best_params['changepoint_prior_scale'],
            best_params['seasonality_prior_scale'],
            best_params['holidays_prior_scale'],
            horizon=horizon
        )

        if forecast.empty:
            print(f"Forecasting failed for ASIN {asin}, skipping.")
            continue

        comparison = format_output_with_forecasts(forecast, forecast_data, horizon=horizon)
        best_weights, best_rmse = auto_find_best_weights(forecast, comparison, step=0.05)
        print(f"Auto best weights for ASIN {asin}: {best_weights} with RMSE={best_rmse}")

        forecast = adjust_forecast_weights(forecast, *best_weights)
        comparison = format_output_with_forecasts(forecast, forecast_data, horizon=horizon)
        comparison['Prophet Forecast'] = forecast['Prophet Forecast']
        comparison['ASIN'] = asin
        comparison['Product Title'] = product_title

        analyze_amazon_buying_habits(comparison, holidays)

        summary_stats, total_forecast_16, total_forecast_8, total_forecast_4, max_forecast, min_forecast, max_week, min_week = calculate_summary_statistics(ts_data, comparison, horizon=horizon)
        visualize_forecast_with_comparison(ts_data, comparison, summary_stats, total_forecast_16, total_forecast_8, total_forecast_4, max_forecast, min_forecast, max_week, min_week, asin, product_title)

        output_file_name = f'forecast_summary_{asin}.xlsx'
        output_file_path = os.path.abspath(output_file_name)
        save_summary_to_excel(
            comparison,
            summary_stats,
            total_forecast_16,
            total_forecast_8,
            total_forecast_4,
            max_forecast,
            min_forecast,
            max_week,
            min_week,
            output_file_path
        )

        consolidated_forecasts[asin] = comparison

    save_forecast_to_excel(output_file, consolidated_forecasts, missing_asin_data)

if __name__ == '__main__':
    main()
